"""
HDR (HelpDesk Report) Views
Monthly reporting system for helpdesk incidents and service requests
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db import models
from datetime import date
from .models import HDRReport, HDREntry


@login_required
def hdr_list(request):
    """List all HDR reports"""
    reports = HDRReport.objects.all().select_related('created_by')

    context = {
        'reports': reports,
    }
    return render(request, 'hdr/hdr_list.html', context)


@login_required
def hdr_create(request):
    """Create a new HDR report"""
    if request.method == 'POST':
        try:
            # Get form data
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))

            # Check if report already exists
            if HDRReport.objects.filter(month=month, year=year).exists():
                messages.error(request, f'A report for {date(year, month, 1).strftime("%B %Y")} already exists.')
                return redirect('hdr_list')

            # Create report (no default entries - user will add manually)
            report = HDRReport.objects.create(
                month=month,
                year=year,
                region=request.POST.get('region', 'VIII'),
                office=request.POST.get('office', 'DPWH Leyte 4th Engineering District'),
                address=request.POST.get('address', 'Ormoc City Leyte'),
                network_admin_name=request.POST.get('network_admin_name', 'BOBBY L. YU'),
                network_admin_contact=request.POST.get('network_admin_contact', '66800|09219290909'),
                network_admin_email=request.POST.get('network_admin_email', 'yu.bobby@dpwh.gov.ph'),
                created_by=request.user
            )

            messages.success(request, f'HDR report for {report.period_display} created successfully. You can now add incidents.')
            return redirect('hdr_edit', report_id=report.id)

        except Exception as e:
            messages.error(request, f'Error creating report: {str(e)}')
            return redirect('hdr_list')

    # GET request - show create form
    context = {
        'months': [
            (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
            (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
            (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
        ],
        'years': range(2024, 2031),
        'current_month': timezone.now().month,
        'current_year': timezone.now().year,
    }
    return render(request, 'hdr/hdr_create.html', context)


@login_required
def hdr_view(request, report_id):
    """View HDR report details"""
    report = get_object_or_404(HDRReport, id=report_id)
    entries = report.entries.order_by('date_reported', 'ref_number')

    context = {
        'report': report,
        'entries': entries
    }
    return render(request, 'hdr/hdr_view.html', context)


@login_required
def hdr_edit(request, report_id):
    """Edit an existing HDR report"""
    report = get_object_or_404(HDRReport, id=report_id)

    # Prevent editing finalized reports
    if report.is_finalized:
        messages.warning(request, 'This report has been finalized and cannot be edited.')
        return redirect('hdr_view', report_id=report.id)

    if request.method == 'POST':
        try:
            # Check if adding new entry
            if 'add_entry' in request.POST:
                # Auto-generate reference number
                ref_number = report.get_next_ref_number()

                HDREntry.objects.create(
                    report=report,
                    ref_number=ref_number,
                    incident_type=request.POST.get('new_incident_type', 'Service Request'),
                    main_category=request.POST.get('new_main_category'),
                    sub_category=request.POST.get('new_sub_category'),
                    description=request.POST.get('new_description'),
                    status=request.POST.get('new_status', 'Pending'),
                    date_reported=request.POST.get('new_date_reported'),
                    reported_by=request.POST.get('new_reported_by'),
                    resolution=request.POST.get('new_resolution', '')
                )
                messages.success(request, f'New incident added successfully with reference number: {ref_number}')
                return redirect('hdr_edit', report_id=report.id)

            # Check if deleting entry
            if 'delete_entry' in request.POST:
                entry_id = int(request.POST.get('entry_id'))
                entry = get_object_or_404(HDREntry, id=entry_id, report=report)
                entry.delete()
                messages.success(request, 'Incident deleted successfully.')
                return redirect('hdr_edit', report_id=report.id)

            # Update report details
            report.region = request.POST.get('region')
            report.office = request.POST.get('office')
            report.address = request.POST.get('address')
            report.network_admin_name = request.POST.get('network_admin_name')
            report.network_admin_contact = request.POST.get('network_admin_contact')
            report.network_admin_email = request.POST.get('network_admin_email')
            report.save()

            # Update entries
            for entry in report.entries.all():
                entry_prefix = f'entry_{entry.id}_'
                entry.ref_number = request.POST.get(f'{entry_prefix}ref_number', entry.ref_number)
                entry.incident_type = request.POST.get(f'{entry_prefix}incident_type', entry.incident_type)
                entry.main_category = request.POST.get(f'{entry_prefix}main_category', entry.main_category)
                entry.sub_category = request.POST.get(f'{entry_prefix}sub_category', entry.sub_category)
                entry.description = request.POST.get(f'{entry_prefix}description', entry.description)
                entry.status = request.POST.get(f'{entry_prefix}status', entry.status)
                entry.date_reported = request.POST.get(f'{entry_prefix}date_reported', entry.date_reported)
                entry.reported_by = request.POST.get(f'{entry_prefix}reported_by', entry.reported_by)
                entry.resolution = request.POST.get(f'{entry_prefix}resolution', entry.resolution)
                entry.save()

            messages.success(request, 'Report updated successfully.')
            return redirect('hdr_view', report_id=report.id)

        except Exception as e:
            messages.error(request, f'Error updating report: {str(e)}')

    context = {
        'report': report,
        'entries': report.entries.order_by('date_reported', 'ref_number'),
        'type_choices': HDREntry.TYPE_CHOICES,
        'category_choices': HDREntry.CATEGORY_CHOICES,
        'status_choices': HDREntry.STATUS_CHOICES,
        'next_ref_number': report.get_next_ref_number(),
    }
    return render(request, 'hdr/hdr_edit.html', context)


@login_required
def hdr_jobsheet_form(request, report_id):
    """Job sheet form for adding new incidents"""
    report = get_object_or_404(HDRReport, id=report_id)

    # Prevent editing finalized reports
    if report.is_finalized:
        messages.warning(request, 'This report has been finalized and cannot be edited.')
        return redirect('hdr_view', report_id=report.id)

    if request.method == 'POST':
        try:
            # Auto-generate reference number
            ref_number = report.get_next_ref_number()

            HDREntry.objects.create(
                report=report,
                ref_number=ref_number,
                incident_type=request.POST.get('incident_type'),
                main_category=request.POST.get('main_category'),
                sub_category=request.POST.get('sub_category'),
                description=request.POST.get('description'),
                status=request.POST.get('status', 'fixed'),
                date_reported=request.POST.get('date_reported'),
                reported_by=request.POST.get('reported_by'),
                section_division=request.POST.get('section_division', ''),
                contact_no=request.POST.get('contact_no', ''),
                resolution=request.POST.get('resolution', ''),
                # Hardware details
                hardware_type=request.POST.get('hardware_type', ''),
                hardware_brand_model=request.POST.get('hardware_brand_model', ''),
                hardware_serial_number=request.POST.get('hardware_serial_number', ''),
                computer_name=request.POST.get('computer_name', ''),
                # Application System / Software
                application_description=request.POST.get('application_description', ''),
                application_version=request.POST.get('application_version', ''),
                # Connectivity
                connectivity_description=request.POST.get('connectivity_description', ''),
                # User Account
                user_account_description=request.POST.get('user_account_description', ''),
                # Assessment
                assessment=request.POST.get('assessment', ''),
                # Mode of Filing and Personnel
                mode_of_filing=request.POST.get('mode_of_filing', ''),
                fulfilled_by=request.POST.get('fulfilled_by', ''),
                reviewed_by=request.POST.get('reviewed_by', ''),
                # Client Evaluation
                concern_addressed=request.POST.get('concern_addressed', ''),
                satisfaction_service=request.POST.get('satisfaction_service', ''),
                satisfaction_solution=request.POST.get('satisfaction_solution', ''),
                client_comments=request.POST.get('client_comments', '')
            )
            messages.success(request, f'Job sheet submitted successfully! Reference Number: {ref_number}')
            return redirect('hdr_edit', report_id=report.id)

        except Exception as e:
            messages.error(request, f'Error submitting job sheet: {str(e)}')

    context = {
        'report': report,
        'type_choices': HDREntry.TYPE_CHOICES,
        'category_choices': HDREntry.CATEGORY_CHOICES,
        'status_choices': HDREntry.STATUS_CHOICES,
        'next_ref_number': report.get_next_ref_number(),
        'today': date.today().isoformat(),
    }
    return render(request, 'hdr/jobsheet_form.html', context)


@login_required
def hdr_entry_detail(request, entry_id):
    """View detailed information for a single HDR entry"""
    entry = get_object_or_404(HDREntry, id=entry_id)

    context = {
        'entry': entry,
        'report': entry.report,
    }
    return render(request, 'hdr/entry_detail.html', context)


@login_required
def hdr_delete(request, report_id):
    """Delete HDR report"""
    if request.method == 'POST':
        report = get_object_or_404(HDRReport, id=report_id)
        period = report.period_display
        report.delete()
        messages.success(request, f'HDR report for {period} deleted successfully.')
        return redirect('hdr_list')
    return redirect('hdr_list')


@login_required
def hdr_finalize(request, report_id):
    """Finalize HDR report (lock it from editing)"""
    if request.method == 'POST':
        report = get_object_or_404(HDRReport, id=report_id)
        report.is_finalized = True
        report.finalized_at = timezone.now()
        report.save()
        messages.success(request, f'HDR report for {report.period_display} has been finalized.')
        return redirect('hdr_view', report_id=report.id)
    return redirect('hdr_list')


@login_required
def hdr_entry_export_diagnostic(request, entry_id):
    """Diagnostic export to show cell structure - helps identify correct cell positions"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    import io

    entry = get_object_or_404(HDREntry, id=entry_id)

    # Load the Standard Job Sheet template
    template_path = 'templates/excel temps/Standard Job Sheet.xlsx'
    wb = load_workbook(template_path, data_only=False, keep_vba=False)
    ws = wb.active

    # Highlight cells with their addresses to help map data
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Add cell addresses as values to identify positions
    diagnostic_cells = {
        'B6': 'B6-RefNo?', 'G6': 'G6-RefNo?', 'H6': 'H6-RefNo?',
        'B8': 'B8-Date?', 'G8': 'G8-Date?', 'H8': 'H8-Date?',
        'B11': 'B11-Name?', 'B12': 'B12-Section?', 'G12': 'G12-Contact?',
        'B14': 'B14-Description?',
        'B17': 'B17-Type?', 'C17': 'C17-Category?',
        'B18': 'B18-SubCat?', 'C18': 'C18-Status?',
        'B21': 'B21-HWType?', 'C21': 'C21-HWBrand?',
        'B22': 'B22-Serial?', 'E22': 'E22-CompName?',
    }

    for cell, label in diagnostic_cells.items():
        ws[cell].value = label
        ws[cell].fill = yellow_fill
        ws[cell].font = Font(bold=True, color='FF0000')

    # Prepare response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="JobSheet_Diagnostic.xlsx"'
    return response


@login_required
def hdr_entry_export(request, entry_id):
    """Export individual job sheet to Excel using Standard Job Sheet template"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    import io

    entry = get_object_or_404(HDREntry, id=entry_id)
    report = entry.report

    # Load the Standard Job Sheet template
    template_path = 'templates/excel temps/Standard Job Sheet.xlsx'
    wb = load_workbook(template_path, data_only=False, keep_vba=False)
    ws = wb.active

    # Fill in the job sheet data based on the actual template structure
    # Looking at the exported form, mapping cells more precisely

    # Header Section (top right)
    ws['G6'] = str(entry.ref_number)  # Ref No field (top right)

    # CLIENT'S INFORMATION Section
    ws['B11'] = str(entry.reported_by)  # Full Name (first field under client info header)
    ws['B12'] = str(entry.section_division)  # Section/Division
    ws['G11'] = entry.date_reported.strftime('%B %d, %Y') if entry.date_reported else ''  # Date of Filing (right side)
    ws['G12'] = str(entry.contact_no)  # Contact No. (right side, same row as section)
    ws['B14'] = str(entry.description)  # Brief description (multi-line field)

    # I.T. SUPPORT TECHNICAL ASSESSMENT Section
    # Incident Classification subsection
    ws['B18'] = str(entry.incident_type)  # Type of Incident
    ws['D18'] = str(entry.main_category)  # Main Category
    ws['B19'] = str(entry.sub_category)  # Sub-Category
    ws['D19'] = str(entry.status)  # Status

    # Hardware subsection
    ws['B22'] = str(entry.hardware_type)  # Type
    ws['D22'] = str(entry.hardware_brand_model)  # Brand and Model
    ws['B23'] = str(entry.hardware_serial_number)  # Serial Number
    ws['E23'] = str(entry.computer_name)  # Computer Name

    # Application System / Software subsection
    ws['B26'] = str(entry.application_description)  # Description
    ws['G26'] = str(entry.application_version)  # Version

    # Connectivity subsection
    ws['B29'] = str(entry.connectivity_description)

    # User Account subsection
    ws['B32'] = str(entry.user_account_description)

    # Assessment subsection
    ws['B35'] = str(entry.assessment)

    # Actions Taken and/or Recommendations
    ws['B38'] = str(entry.resolution)

    # Mode of Filing and Personnel Section
    # Mode checkboxes around row 42, personnel names to the right
    ws['E43'] = str(entry.fulfilled_by)  # Fulfilled by
    ws['E44'] = str(entry.reviewed_by)  # Reviewed by

    # CLIENT'S EVALUATION Section (bottom)
    # Questions with Yes/No and satisfaction options
    ws['B48'] = str(entry.concern_addressed)  # Question 1
    ws['B49'] = str(entry.satisfaction_service)  # Question 2
    ws['B50'] = str(entry.satisfaction_solution)  # Question 3
    ws['B52'] = str(entry.client_comments)  # Comments/Suggestions

    # Prepare response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'JobSheet_{entry.ref_number}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def hdr_export_excel(request, report_id):
    """Export HDR report to Excel using template - fills data into pre-formatted cells"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
    from copy import copy
    from datetime import datetime
    import io

    report = get_object_or_404(HDRReport, id=report_id)
    entries = report.entries.order_by('date_reported', 'ref_number')

    # Load the new template
    template_path = 'media/pm_reports/new_hdr.xlsx'
    wb = load_workbook(template_path, data_only=False, keep_vba=False)

    # Use the template's first sheet
    ws = wb.worksheets[0]
    wb.active = ws  # Make it the active sheet

    # Unmerge all cells to avoid "MergedCell is read-only" errors
    # Create a list of merged ranges to unmerge
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    # Fill in header information in the template
    ws['B2'] = report.period_display
    ws['B4'] = report.region
    ws['H4'] = report.network_admin_name
    ws['B5'] = report.office
    ws['H5'] = report.network_admin_contact
    ws['B6'] = report.address
    ws['H6'] = report.network_admin_email
    # Data starts at row 9 (assuming row 8 has headers in template)
    data_start_row = 9

    # Write each entry to the template
    for idx, entry in enumerate(entries):
        row_num = data_start_row + idx

        # Write data directly to template cells
        ws[f'A{row_num}'] = str(entry.ref_number)
        ws[f'B{row_num}'] = str(entry.incident_type)
        ws[f'C{row_num}'] = str(entry.main_category)
        ws[f'D{row_num}'] = str(entry.sub_category)
        ws[f'E{row_num}'] = str(entry.description)
        ws[f'F{row_num}'] = str(entry.status)
        ws[f'G{row_num}'] = entry.date_reported if entry.date_reported else ''
        ws[f'H{row_num}'] = str(entry.reported_by)
        ws[f'I{row_num}'] = str(entry.resolution) if entry.resolution else ''

        # Enable text wrapping for all data cells in this row
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws[f'{col}{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')

    # Apply formatting
    # Merge specific cells
    # Apply formatting
    # Merge specific cells
    ws.merge_cells('G14:H14')
    ws.merge_cells('G15:H15')
    ws.merge_cells('G16:H16')
    ws.merge_cells('G17:H17')

    ws.merge_cells('A12:B12')
    ws.merge_cells('A14:B14')
    ws.merge_cells('A15:B15')
    ws.merge_cells('A16:B16')
    ws.merge_cells('A17:B17')

    # Prepare response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'HDR_{report.period_display.replace(" ", "_")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
