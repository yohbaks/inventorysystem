"""
ASIR (Application Systems Implementation Report) Views
Monthly reporting system for application systems deployed in the office
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db import models
from datetime import date
from .models import ASIRReport, ASIREntry


@login_required
def asir_list(request):
    """List all ASIR reports"""
    reports = ASIRReport.objects.all().select_related('created_by')

    context = {
        'reports': reports,
    }
    return render(request, 'asir/asir_list.html', context)


@login_required
def asir_create(request):
    """Create a new ASIR report"""
    if request.method == 'POST':
        try:
            # Get form data
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))

            # Check if report already exists
            if ASIRReport.objects.filter(month=month, year=year).exists():
                messages.error(request, f'A report for {date(year, month, 1).strftime("%B %Y")} already exists.')
                return redirect('asir_list')

            # Create report
            report = ASIRReport.objects.create(
                month=month,
                year=year,
                region=request.POST.get('region', 'VIII'),
                office=request.POST.get('office', 'DPWH, Leyte Fourth District Engineering Office'),
                address=request.POST.get('address', 'Ormoc City, Leyte'),
                network_admin_name=request.POST.get('network_admin_name', 'BOBBY L. YU'),
                network_admin_contact=request.POST.get('network_admin_contact', '09219290909'),
                network_admin_email=request.POST.get('network_admin_email', 'yu.bobby@dpwh.gov.ph'),
                created_by=request.user
            )

            # Create default entries based on common application systems
            default_applications = [
                {'name': 'CWR - Civil Works Registry', 'users': 2},
                {'name': 'eBudget - Electronic Budget System', 'users': 12},
                {'name': 'eNGAS - Electronic New Government Accounting System', 'users': 12},
                {'name': 'RBIA-Util - RBIA Utilities', 'users': 3},
                {'name': 'PCMA - Project Construction Management Application', 'users': 15},
                {'name': 'Autodesk Autocad', 'users': 3},
                {'name': 'Civil 3D', 'users': 11},
                {'name': 'CCIS - Construction Cost Information System', 'users': 2},
                {'name': 'RTIA - Road Traffic Information Application', 'users': 1},
                {'name': 'Sophos Client Agent', 'users': 110},
                {'name': 'RPS - Regular Payroll System', 'users': 3},
                {'name': 'BMC Client', 'users': 10},
                {'name': 'MYPS - Multi Year Programming and Scheduling Application', 'users': 1},
                {'name': 'ArcGIS', 'users': 2},
                {'name': 'DoTS - Document Tracking System', 'users': 9},
            ]

            for idx, app in enumerate(default_applications, start=1):
                ASIREntry.objects.create(
                    report=report,
                    item_number=idx,
                    application_name=app['name'],
                    number_of_users=app['users'],
                    status='Deployed and being used',
                    activity_details='None',
                    activity_date='N/A',
                    remarks='None'
                )

            messages.success(request, f'ASIR report for {report.period_display} created successfully.')
            return redirect('asir_edit', report_id=report.id)

        except Exception as e:
            messages.error(request, f'Error creating report: {str(e)}')
            return redirect('asir_list')

    # GET request - show create form
    context = {
        'months': [
            (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
            (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
            (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
        ],
        'years': range(2024, 2031),
        'current_month': timezone.now().month,
        'current_year': timezone.now().year,
    }
    return render(request, 'asir/asir_create.html', context)


@login_required
def asir_view(request, report_id):
    """View ASIR report details"""
    report = get_object_or_404(ASIRReport, id=report_id)
    entries = report.entries.order_by('item_number')

    context = {
        'report': report,
        'entries': entries
    }
    return render(request, 'asir/asir_view.html', context)


@login_required
def asir_edit(request, report_id):
    """Edit an existing ASIR report"""
    report = get_object_or_404(ASIRReport, id=report_id)

    # Prevent editing finalized reports
    if report.is_finalized:
        messages.warning(request, 'This report has been finalized and cannot be edited.')
        return redirect('asir_view', report_id=report.id)

    if request.method == 'POST':
        try:
            # Check if adding new entry
            if 'add_entry' in request.POST:
                # Get the highest item number
                max_item = report.entries.aggregate(models.Max('item_number'))['item_number__max'] or 0
                ASIREntry.objects.create(
                    report=report,
                    item_number=max_item + 1,
                    application_name=request.POST.get('new_application_name', 'New Application'),
                    number_of_users=int(request.POST.get('new_number_of_users', 0)),
                    status=request.POST.get('new_status', 'Deployed and being used'),
                    activity_details=request.POST.get('new_activity_details', 'None'),
                    activity_date=request.POST.get('new_activity_date', 'N/A'),
                    remarks=request.POST.get('new_remarks', 'None')
                )
                messages.success(request, 'New entry added successfully.')
                return redirect('asir_edit', report_id=report.id)

            # Check if deleting entry
            if 'delete_entry' in request.POST:
                entry_id = int(request.POST.get('entry_id'))
                entry = get_object_or_404(ASIREntry, id=entry_id, report=report)
                entry.delete()
                # Renumber remaining entries
                for idx, entry in enumerate(report.entries.order_by('item_number'), start=1):
                    entry.item_number = idx
                    entry.save()
                messages.success(request, 'Entry deleted successfully.')
                return redirect('asir_edit', report_id=report.id)

            # Update report details
            report.region = request.POST.get('region')
            report.office = request.POST.get('office')
            report.address = request.POST.get('address')
            report.network_admin_name = request.POST.get('network_admin_name')
            report.network_admin_contact = request.POST.get('network_admin_contact')
            report.network_admin_email = request.POST.get('network_admin_email')
            report.save()

            # Update entries
            for entry in report.entries.all():
                entry_prefix = f'entry_{entry.id}_'
                entry.application_name = request.POST.get(f'{entry_prefix}application_name', entry.application_name)
                entry.number_of_users = int(request.POST.get(f'{entry_prefix}number_of_users', entry.number_of_users))
                entry.status = request.POST.get(f'{entry_prefix}status', entry.status)
                entry.activity_details = request.POST.get(f'{entry_prefix}activity_details', entry.activity_details)
                entry.activity_date = request.POST.get(f'{entry_prefix}activity_date', entry.activity_date)
                entry.remarks = request.POST.get(f'{entry_prefix}remarks', entry.remarks)
                entry.save()

            messages.success(request, 'Report updated successfully.')
            return redirect('asir_view', report_id=report.id)

        except Exception as e:
            messages.error(request, f'Error updating report: {str(e)}')

    context = {
        'report': report,
        'entries': report.entries.order_by('item_number'),
        'status_choices': ASIREntry.STATUS_CHOICES,
    }
    return render(request, 'asir/asir_edit.html', context)


@login_required
def asir_delete(request, report_id):
    """Delete ASIR report"""
    if request.method == 'POST':
        report = get_object_or_404(ASIRReport, id=report_id)
        period = report.period_display
        report.delete()
        messages.success(request, f'ASIR report for {period} deleted successfully.')
        return redirect('asir_list')
    return redirect('asir_list')


@login_required
def asir_finalize(request, report_id):
    """Finalize ASIR report (lock it from editing)"""
    if request.method == 'POST':
        report = get_object_or_404(ASIRReport, id=report_id)
        report.is_finalized = True
        report.finalized_at = timezone.now()
        report.save()
        messages.success(request, f'ASIR report for {report.period_display} has been finalized.')
        return redirect('asir_view', report_id=report.id)
    return redirect('asir_list')


@login_required
def asir_export_excel(request, report_id):
    """Export ASIR report to Excel using template - fills data into pre-formatted cells"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from django.db import models
    import io

    report = get_object_or_404(ASIRReport, id=report_id)
    entries = report.entries.order_by('item_number')

    # Load template
    template_path = 'media/pm_reports/Ley4_ASIR.xlsx'
    wb = load_workbook(template_path)

    # Use the "Do not delete" sheet as template or create new sheet
    # We'll use "October 2025" sheet as template and modify it
    if 'Do not delete' in wb.sheetnames:
        ws = wb['Do not delete']
    else:
        ws = wb.active

    # Fill in header information
    ws['A2'] = f'For the Month of {report.period_display}'
    ws['B4'] = report.region
    ws['B5'] = report.office
    ws['B6'] = report.address
    ws['F4'] = report.network_admin_name
    ws['F5'] = report.network_admin_contact
    ws['F6'] = report.network_admin_email

    # Data starts at row 10 (template already has headers and formatting)
    data_start_row = 10

    # Fill in data entries - preserve template formatting
    for idx, entry in enumerate(entries):
        row_num = data_start_row + idx

        # Fill in the data for this entry
        ws[f'A{row_num}'] = entry.item_number
        ws[f'B{row_num}'] = entry.application_name
        ws[f'C{row_num}'] = entry.number_of_users if entry.number_of_users > 0 else ''
        ws[f'D{row_num}'] = entry.status
        ws[f'E{row_num}'] = entry.activity_details
        ws[f'F{row_num}'] = entry.activity_date
        ws[f'G{row_num}'] = entry.remarks

        # Enable text wrapping for all data cells to handle long content
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            cell = ws[f'{col}{row_num}']
            if cell.alignment:
                # Preserve existing alignment but enable wrapping
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=True
                )
            else:
                # Default alignment with wrapping
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )

    # Prepare response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ASIR_{report.period_display.replace(" ", "_")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
