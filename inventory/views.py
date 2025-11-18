# Standard library
import io
import os
import json
import qrcode
import traceback
from io import BytesIO
from datetime import datetime, timedelta
from calendar import month_abbr
from collections import defaultdict
try:
    import pythoncom
    from win32com.client import Dispatch
    HAS_WIN32 = True
except ImportError:
    # Windows-specific modules not available on Linux
    HAS_WIN32 = False
from django.http import HttpResponse, Http404


# Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.db import IntegrityError, transaction
from django.db.models import Count, F, Prefetch, Max, Q
from django.db.models.functions import TruncDay, TruncMonth, Upper, Trim
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from django.utils.dateformat import DateFormat
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

# Third-party (keep if used anywhere in file)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
try:
    from weasyprint import HTML
    HAS_WEASYPRINT = True
except ImportError:
    HAS_WEASYPRINT = False
try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False
try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
try:
    from docx2pdf import convert
    HAS_DOCX2PDF = True
except ImportError:
    HAS_DOCX2PDF = False

# Local
from inventory.models import (
    Equipment_Package, DesktopDetails, KeyboardDetails, MouseDetails, MonitorDetails,
    UPSDetails, UserDetails, DocumentsDetails, Employee, Brand,
    DisposedDesktopDetail, DisposedKeyboard, DisposedMouse, DisposedMonitor, DisposedUPS,
    SalvagedMonitor, SalvagedKeyboard, SalvagedMouse, SalvagedUPS,
    SalvagedMonitorHistory, SalvagedKeyboardHistory, SalvagedMouseHistory, SalvagedUPSHistory,
    EndUserChangeHistory, AssetOwnerChangeHistory,
    PreventiveMaintenance, PMScheduleAssignment, MaintenanceChecklistItem,
    QuarterSchedule, PMSectionSchedule, OfficeSection, Profile,
    LaptopPackage, LaptopDetails, DisposedLaptop, PrinterPackage, PrinterDetails, DisposedPrinter, Notification,
    OfficeSuppliesPackage, OfficeSuppliesDetails, DisposedOfficeSupplies
)


from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from django.db.models import F, Value
from django.db.models.functions import Upper, Trim

from inventory.utils.pm_helpers import transfer_pm_schedule_on_user_change

from django.contrib.contenttypes.models import ContentType


# ==========================================
# HELPER FUNCTIONS FOR CHANGE HISTORY
# ==========================================

def get_asset_owner_history(package):
    """Get asset owner change history for any package type (Desktop, Laptop, Printer, etc.)"""
    from django.contrib.contenttypes.models import ContentType
    content_type = ContentType.objects.get_for_model(package)
    return AssetOwnerChangeHistory.objects.filter(
        content_type=content_type,
        object_id=package.id
    ).select_related('old_assetowner', 'new_assetowner', 'changed_by').order_by('-changed_at')


def get_end_user_history(package):
    """Get end user change history for any package type (Desktop, Laptop, Printer, etc.)"""
    from django.contrib.contenttypes.models import ContentType
    content_type = ContentType.objects.get_for_model(package)
    return EndUserChangeHistory.objects.filter(
        content_type=content_type,
        object_id=package.id
    ).select_related('old_enduser', 'new_enduser', 'changed_by').order_by('-changed_at')


def auto_transfer_pm_schedule(equipment_package, new_enduser):
    """
    Auto-transfer PM schedule to the new End User's section.
    - Keeps completed PMs in their ORIGINAL section (history preservation)
    - Transfers only PENDING schedules to the new section
    - Creates new schedules for quarters not yet assigned
    """
    from inventory.models import PMScheduleAssignment, PMSectionSchedule, PreventiveMaintenance
    from django.db import transaction

    try:
        if not new_enduser or not new_enduser.employee_office_section:
            return "⚠️ End user has no assigned office section."

        new_section = new_enduser.employee_office_section

        with transaction.atomic():

            # 🔹 1. Get all assignments for this equipment
            all_assignments = PMScheduleAssignment.objects.filter(
                equipment_package=equipment_package
            ).select_related('pm_section_schedule__quarter_schedule', 'pm_section_schedule__section')

            # Track which quarters are already completed (keep in original section)
            completed_quarters = set()
            pending_assignments = []

            for assignment in all_assignments:
                quarter_key = (
                    assignment.pm_section_schedule.quarter_schedule.year,
                    assignment.pm_section_schedule.quarter_schedule.quarter
                )
                
                if assignment.is_completed:
                    # ✅ Keep completed assignments in their original section
                    completed_quarters.add(quarter_key)
                else:
                    # ❌ Mark pending assignments for deletion and recreation
                    pending_assignments.append(assignment)

            # 🔹 2. Delete only PENDING assignments (will be recreated in new section)
            deleted_count = len(pending_assignments)
            for assignment in pending_assignments:
                assignment.delete()

            # 🔹 3. Get all PM section schedules for the new section
            new_section_schedules = PMSectionSchedule.objects.filter(
                section=new_section
            ).select_related('quarter_schedule').order_by(
                'quarter_schedule__year',
                'quarter_schedule__quarter'
            )

            if not new_section_schedules.exists():
                return f"⚠️ No PM schedules found for {new_section.name} section."

            # 🔹 4. Create assignments in NEW section for quarters that aren't completed
            created_count = 0
            for sched in new_section_schedules:
                quarter_key = (sched.quarter_schedule.year, sched.quarter_schedule.quarter)
                
                # Skip if this quarter was already completed in the old section
                if quarter_key in completed_quarters:
                    continue
                
                # Create new assignment for pending quarters
                PMScheduleAssignment.objects.get_or_create(
                    equipment_package=equipment_package,
                    pm_section_schedule=sched,
                    defaults={
                        'is_completed': False,
                        'remarks': f'Transferred from previous section to {new_section.name}'
                    }
                )
                created_count += 1

            # 🔹 5. Build informative message
            if created_count > 0:
                return (
                    f"✅ PM schedule transferred to {new_section.name}. "
                    f"Deleted {deleted_count} pending assignment(s), "
                    f"created {created_count} new assignment(s). "
                    f"Completed PMs remain in original section."
                )
            else:
                return (
                    f"✅ All quarters already completed. "
                    f"No pending schedules to transfer to {new_section.name}."
                )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"❌ PM transfer error: {e}"
# ✅ AJAX: Check if serial number exists across multiple models (with exclude_id support)


def auto_transfer_pm_schedule_laptop(laptop_package, new_enduser):
    """
    Auto-transfer PM schedule for LAPTOP to the new End User's section.
    - Keeps completed PMs in their ORIGINAL section (history preservation)
    - Transfers only PENDING schedules to the new section
    - Creates new schedules for quarters not yet assigned
    """
    from inventory.models import PMScheduleAssignment, PMSectionSchedule, PreventiveMaintenance
    from django.db import transaction

    try:
        if not new_enduser or not new_enduser.employee_office_section:
            return "⚠️ End user has no assigned office section."

        new_section = new_enduser.employee_office_section

        with transaction.atomic():

            # 🔹 1. Get all assignments for this LAPTOP
            all_assignments = PMScheduleAssignment.objects.filter(
                laptop_package=laptop_package
            ).select_related('pm_section_schedule__quarter_schedule', 'pm_section_schedule__section')

            # Track which quarters are already completed (keep in original section)
            completed_quarters = set()
            pending_assignments = []

            for assignment in all_assignments:
                quarter_key = (
                    assignment.pm_section_schedule.quarter_schedule.year,
                    assignment.pm_section_schedule.quarter_schedule.quarter
                )
                
                if assignment.is_completed:
                    # ✅ Keep completed assignments in their original section
                    completed_quarters.add(quarter_key)
                else:
                    # ❌ Mark pending assignments for deletion and recreation
                    pending_assignments.append(assignment)

            # 🔹 2. Delete only PENDING assignments (will be recreated in new section)
            deleted_count = len(pending_assignments)
            for assignment in pending_assignments:
                assignment.delete()

            # 🔹 3. Get all PM section schedules for the new section
            new_section_schedules = PMSectionSchedule.objects.filter(
                section=new_section
            ).select_related('quarter_schedule').order_by(
                'quarter_schedule__year',
                'quarter_schedule__quarter'
            )

            if not new_section_schedules.exists():
                return f"⚠️ No PM schedules found for {new_section.name} section."

            # 🔹 4. Create assignments in NEW section for quarters that aren't completed
            created_count = 0
            for sched in new_section_schedules:
                quarter_key = (sched.quarter_schedule.year, sched.quarter_schedule.quarter)
                
                # Skip if this quarter was already completed in the old section
                if quarter_key in completed_quarters:
                    continue
                
                # Create new assignment for pending quarters
                PMScheduleAssignment.objects.get_or_create(
                    laptop_package=laptop_package,
                    pm_section_schedule=sched,
                    defaults={
                        'is_completed': False,
                        'remarks': f'Transferred from previous section to {new_section.name}'
                    }
                )
                created_count += 1

            # 🔹 5. Build informative message
            if created_count > 0:
                return (
                    f"✅ PM schedule transferred to {new_section.name}. "
                    f"Deleted {deleted_count} pending assignment(s), "
                    f"created {created_count} new assignment(s). "
                    f"Completed PMs remain in original section."
                )
            else:
                return (
                    f"✅ All quarters already completed. "
                    f"No pending schedules to transfer to {new_section.name}."
                )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"❌ PM transfer error: {e}"

def check_serial_no(request):
    serial = (request.GET.get('serial') or '').strip()
    category = (request.GET.get('category') or '').strip()
    exclude_id = request.GET.get('exclude_id')

    # Quick exit if missing
    if not serial or not category:
        return JsonResponse({'exists': False})

    serial_norm = serial.upper()

    # Map model + field by category
    model_map = {
        "Desktop": (DesktopDetails, "serial_no"),
        "Monitor": (MonitorDetails, "monitor_sn_db"),
        "Keyboard": (KeyboardDetails, "keyboard_sn_db"),
        "Mouse": (MouseDetails, "mouse_sn_db"),
        "UPS": (UPSDetails, "ups_sn_db"),
        "Laptop": (LaptopDetails, "laptop_sn_db"),
        "Printer": (PrinterDetails, "printer_sn_db"),
    }

    Model, field_name = model_map.get(category, (None, None))
    if not Model:
        return JsonResponse({'exists': False})

    # Base query normalized (case + whitespace insensitive)
    query = Model.objects.annotate(
        sn_norm=Upper(Trim(F(field_name)))
    ).filter(sn_norm=serial_norm)

    # ✅ Exclude current record if editing (prevents false duplicate)
    if exclude_id:
        query = query.exclude(id=exclude_id)

    exists = query.exists()
    return JsonResponse({'exists': exists})

# AJAX: Check if Monitor serial number exists in desktop details view
def check_monitor_sn(request):
    """AJAX: Check if a Monitor serial number already exists."""
    sn = request.GET.get("sn", "").strip().upper()
    package_id = request.GET.get("package_id")
    qs = MonitorDetails.objects.filter(monitor_sn_norm=sn)
    if package_id:
        qs = qs.exclude(equipment_package_id=package_id)
    return JsonResponse({"exists": qs.exists()})

# AJAX: Check if Keyboard serial number exists in desktop details view
def check_keyboard_sn(request):
    """AJAX: Check if a Keyboard serial number already exists."""
    sn = request.GET.get("sn", "").strip().upper()
    package_id = request.GET.get("package_id")
    qs = KeyboardDetails.objects.filter(keyboard_sn_norm=sn)
    if package_id:
        qs = qs.exclude(equipment_package_id=package_id)
    return JsonResponse({"exists": qs.exists()})

# AJAX: Check if Mouse serial number exists in desktop details view
def check_mouse_sn(request):
    """AJAX: Check if a Mouse serial number already exists."""
    sn = request.GET.get("sn", "").strip().upper()
    package_id = request.GET.get("package_id")
    qs = MouseDetails.objects.filter(mouse_sn_norm=sn)
    if package_id:
        qs = qs.exclude(equipment_package_id=package_id)
    return JsonResponse({"exists": qs.exists()})

# AJAX: Check if UPS serial number exists in desktop details view
def check_ups_sn(request):
    """AJAX: Check if a UPS serial number already exists."""
    sn = request.GET.get("sn", "").strip().upper()
    package_id = request.GET.get("package_id")
    qs = UPSDetails.objects.filter(ups_sn_norm=sn)
    if package_id:
        qs = qs.exclude(equipment_package_id=package_id)
    return JsonResponse({"exists": qs.exists()})

# Salvage logic for components
def salvage_monitor_logic(monitor, new_package=None, notes=None):
    # Ensure uniqueness by monitor_sn
    salvaged_monitor, created = SalvagedMonitor.objects.get_or_create(
        monitor_sn=monitor.monitor_sn_db,
        defaults={
            'monitor': monitor,
            'equipment_package': monitor.equipment_package,
            'monitor_brand': str(monitor.monitor_brand_db) if monitor.monitor_brand_db else None,
            'monitor_model': monitor.monitor_model_db,
            'monitor_size': monitor.monitor_size_db,
            'computer_name': monitor.equipment_package.computer_name,
            'asset_owner': getattr(monitor.equipment_package.desktop_details.first(), "asset_owner", None),
            'notes': notes or "Salvaged instead of disposed",
        }
    )

    # If row already exists, update it with latest info
    if not created:
        salvaged_monitor.monitor = monitor
        salvaged_monitor.equipment_package = monitor.equipment_package
        salvaged_monitor.monitor_brand = str(monitor.monitor_brand_db) if monitor.monitor_brand_db else None
        salvaged_monitor.monitor_model = monitor.monitor_model_db
        salvaged_monitor.monitor_size = monitor.monitor_size_db
        salvaged_monitor.computer_name = monitor.equipment_package.computer_name
        salvaged_monitor.asset_owner = getattr(monitor.equipment_package.desktop_details.first(), "asset_owner", None)
        if notes:
            salvaged_monitor.notes = notes

    if new_package:  # reassignment
        salvaged_monitor.is_reassigned = True
        salvaged_monitor.reassigned_to = new_package
        salvaged_monitor.computer_name = new_package.computer_name
        salvaged_monitor.asset_owner = getattr(new_package.desktop_details.first(), "asset_owner", None)
        salvaged_monitor.save()

        SalvagedMonitorHistory.objects.create(
            salvaged_monitor=salvaged_monitor,
            reassigned_to=new_package
        )
    else:  # available in salvage stock
        salvaged_monitor.is_reassigned = False
        salvaged_monitor.reassigned_to = None
        salvaged_monitor.save()

    return salvaged_monitor

def salvage_keyboard_logic(keyboard, new_package=None, notes=None):
    salvaged_keyboard, created = SalvagedKeyboard.objects.get_or_create(
        keyboard_sn=keyboard.keyboard_sn_db,
        defaults={
            'keyboard': keyboard,
            'equipment_package': keyboard.equipment_package,
            'keyboard_brand': str(keyboard.keyboard_brand_db) if keyboard.keyboard_brand_db else None,
            'keyboard_model': keyboard.keyboard_model_db,
            'computer_name': keyboard.equipment_package.computer_name,
            'asset_owner': getattr(keyboard.equipment_package.desktop_details.first(), "asset_owner", None),
            'notes': notes or "Salvaged instead of disposed",
        }
    )

    if not created:
        salvaged_keyboard.keyboard = keyboard
        salvaged_keyboard.equipment_package = keyboard.equipment_package
        salvaged_keyboard.keyboard_brand = str(keyboard.keyboard_brand_db) if keyboard.keyboard_brand_db else None
        salvaged_keyboard.keyboard_model = keyboard.keyboard_model_db
        salvaged_keyboard.computer_name = keyboard.equipment_package.computer_name
        salvaged_keyboard.asset_owner = getattr(keyboard.equipment_package.desktop_details.first(), "asset_owner", None)
        if notes:
            salvaged_keyboard.notes = notes

    if new_package:
        salvaged_keyboard.is_reassigned = True
        salvaged_keyboard.reassigned_to = new_package
        salvaged_keyboard.computer_name = new_package.computer_name
        salvaged_keyboard.asset_owner = getattr(new_package.desktop_details.first(), "asset_owner", None)
        salvaged_keyboard.save()

        SalvagedKeyboardHistory.objects.create(
            salvaged_keyboard=salvaged_keyboard,
            reassigned_to=new_package
        )
    else:
        salvaged_keyboard.is_reassigned = False
        salvaged_keyboard.reassigned_to = None
        salvaged_keyboard.save()

    return salvaged_keyboard


def salvage_mouse_logic(mouse, new_package=None, notes=None):
    salvaged_mouse, created = SalvagedMouse.objects.get_or_create(
        mouse_sn=mouse.mouse_sn_db,
        defaults={
            'mouse': mouse,
            'equipment_package': mouse.equipment_package,
            'mouse_brand': str(mouse.mouse_brand_db) if mouse.mouse_brand_db else None,
            'mouse_model': mouse.mouse_model_db,
            'computer_name': mouse.equipment_package.computer_name,
            'asset_owner': getattr(mouse.equipment_package.desktop_details.first(), "asset_owner", None),
            'notes': notes or "Salvaged instead of disposed",
        }
    )

    if not created:
        salvaged_mouse.mouse = mouse
        salvaged_mouse.equipment_package = mouse.equipment_package
        salvaged_mouse.mouse_brand = str(mouse.mouse_brand_db) if mouse.mouse_brand_db else None
        salvaged_mouse.mouse_model = mouse.mouse_model_db
        salvaged_mouse.computer_name = mouse.equipment_package.computer_name
        salvaged_mouse.asset_owner = getattr(mouse.equipment_package.desktop_details.first(), "asset_owner", None)
        if notes:
            salvaged_mouse.notes = notes

    if new_package:
        salvaged_mouse.is_reassigned = True
        salvaged_mouse.reassigned_to = new_package
        salvaged_mouse.computer_name = new_package.computer_name
        salvaged_mouse.asset_owner = getattr(new_package.desktop_details.first(), "asset_owner", None)
        salvaged_mouse.save()

        SalvagedMouseHistory.objects.create(
            salvaged_mouse=salvaged_mouse,
            reassigned_to=new_package
        )
    else:
        salvaged_mouse.is_reassigned = False
        salvaged_mouse.reassigned_to = None
        salvaged_mouse.save()

    return salvaged_mouse


def salvage_ups_logic(ups, new_package=None, notes=None):
    salvaged_ups, created = SalvagedUPS.objects.get_or_create(
        ups_sn=ups.ups_sn_db,
        defaults={
            'ups': ups,
            'equipment_package': ups.equipment_package,
            'ups_brand': str(ups.ups_brand_db) if ups.ups_brand_db else None,
            'ups_model': ups.ups_model_db,
            'ups_capacity': ups.ups_capacity_db,
            'computer_name': ups.equipment_package.computer_name,
            'asset_owner': getattr(ups.equipment_package.desktop_details.first(), "asset_owner", None),
            'notes': notes or "Salvaged instead of disposed",
        }
    )

    if not created:
        salvaged_ups.ups = ups
        salvaged_ups.equipment_package = ups.equipment_package
        salvaged_ups.ups_brand = str(ups.ups_brand_db) if ups.ups_brand_db else None
        salvaged_ups.ups_model = ups.ups_model_db
        salvaged_ups.ups_capacity = ups.ups_capacity_db
        salvaged_ups.computer_name = ups.equipment_package.computer_name
        salvaged_ups.asset_owner = getattr(ups.equipment_package.desktop_details.first(), "asset_owner", None)
        if notes:
            salvaged_ups.notes = notes

    if new_package:
        salvaged_ups.is_reassigned = True
        salvaged_ups.reassigned_to = new_package
        salvaged_ups.computer_name = new_package.computer_name
        salvaged_ups.asset_owner = getattr(new_package.desktop_details.first(), "asset_owner", None)
        salvaged_ups.save()

        SalvagedUPSHistory.objects.create(
            salvaged_ups=salvaged_ups,
            reassigned_to=new_package
        )
    else:
        salvaged_ups.is_reassigned = False
        salvaged_ups.reassigned_to = None
        salvaged_ups.save()

    return salvaged_ups


@login_required  

def success_page(request, package_id):
    """Success page after adding desktop, laptop, or printer package."""
    # Get the equipment type from querystring (?type=Desktop/Laptop/Printer)
    equipment_type = request.GET.get("type", "Unknown")

    # Default redirect (fallback if type is missing)
    redirect_url = reverse("dashboard")

    # Generate the correct redirect URL depending on equipment type
    if equipment_type == "Desktop":
        redirect_url = reverse("desktop_details_view", kwargs={"package_id": package_id})
    elif equipment_type == "Laptop":
        redirect_url = reverse("laptop_details_view", kwargs={"package_id": package_id})
    elif equipment_type == "Printer":
        redirect_url = reverse("printer_details_view", kwargs={"printer_id": package_id})  # ✅ FIXED HERE

    # Pass data to the template
    context = {
        "package_id": package_id,
        "equipment_type": equipment_type,
        "redirect_url": redirect_url,
    }
    return render(request, "success_add.html", context)
    


#Template: Desktop_details_view
@login_required
def equipment_package_base(request):
    # Fetch all desktop details
    desktop_details = DesktopDetails.objects.all()
    
    # Create a combined list where each desktop is paired with its keyboards
    desktops_with_items = []
    for desktop in desktop_details:
        keyboards = KeyboardDetails.objects.filter(equipment_package=desktop.equipment_package, is_disposed=False)
        user = UserDetails.objects.filter(equipment_package=desktop.equipment_package)
        desktops_with_items.append({
            'desktop': desktop,
            'keyboards': keyboards,  # This can have multiple entries per desktop
            'user': user
        })

    return render(request, 'desktop_details.html', {
        'desktops_with_items': desktops_with_items,
    })   



@login_required
def desktop_details_view(request, package_id):
    equipment_package = get_object_or_404(Equipment_Package, id=package_id)
    desktop_details = (
        DesktopDetails.objects.filter(equipment_package=equipment_package)
        .order_by('-id')
        .first()
    )
    if desktop_details:
        desktop_details.refresh_from_db()

    # Generate QR code if missing
    if not equipment_package.qr_code:
        equipment_package.generate_qr_code()
        equipment_package.save()

    # Active component details
    keyboard_detailsx = KeyboardDetails.objects.filter(equipment_package=equipment_package, is_disposed=False)
    mouse_details = MouseDetails.objects.filter(equipment_package=equipment_package, is_disposed=False)
    monitor_detailsx = MonitorDetails.objects.filter(equipment_package=equipment_package, is_disposed=False)
    ups_details = UPSDetails.objects.filter(equipment_package=equipment_package, is_disposed=False)
    documents_details = DocumentsDetails.objects.filter(equipment_package=equipment_package)
    user_details = UserDetails.objects.filter(equipment_package=equipment_package).first()

    # Disposed components
    disposed_desktop = DisposedDesktopDetail.objects.filter(desktop__equipment_package=equipment_package)
    disposed_keyboards = DisposedKeyboard.objects.filter(keyboard_dispose_db__equipment_package=equipment_package)
    disposed_mouse = DisposedMouse.objects.filter(mouse_db__equipment_package=equipment_package)
    disposed_monitor = DisposedMonitor.objects.filter(monitor_disposed_db__equipment_package=equipment_package)
    disposed_ups = DisposedUPS.objects.filter(ups_db__equipment_package=equipment_package)

    # Flags for existence
    has_active_desktop = DesktopDetails.objects.filter(equipment_package=equipment_package, is_disposed=False).exists()
    has_active_keyboards = keyboard_detailsx.exists()
    has_active_mouse = mouse_details.exists()
    has_active_monitor = monitor_detailsx.exists()
    has_active_ups = ups_details.exists()
    desktops_disposed_filter = DesktopDetails.objects.filter(equipment_package=equipment_package, is_disposed=False)

    # Brand filters
    desktop_brands  = Brand.objects.filter(is_desktop=True)
    monitor_brands  = Brand.objects.filter(is_monitor=True)
    keyboard_brands = Brand.objects.filter(is_keyboard=True)
    mouse_brands    = Brand.objects.filter(is_mouse=True)
    ups_brands      = Brand.objects.filter(is_ups=True)

    #Preventive Maintenance
    # Preventive Maintenance Schedule Assignments
    pm_assignments = PMScheduleAssignment.objects.filter(
        equipment_package=equipment_package
    ).select_related(
        'pm_section_schedule__section',
        'pm_section_schedule__quarter_schedule'
    )

    # Preventive Maintenance History (Completed records)
    maintenance_records = (
        PreventiveMaintenance.objects
        .filter(equipment_package=equipment_package)
        .select_related(
            "pm_schedule_assignment__pm_section_schedule__quarter_schedule",
            "pm_schedule_assignment__pm_section_schedule__section"
        )
        .order_by("-date_accomplished")
    )

    # Change history
    enduser_history = get_end_user_history(equipment_package)
    assetowner_history = get_asset_owner_history(equipment_package)

    # Employees for dropdowns
    employees = Employee.objects.all()

     # ✅ Salvaged for re-adding
    salvaged_monitors = SalvagedMonitor.objects.filter(is_reassigned=False).order_by("-salvage_date")
    salvaged_keyboards = SalvagedKeyboard.objects.filter(is_reassigned=False).order_by("-salvage_date")
    salvaged_mice = SalvagedMouse.objects.filter(is_reassigned=False).order_by("-salvage_date")
    salvaged_ups = SalvagedUPS.objects.filter(is_reassigned=False).order_by("-salvage_date")
    
    # ✅ Salvaged components in view details
    salvaged_monitors_view = SalvagedMonitor.objects.filter(
        equipment_package=equipment_package,
        is_reassigned=False,
        is_disposed=False
    ).order_by("-salvage_date")

    salvaged_keyboards_view = SalvagedKeyboard.objects.filter(
        equipment_package=equipment_package,
        is_reassigned=False,
        is_disposed=False
    ).order_by("-salvage_date")

    salvaged_mice_view = SalvagedMouse.objects.filter(
        equipment_package=equipment_package,
        is_reassigned=False,
        is_disposed=False
    ).order_by("-salvage_date")

    salvaged_ups_view = SalvagedUPS.objects.filter(
        equipment_package=equipment_package,
        is_reassigned=False,
        is_disposed=False
    ).order_by("-salvage_date")

    return render(request, 'desktop_details_view.html', {
        'desktop_detailsx': desktop_details,
        'equipment_package': equipment_package,
        'keyboard_detailse': keyboard_detailsx.first(),
        'monitor_detailse': monitor_detailsx.first(),
        'mouse_detailse': mouse_details.first(),
        'ups_detailse': ups_details.first(),
        'documents_detailse': documents_details.first(),
        'user_details': user_details,
        'disposed_desktop': disposed_desktop,
        'disposed_keyboards': disposed_keyboards,
        'disposed_mouse': disposed_mouse,
        'disposed_monitor': disposed_monitor,
        'disposed_ups': disposed_ups,
        'has_active_desktop': has_active_desktop,
        'has_active_keyboards': has_active_keyboards,
        'has_active_mouse': has_active_mouse,
        'has_active_monitor': has_active_monitor,
        'has_active_ups': has_active_ups,
        'desktops_disposed_filter': desktops_disposed_filter,
        'desktop_brands': desktop_brands,
        'monitor_brands': monitor_brands,
        'keyboard_brands': keyboard_brands,
        'mouse_brands': mouse_brands,
        'ups_brands': ups_brands,

        'employees': employees,
        'enduser_history': enduser_history,
        'assetowner_history': assetowner_history,
        'pm_assignments': pm_assignments,
        'maintenance_records': maintenance_records,
        'salvaged_monitors': salvaged_monitors,
        'salvaged_keyboards': salvaged_keyboards,
        'salvaged_mice': salvaged_mice,
        'salvaged_ups': salvaged_ups,
        'salvaged_monitors_view': salvaged_monitors_view,
        'salvaged_keyboards_view': salvaged_keyboards_view,
        'salvaged_mice_view': salvaged_mice_view,
        'salvaged_ups_view': salvaged_ups_view,
    })






def keyboard_detailed_view(request, keyboard_id):
    # Get the specific keyboard using its ID
    keyboard = get_object_or_404(KeyboardDetails, id=keyboard_id)
    # Render the detailed view of the keyboard
    return render(request, 'keyboard_detailed_view.html', {'keyboard': keyboard})




@require_POST
def update_desktop(request, pk):
    try:
        desktop = get_object_or_404(DesktopDetails, pk=pk)

        # ✅ Basic Info
        desktop.serial_no = request.POST.get('desktop_sn_form', '').strip()
        desktop.model = request.POST.get('desktop_model_form', '').strip()
        desktop.processor = request.POST.get('desktop_proccessor_form', '').strip()
        desktop.memory = request.POST.get('desktop_memory_form', '').strip()
        desktop.drive = request.POST.get('desktop_drive_form', '').strip()

        # ✅ Brand Update (validated)
        brand_id = request.POST.get('desktop_brand_form')
        if brand_id:
            try:
                desktop.brand_name = get_object_or_404(Brand, pk=brand_id)
            except:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid brand selected.'
                }, status=400)

        # ✅ OS and Office Versions + Keys
        desktop.desktop_OS = request.POST.get('desktop_OS', desktop.desktop_OS)
        desktop.desktop_OS_keys = request.POST.get('desktop_OS_keys', desktop.desktop_OS_keys)
        desktop.desktop_Office = request.POST.get('desktop_Office', desktop.desktop_Office)
        desktop.desktop_Office_keys = request.POST.get('desktop_Office_keys', desktop.desktop_Office_keys)

        # ✅ Validate required fields
        if not desktop.serial_no or not desktop.model:
            return JsonResponse({
                'success': False,
                'error': 'Serial number and model are required.'
            }, status=400)

        # ✅ Save all updates
        desktop.save()

        # ✅ Return JSON response for AJAX
        redirect_url = reverse('desktop_details_view', kwargs={'package_id': desktop.equipment_package.pk})
        return JsonResponse({
            'success': True,
            'message': 'Desktop details updated successfully!',
            'redirect_url': f'{redirect_url}#pills-desktop'
        })

    except DesktopDetails.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Desktop not found.'
        }, status=404)
    
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error updating desktop {pk}: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'error': f'An error occurred while updating: {str(e)}'
        }, status=500)


# Upload desktop photo
@require_POST
def upload_desktop_photo(request, pk):
    try:
        desktop = get_object_or_404(DesktopDetails, pk=pk)

        if 'desktop_photo' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'No photo file provided.'
            }, status=400)

        # Handle photo upload
        desktop.desktop_photo = request.FILES['desktop_photo']
        desktop.save()

        return JsonResponse({
            'success': True,
            'message': 'Desktop photo uploaded successfully!',
            'photo_url': desktop.desktop_photo.url if desktop.desktop_photo else None
        })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error uploading desktop photo {pk}: {str(e)}")

        return JsonResponse({
            'success': False,
            'error': f'Failed to upload photo: {str(e)}'
        }, status=500)


# Delete desktop photo
@require_POST
def delete_desktop_photo(request, pk):
    try:
        desktop = get_object_or_404(DesktopDetails, pk=pk)

        if desktop.desktop_photo:
            # Delete the file from storage
            desktop.desktop_photo.delete(save=False)
            desktop.desktop_photo = None
            desktop.save()

            return JsonResponse({
                'success': True,
                'message': 'Desktop photo deleted successfully!'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No photo to delete.'
            }, status=400)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error deleting desktop photo {pk}: {str(e)}")

        return JsonResponse({
            'success': False,
            'error': f'Failed to delete photo: {str(e)}'
        }, status=500)


#update monitor details
@require_POST
def update_monitor(request, pk):
    try:
        monitor = get_object_or_404(MonitorDetails, pk=pk)
        monitor.monitor_sn_db = request.POST.get('monitor_sn_db')
        brand_id = request.POST.get('monitor_brand_db')
        monitor.monitor_brand_db = get_object_or_404(Brand, pk=brand_id)
        monitor.monitor_model_db = request.POST.get('monitor_model_db')
        monitor.monitor_size_db = request.POST.get('monitor_size_db')
        monitor.save()

        base_url = reverse('desktop_details_view', kwargs={'package_id': monitor.equipment_package.pk})
        redirect_url = f"{base_url}#pills-monitor"

        return JsonResponse({
            'success': True,
            'message': 'Monitor updated successfully!',
            'redirect_url': redirect_url
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating monitor: {str(e)}'
        })


#update keyboard details

def update_keyboard(request, pk):
    try:
        keyboard = get_object_or_404(KeyboardDetails, pk=pk)

        # 🧩 Update fields
        keyboard.keyboard_sn_db = request.POST.get('keyboard_sn_db')
        brand_id = request.POST.get('keyboard_brand_db')
        keyboard.keyboard_brand_db = get_object_or_404(Brand, pk=brand_id)
        keyboard.keyboard_model_db = request.POST.get('keyboard_model_db')

        keyboard.save()

        # ✅ Prepare redirect URL for same page + same pill
        base_url = reverse('desktop_details_view', kwargs={'package_id': keyboard.equipment_package.pk})
        redirect_url = f"{base_url}#pills-keyboard"

        # ✅ Return JSON for AJAX success
        return JsonResponse({
            'success': True,
            'message': 'Keyboard updated successfully!',
            'redirect_url': redirect_url
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating keyboard: {str(e)}'
        })




@require_POST
def update_mouse(request, pk):
    try:
        mouse = get_object_or_404(MouseDetails, pk=pk)
        mouse.mouse_sn_db = request.POST.get('mouse_sn_db')

        brand_id = request.POST.get('mouse_brand_db')
        mouse.mouse_brand_db = get_object_or_404(Brand, pk=brand_id)
        mouse.mouse_model_db = request.POST.get('mouse_model_db')

        mouse.save()

        base_url = reverse('desktop_details_view', kwargs={'package_id': mouse.equipment_package.pk})
        redirect_url = f"{base_url}#pills-mouse"

        return JsonResponse({
            'success': True,
            'message': 'Mouse updated successfully!',
            'redirect_url': redirect_url
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating mouse: {str(e)}'
        })


# 🧱 UPDATE UPS
@require_POST
def update_ups(request, pk):
    try:
        ups = get_object_or_404(UPSDetails, pk=pk)
        ups.ups_sn_db = request.POST.get('ups_sn_db')

        brand_id = request.POST.get('ups_brand_db')
        ups.ups_brand_db = get_object_or_404(Brand, pk=brand_id)
        ups.ups_model_db = request.POST.get('ups_model_db')
        ups.ups_capacity_db = request.POST.get('ups_capacity_db')
        ups.save()

        base_url = reverse('desktop_details_view', kwargs={'package_id': ups.equipment_package.pk})
        redirect_url = f"{base_url}#pills-ups"

        return JsonResponse({
            'success': True,
            'message': 'UPS details updated successfully!',
            'redirect_url': redirect_url
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating UPS: {str(e)}'
        })

@require_POST
def update_documents(request, pk):
    try:
        documents = get_object_or_404(DocumentsDetails, pk=pk)
        documents.docs_PAR = request.POST.get('docs_PAR')
        documents.docs_Propertyno = request.POST.get('docs_Propertyno')
        documents.docs_Acquisition_Type = request.POST.get('docs_Acquisition_Type')
        documents.docs_Value = request.POST.get('docs_Value')
        documents.docs_Datereceived = request.POST.get('docs_Datereceived')
        documents.docs_Dateinspected = request.POST.get('docs_Dateinspected')
        documents.docs_Supplier = request.POST.get('docs_Supplier')
        documents.docs_Status = request.POST.get('docs_Status')

        documents.save()

        base_url = reverse('desktop_details_view', kwargs={'package_id': documents.equipment_package.pk})
        redirect_url = f"{base_url}#pills-documents"

        return JsonResponse({
            'success': True,
            'message': 'Document details updated successfully!',
            'redirect_url': redirect_url
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating documents: {str(e)}'
        })

                                            ######## SINGLE DISPOSAL TAB ###########


@require_POST
def keyboard_disposed(request, keyboard_id):
    try:
        keyboard = get_object_or_404(KeyboardDetails, id=keyboard_id)
        keyboard.is_disposed = True
        keyboard.save()

        # Create DisposedKeyboard record
        DisposedKeyboard.objects.create(
            keyboard_dispose_db=keyboard,
            equipment_package=keyboard.equipment_package,
            disposed_under=None,
            disposal_date=timezone.now()
        )

        # Mark salvaged record as disposed (if any)
        SalvagedKeyboard.objects.filter(keyboard_sn=keyboard.keyboard_sn_db).update(
            is_disposed=True,
            disposed_date=timezone.now()
        )

        # Redirect URL (for reloading the Keyboard tab)
        base_url = reverse('desktop_details_view', kwargs={'package_id': keyboard.equipment_package.pk})
        redirect_url = f"{base_url}#pills-keyboard"

        return JsonResponse({
            'success': True,
            'message': 'Keyboard disposed successfully!',
            'redirect_url': redirect_url
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error disposing keyboard: {str(e)}'
        })


# Mouse disposal under Mouse pill page
@require_POST
def mouse_disposed(request, mouse_id):
    try:
        mouse = get_object_or_404(MouseDetails, id=mouse_id)
        mouse.is_disposed = True
        mouse.save()

        DisposedMouse.objects.create(
            mouse_db=mouse,
            equipment_package=mouse.equipment_package,
            disposed_under=None,
            disposal_date=timezone.now()
        )

        SalvagedMouse.objects.filter(mouse_sn=mouse.mouse_sn_db).update(
            is_disposed=True,
            disposed_date=timezone.now()
        )

        base_url = reverse('desktop_details_view', kwargs={'package_id': mouse.equipment_package.pk})
        redirect_url = f"{base_url}#pills-mouse"

        return JsonResponse({
            'success': True,
            'message': f"✅ Mouse '{mouse.mouse_model_db}' disposed successfully!",
            'redirect_url': redirect_url
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f"Error disposing mouse: {str(e)}"
        })



                            ############### UPS disposal under UPS pill page  #############
# 🧱 DISPOSE UPS
@require_POST
def ups_disposed(request, ups_id):
    try:
        ups = get_object_or_404(UPSDetails, id=ups_id)
        ups.is_disposed = True
        ups.save()

        DisposedUPS.objects.create(
            ups_db=ups,
            equipment_package=ups.equipment_package,
            disposed_under=None,
            disposal_date=timezone.now()
        )

        SalvagedUPS.objects.filter(ups_sn=ups.ups_sn_db).update(
            is_disposed=True,
            disposed_date=timezone.now()
        )

        base_url = reverse('desktop_details_view', kwargs={'package_id': ups.equipment_package.pk})
        redirect_url = f"{base_url}#pills-ups"

        return JsonResponse({
            'success': True,
            'message': f"✅ UPS '{ups.ups_model_db}' disposed successfully!",
            'redirect_url': redirect_url
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error disposing UPS: {str(e)}'
        })

                                            ######## SINGLE END TAB ###########




#viewing of all keyboard disposed sa page ni sa keyboard na naay disposed pud.
def disposed_keyboards(request):
    # Get all disposed keyboards
    disposed_keyboards = DisposedKeyboard.objects.all()
    # Render the list of disposed keyboards to the template
    return render(request, 'disposed_keyboards.html', {'disposed_keyboards': disposed_keyboards})



# END ################ (KEYBOARD END)


#This function retrieves all mouse records and renders them in a similar way as mouse_details.
def monitor_details(request):
    """
    Show all monitor states per serial:
    - Active: MonitorDetails (is_disposed=False) whose serial is NOT in salvaged list
    - Salvaged/Reassigned/Disposed: ALL SalvagedMonitor records (regardless of flags)
    """
    # 1) Get ALL salvaged monitors (no filter needed!)
    salvaged_qs = (
        SalvagedMonitor.objects
        .all()  # ← Get ALL salvaged monitors including disposed
        .select_related("reassigned_to")
        .prefetch_related("reassigned_to__user_details__user_Enduser")
    )

    # Normalize serials to avoid duplicates
    salvaged_serials = {
        (s.monitor_sn or "").strip().lower()
        for s in salvaged_qs
    }

    # 2) Active monitors EXCLUDING any serial that's in salvaged table
    active_qs = (
        MonitorDetails.objects
        .filter(is_disposed=False)
        .select_related("equipment_package", "monitor_brand_db")
        .prefetch_related("equipment_package__user_details__user_Enduser")
    )

    active_data = []
    for m in active_qs:
        sn_norm = (m.monitor_sn_db or "").strip().lower()
        if sn_norm in salvaged_serials:
            continue

        active_data.append({
            "id": m.id,
            "monitor_sn_db": m.monitor_sn_db,
            "monitor_brand_db": m.monitor_brand_db,
            "monitor_model_db": m.monitor_model_db,
            "monitor_size_db": m.monitor_size_db,
            "equipment_package": m.equipment_package,
            "status": "active",
            "salvaged_id": None,
        })

    # 3) Format salvaged rows - ALL of them!
    salvaged_data = []
    for s in salvaged_qs:
        # ✅ Determine status based on flags
        if s.is_disposed:
            status = "disposed"      # 🆕 fixed to show disposed properly
        elif s.is_reassigned:
            status = "reassigned"
        else:
            status = "salvaged"      # newly salvaged and available

        # Use reassigned_to if available, otherwise original equipment_package
        display_package = s.reassigned_to if s.reassigned_to else s.equipment_package

        salvaged_data.append({
            "id": s.id,
            "monitor_sn_db": s.monitor_sn,
            "monitor_brand_db": s.monitor_brand,
            "monitor_model_db": s.monitor_model,
            "monitor_size_db": s.monitor_size,
            "equipment_package": display_package,
            "status": status,
            "salvaged_id": s.id,
        })

    monitors = active_data + salvaged_data

    return render(request, "monitor_details.html", {"monitors": monitors})



def monitor_timeline_detail(request, salvaged_id):
    """Detail view for salvaged/reassigned monitor with history"""
    salvaged_monitor = get_object_or_404(SalvagedMonitor, pk=salvaged_id)
    history = salvaged_monitor.history.order_by('-reassigned_at')
    
    # Determine status for display
    if salvaged_monitor.is_disposed:
        status_label = "Salvaged"
        status_class = "danger"
    elif salvaged_monitor.is_reassigned:
        status_label = "Reassigned"
        status_class = "secondary"
    else:
        status_label = "Available"
        status_class = "success"
    
    return render(request, 'salvage/monitor_timeline_detail.html', {
        "monitor": salvaged_monitor,
        "history": history,
        "status_label": status_label,
        "status_class": status_class,
    })


# ==================== KEYBOARD DETAILS ====================
def keyboard_details(request):
    """
    Show all keyboard states per serial:
    - Active: KeyboardDetails (is_disposed=False) whose serial is NOT in salvaged list
    - Salvaged/Reassigned: ALL SalvagedKeyboard records (regardless of flags)
    """
    # 1) Get ALL salvaged keyboards (no filter!)
    # This includes newly salvaged (available), reassigned, and disposed
    salvaged_qs = (
        SalvagedKeyboard.objects
        .all()  # ← FIXED: Get ALL salvaged keyboards
        .select_related("reassigned_to")
        .prefetch_related("reassigned_to__user_details__user_Enduser")
    )

    # Normalize serials to avoid duplicates
    salvaged_serials = {
        (s.keyboard_sn or "").strip().lower()
        for s in salvaged_qs
    }

    # 2) Active keyboards EXCLUDING any serial that's in salvaged table
    active_qs = (
        KeyboardDetails.objects
        .filter(is_disposed=False)
        .select_related("equipment_package", "keyboard_brand_db")
        .prefetch_related("equipment_package__user_details__user_Enduser")
    )

    active_data = []
    for k in active_qs:
        sn_norm = (k.keyboard_sn_db or "").strip().lower()
        if sn_norm in salvaged_serials:
            continue

        active_data.append({
            "id": k.id,
            "keyboard_sn_db": k.keyboard_sn_db,
            "keyboard_brand_db": k.keyboard_brand_db,
            "keyboard_model_db": k.keyboard_model_db,
            "equipment_package": k.equipment_package,
            "status": "active",
            "salvaged_id": None,
        })

    # 3) Format salvaged keyboards - ALL of them!
    salvaged_data = []
    for s in salvaged_qs:
        # Determine status
        if s.is_disposed:
            status = "disposed"
        elif s.is_reassigned:
            status = "reassigned"
        else:
            # Newly salvaged (available for reassignment)
            status = "salvaged"
        
        # Use reassigned_to if available, otherwise original equipment_package
        display_package = s.reassigned_to if s.reassigned_to else s.equipment_package
        
        salvaged_data.append({
            "id": s.id,
            "keyboard_sn_db": s.keyboard_sn,
            "keyboard_brand_db": s.keyboard_brand,
            "keyboard_model_db": s.keyboard_model,
            "equipment_package": display_package,
            "status": status,
            "salvaged_id": s.id,
        })

    keyboards = active_data + salvaged_data
    return render(request, "keyboard_details.html", {"keyboards": keyboards})


def keyboard_timeline_detail(request, salvaged_id):
    """Detail view for salvaged/reassigned keyboard with history"""
    salvaged_keyboard = get_object_or_404(SalvagedKeyboard, pk=salvaged_id)
    history = salvaged_keyboard.history.order_by('-reassigned_at')
    
    # Determine status for display
    if salvaged_keyboard.is_disposed:
        status_label = "Salvaged"
        status_class = "danger"
    elif salvaged_keyboard.is_reassigned:
        status_label = "Reassigned"
        status_class = "secondary"
    else:
        status_label = "Available"
        status_class = "success"
    
    return render(request, 'salvage/keyboard_timeline_detail.html', {
        "keyboard": salvaged_keyboard,
        "history": history,
        "status_label": status_label,
        "status_class": status_class,
    })

# ==================== MOUSE DETAILS ====================
# ==================== MOUSE DETAILS ====================
def mouse_details(request):
    """
    Show all mouse states per serial:
    - Active: MouseDetails (is_disposed=False) whose serial is NOT in salvaged list
    - Salvaged/Reassigned/Disposed: ALL SalvagedMouse records (regardless of flags)
    """
    # 1) Get ALL salvaged mice (including newly salvaged, reassigned, and disposed)
    salvaged_qs = (
        SalvagedMouse.objects
        .all()
        .select_related("reassigned_to")
        .prefetch_related("reassigned_to__user_details__user_Enduser")
    )

    # Normalize serials to avoid duplicates
    salvaged_serials = {
        (s.mouse_sn or "").strip().lower()
        for s in salvaged_qs
    }

    # 2) Active mice EXCLUDING those already salvaged
    active_qs = (
        MouseDetails.objects
        .filter(is_disposed=False)
        .select_related("equipment_package", "mouse_brand_db")
        .prefetch_related("equipment_package__user_details__user_Enduser")
    )

    active_data = []
    for m in active_qs:
        sn_norm = (m.mouse_sn_db or "").strip().lower()
        if sn_norm in salvaged_serials:
            continue
        active_data.append({
            "id": m.id,
            "mouse_sn_db": m.mouse_sn_db,
            "mouse_brand_db": m.mouse_brand_db,
            "mouse_model_db": m.mouse_model_db,
            "equipment_package": m.equipment_package,
            "status": "active",
            "salvaged_id": None,
        })

    # 3) Salvaged mice records
    salvaged_data = []
    for s in salvaged_qs:
        if s.is_disposed:
            status = "disposed"
        elif s.is_reassigned:
            status = "reassigned"
        else:
            status = "salvaged"

        display_package = s.reassigned_to if s.reassigned_to else s.equipment_package

        salvaged_data.append({
            "id": s.id,
            "mouse_sn_db": s.mouse_sn,
            "mouse_brand_db": s.mouse_brand,
            "mouse_model_db": s.mouse_model,
            "equipment_package": display_package,
            "status": status,
            "salvaged_id": s.id,
        })

    mice = active_data + salvaged_data
    return render(request, "mouse_details.html", {"mice": mice})



def mouse_timeline_detail(request, salvaged_id):
    salvaged_mouse = get_object_or_404(SalvagedMouse, pk=salvaged_id)
    history = salvaged_mouse.history.order_by('-reassigned_at')

    if salvaged_mouse.is_disposed:
        status_label = "Disposed"
        status_class = "danger"
    elif salvaged_mouse.is_reassigned:
        status_label = "Reassigned"
        status_class = "secondary"
    else:
        status_label = "Salvaged"
        status_class = "success"

    return render(request, 'salvage/mouse_timeline_detail.html', {
        "mouse": salvaged_mouse,
        "history": history,
        "status_label": status_label,
        "status_class": status_class,
    })



# ==================== UPS DETAILS ====================
def ups_details(request):
    """
    Show all UPS states per serial:
    - Active: UPSDetails (is_disposed=False) whose serial is NOT in salvaged list
    - Salvaged/Reassigned: ALL SalvagedUPS records (regardless of flags)
    """
    # 1) Get ALL salvaged UPS units (no filter!)
    salvaged_qs = (
        SalvagedUPS.objects
        .all()  # ← FIXED: Get ALL salvaged UPS
        .select_related("reassigned_to")
        .prefetch_related("reassigned_to__user_details__user_Enduser")
    )

    # Normalize serials
    salvaged_serials = {
        (s.ups_sn or "").strip().lower()
        for s in salvaged_qs
    }

    # 2) Active UPS units EXCLUDING salvaged serials
    active_qs = (
        UPSDetails.objects
        .filter(is_disposed=False)
        .select_related("equipment_package", "ups_brand_db")
        .prefetch_related("equipment_package__user_details__user_Enduser")
    )

    active_data = []
    for u in active_qs:
        sn_norm = (u.ups_sn_db or "").strip().lower()
        if sn_norm in salvaged_serials:
            continue

        active_data.append({
            "id": u.id,
            "ups_sn_db": u.ups_sn_db,
            "ups_brand_db": u.ups_brand_db,
            "ups_model_db": u.ups_model_db,
            "equipment_package": u.equipment_package,
            "status": "active",
            "salvaged_id": None,
        })

    # 3) Format salvaged UPS units
    salvaged_data = []
    for s in salvaged_qs:
        # Determine status
        if s.is_disposed:
            status = "disposed"
        elif s.is_reassigned:
            status = "reassigned"
        else:
            status = "salvaged"
        
        display_package = s.reassigned_to if s.reassigned_to else s.equipment_package
        
        salvaged_data.append({
            "id": s.id,
            "ups_sn_db": s.ups_sn,
            "ups_brand_db": s.ups_brand,
            "ups_model_db": s.ups_model,
            "equipment_package": display_package,
            "status": status,
            "salvaged_id": s.id,
        })

    ups_list = active_data + salvaged_data
    return render(request, "ups_details.html", {"ups_list": ups_list})


def ups_timeline_detail(request, salvaged_id):
    """Detail view for salvaged/reassigned UPS with history"""
    salvaged_ups = get_object_or_404(SalvagedUPS, pk=salvaged_id)
    history = salvaged_ups.history.order_by('-reassigned_at')
    
    # Determine status for display
    if salvaged_ups.is_disposed:
        status_label = "Salvaged"
        status_class = "danger"
    elif salvaged_ups.is_reassigned:
        status_label = "Reassigned"
        status_class = "secondary"
    else:
        status_label = "Available"
        status_class = "success"
    
    return render(request, 'salvage/ups_timeline_detail.html', {
        "ups": salvaged_ups,
        "history": history,
        "status_label": status_label,
        "status_class": status_class,
    })


#This function retrieves the details of a specific mouse by its ID.
def mouse_detailed_view(request, mouse_id):
    # Get the specific mouse using its ID
    mouse = get_object_or_404(MouseDetails, id=mouse_id)
    # Render the detailed view of the mouse
    return render(request, 'mouse_detailed_view.html', {'mouse': mouse})



@require_POST
def monitor_disposed(request, monitor_id):
    try:
        monitor = get_object_or_404(MonitorDetails, id=monitor_id)
        monitor.is_disposed = True
        monitor.save()

        DisposedMonitor.objects.create(
            monitor_disposed_db=monitor,
            equipment_package=monitor.equipment_package,
            monitor_sn=monitor.monitor_sn_db,
            monitor_brand=str(monitor.monitor_brand_db) if monitor.monitor_brand_db else None,
            monitor_model=monitor.monitor_model_db,
            monitor_size=monitor.monitor_size_db,
            reason=request.POST.get("reason") or "Disposed individually",
        )

        SalvagedMonitor.objects.filter(monitor_sn=monitor.monitor_sn_db).update(
            is_disposed=True,
            disposed_date=timezone.now()
        )

        base_url = reverse('desktop_details_view', kwargs={'package_id': monitor.equipment_package.pk})
        redirect_url = f"{base_url}#pills-monitor"

        return JsonResponse({
            'success': True,
            'message': f"✅ Monitor '{monitor.monitor_model_db}' disposed successfully!",
            'redirect_url': redirect_url
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f"Error disposing monitor: {str(e)}"
        })




@login_required
def add_monitor_to_package(request, package_id):
    """
    FIXED: Properly handle salvaged monitor reassignment
    - Don't create duplicate MonitorDetails
    - Reactivate existing disposed monitor instead
    """
    equipment_package = get_object_or_404(Equipment_Package, pk=package_id)

    if request.method == "POST":
        salvaged_monitor_id = request.POST.get("salvaged_monitor_id")

        try:
            # ==================== CASE 1: SALVAGED MONITOR ====================
            if salvaged_monitor_id:
                salvaged_monitor = get_object_or_404(SalvagedMonitor, id=salvaged_monitor_id)
                
                # Check if already reassigned
                if salvaged_monitor.is_reassigned:
                    msg = "❌ This salvaged monitor has already been reassigned."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                sn_norm = salvaged_monitor.monitor_sn.strip().upper()
                
                # ✅ FIXED: Check if monitor exists (might be disposed)
                existing_monitor = MonitorDetails.objects.filter(monitor_sn_norm=sn_norm).first()
                
                if existing_monitor:
                    # ✅ REACTIVATE existing monitor instead of creating new one
                    if existing_monitor.is_disposed:
                        # Reactivate the disposed monitor
                        existing_monitor.equipment_package = equipment_package
                        existing_monitor.is_disposed = False
                        existing_monitor.save()
                        
                        msg = "✅ Salvaged monitor reactivated and reassigned successfully."
                    else:
                        # Monitor is already active in another package
                        msg = f"❌ Monitor '{sn_norm}' is already active in another package."
                        if request.headers.get("x-requested-with") == "XMLHttpRequest":
                            return JsonResponse({'success': False, 'message': msg})
                        messages.error(request, msg)
                        return redirect("desktop_details_view", package_id=equipment_package.id)
                else:
                    # ✅ No existing record - create new one
                    existing_monitor = MonitorDetails.objects.create(
                        equipment_package=equipment_package,
                        monitor_sn_db=salvaged_monitor.monitor_sn,
                        monitor_brand_db=Brand.objects.filter(name=salvaged_monitor.monitor_brand).first(),
                        monitor_model_db=salvaged_monitor.monitor_model,
                        monitor_size_db=salvaged_monitor.monitor_size,
                        is_disposed=False,
                    )
                    msg = "✅ Salvaged monitor reassigned and logged."

                # Update salvaged monitor record
                salvaged_monitor.is_reassigned = True
                salvaged_monitor.reassigned_to = equipment_package
                salvaged_monitor.save()

                # Log reassignment history
                SalvagedMonitorHistory.objects.create(
                    salvaged_monitor=salvaged_monitor,
                    reassigned_to=equipment_package,
                )

            # ==================== CASE 2: MANUAL INPUT ====================
            else:
                monitor_sn = request.POST.get("monitor_sn", "").strip()
                monitor_brand_id = request.POST.get("monitor_brand_db")
                monitor_model = request.POST.get("monitor_model")
                monitor_size = request.POST.get("monitor_size")

                if not monitor_sn or not monitor_model:
                    msg = "❌ Please fill in all required fields."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                sn_norm = monitor_sn.upper()
                
                # ✅ FIXED: Check for both active and disposed monitors
                existing_monitor = MonitorDetails.objects.filter(monitor_sn_norm=sn_norm).first()
                
                if existing_monitor and not existing_monitor.is_disposed:
                    # Monitor is already active
                    msg = f"❌ A monitor with serial number '{monitor_sn}' already exists and is active."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                brand_instance = Brand.objects.filter(id=monitor_brand_id).first() if monitor_brand_id else None

                # Create new monitor (or reactivate if exists but disposed)
                if existing_monitor and existing_monitor.is_disposed:
                    # Reactivate disposed monitor
                    existing_monitor.equipment_package = equipment_package
                    existing_monitor.monitor_brand_db = brand_instance
                    existing_monitor.monitor_model_db = monitor_model
                    existing_monitor.monitor_size_db = monitor_size
                    existing_monitor.is_disposed = False
                    existing_monitor.save()
                else:
                    # Create new monitor
                    MonitorDetails.objects.create(
                        equipment_package=equipment_package,
                        monitor_sn_db=monitor_sn,
                        monitor_brand_db=brand_instance,
                        monitor_model_db=monitor_model,
                        monitor_size_db=monitor_size,
                        is_disposed=False,
                    )

                msg = "✅ New monitor added successfully."

            base_url = reverse('desktop_details_view', kwargs={'package_id': equipment_package.pk})
            redirect_url = f"{base_url}#pills-monitor"

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': True, 'message': msg, 'redirect_url': redirect_url})

            messages.success(request, msg)
            return redirect(redirect_url)

        except Exception as e:
            msg = f"❌ Error: {str(e)}"
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': False, 'message': msg})
            messages.error(request, msg)
            return redirect("desktop_details_view", package_id=equipment_package.id)

    return redirect("desktop_details_view", package_id=equipment_package.id)


@login_required
def add_keyboard_to_package(request, package_id):
    """Add or reassign keyboard to equipment package"""
    equipment_package = get_object_or_404(Equipment_Package, pk=package_id)

    if request.method == "POST":
        salvaged_keyboard_id = request.POST.get("salvaged_keyboard_id")

        try:
            # CASE 1: Salvaged Keyboard
            if salvaged_keyboard_id:
                salvaged_keyboard = get_object_or_404(SalvagedKeyboard, id=salvaged_keyboard_id)
                
                if salvaged_keyboard.is_reassigned:
                    msg = "❌ This salvaged keyboard has already been reassigned."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                sn_norm = salvaged_keyboard.keyboard_sn.strip().upper()
                existing_keyboard = KeyboardDetails.objects.filter(keyboard_sn_norm=sn_norm).first()
                
                if existing_keyboard:
                    if existing_keyboard.is_disposed:
                        # ✅ REACTIVATE disposed keyboard
                        existing_keyboard.equipment_package = equipment_package
                        existing_keyboard.is_disposed = False
                        existing_keyboard.save()
                        msg = "✅ Salvaged keyboard reactivated and reassigned successfully."
                    else:
                        msg = f"❌ Keyboard '{sn_norm}' is already active in another package."
                        if request.headers.get("x-requested-with") == "XMLHttpRequest":
                            return JsonResponse({'success': False, 'message': msg})
                        messages.error(request, msg)
                        return redirect("desktop_details_view", package_id=equipment_package.id)
                else:
                    # ✅ Create new record
                    KeyboardDetails.objects.create(
                        equipment_package=equipment_package,
                        keyboard_sn_db=salvaged_keyboard.keyboard_sn,
                        keyboard_brand_db=Brand.objects.filter(name=salvaged_keyboard.keyboard_brand).first(),
                        keyboard_model_db=salvaged_keyboard.keyboard_model,
                        is_disposed=False,
                    )
                    msg = "✅ Salvaged keyboard reassigned successfully."

                # Update salvaged record
                salvaged_keyboard.is_reassigned = True
                salvaged_keyboard.reassigned_to = equipment_package
                salvaged_keyboard.save()

                # Log history
                SalvagedKeyboardHistory.objects.create(
                    salvaged_keyboard=salvaged_keyboard,
                    reassigned_to=equipment_package,
                )

            # CASE 2: Manual Input
            else:
                keyboard_sn = request.POST.get("keyboard_sn", "").strip()
                keyboard_brand_id = request.POST.get("keyboard_brand_db")
                keyboard_model = request.POST.get("keyboard_model")

                if not keyboard_sn or not keyboard_model:
                    msg = "❌ Please fill in all required fields."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                sn_norm = keyboard_sn.upper()
                existing_keyboard = KeyboardDetails.objects.filter(keyboard_sn_norm=sn_norm).first()
                
                if existing_keyboard and not existing_keyboard.is_disposed:
                    msg = f"❌ Keyboard '{keyboard_sn}' is already active."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                brand_instance = Brand.objects.filter(id=keyboard_brand_id).first() if keyboard_brand_id else None

                if existing_keyboard and existing_keyboard.is_disposed:
                    # ✅ Reactivate
                    existing_keyboard.equipment_package = equipment_package
                    existing_keyboard.keyboard_brand_db = brand_instance
                    existing_keyboard.keyboard_model_db = keyboard_model
                    existing_keyboard.is_disposed = False
                    existing_keyboard.save()
                else:
                    # Create new
                    KeyboardDetails.objects.create(
                        equipment_package=equipment_package,
                        keyboard_sn_db=keyboard_sn,
                        keyboard_brand_db=brand_instance,
                        keyboard_model_db=keyboard_model,
                        is_disposed=False,
                    )

                msg = "✅ Keyboard added successfully."

            base_url = reverse('desktop_details_view', kwargs={'package_id': equipment_package.pk})
            redirect_url = f"{base_url}#pills-keyboard"

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': True, 'message': msg, 'redirect_url': redirect_url})

            messages.success(request, msg)
            return redirect(redirect_url)

        except Exception as e:
            msg = f"❌ Error: {str(e)}"
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': False, 'message': msg})
            messages.error(request, msg)
            return redirect("desktop_details_view", package_id=equipment_package.id)

    return redirect("desktop_details_view", package_id=equipment_package.id)



#This function allows adding a new mouse to a specific desktop package, then redirects back to the "Mouse" tab of the desktop details view.
@login_required
def add_mouse_to_package(request, package_id):
    """Add or reassign mouse to equipment package"""
    equipment_package = get_object_or_404(Equipment_Package, pk=package_id)

    if request.method == "POST":
        salvaged_mouse_id = request.POST.get("salvaged_mouse_id")

        try:
            # CASE 1: Salvaged Mouse
            if salvaged_mouse_id:
                salvaged_mouse = get_object_or_404(SalvagedMouse, id=salvaged_mouse_id)
                
                if salvaged_mouse.is_reassigned:
                    msg = "❌ This salvaged mouse has already been reassigned."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                sn_norm = salvaged_mouse.mouse_sn.strip().upper()
                existing_mouse = MouseDetails.objects.filter(mouse_sn_norm=sn_norm).first()
                
                if existing_mouse:
                    if existing_mouse.is_disposed:
                        # ✅ REACTIVATE
                        existing_mouse.equipment_package = equipment_package
                        existing_mouse.is_disposed = False
                        existing_mouse.save()
                        msg = "✅ Salvaged mouse reactivated and reassigned successfully."
                    else:
                        msg = f"❌ Mouse '{sn_norm}' is already active in another package."
                        if request.headers.get("x-requested-with") == "XMLHttpRequest":
                            return JsonResponse({'success': False, 'message': msg})
                        messages.error(request, msg)
                        return redirect("desktop_details_view", package_id=equipment_package.id)
                else:
                    # ✅ Create new
                    MouseDetails.objects.create(
                        equipment_package=equipment_package,
                        mouse_sn_db=salvaged_mouse.mouse_sn,
                        mouse_brand_db=Brand.objects.filter(name=salvaged_mouse.mouse_brand).first(),
                        mouse_model_db=salvaged_mouse.mouse_model,
                        is_disposed=False,
                    )
                    msg = "✅ Salvaged mouse reassigned successfully."

                # Update salvaged record
                salvaged_mouse.is_reassigned = True
                salvaged_mouse.reassigned_to = equipment_package
                salvaged_mouse.save()

                # Log history
                SalvagedMouseHistory.objects.create(
                    salvaged_mouse=salvaged_mouse,
                    reassigned_to=equipment_package,
                )

            # CASE 2: Manual Input
            else:
                mouse_sn = request.POST.get("mouse_sn", "").strip()
                mouse_brand_id = request.POST.get("mouse_brand_db")
                mouse_model = request.POST.get("mouse_model")

                if not mouse_sn or not mouse_model:
                    msg = "❌ Please fill in all required fields."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                sn_norm = mouse_sn.upper()
                existing_mouse = MouseDetails.objects.filter(mouse_sn_norm=sn_norm).first()
                
                if existing_mouse and not existing_mouse.is_disposed:
                    msg = f"❌ Mouse '{mouse_sn}' is already active."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                brand_instance = Brand.objects.filter(id=mouse_brand_id).first() if mouse_brand_id else None

                if existing_mouse and existing_mouse.is_disposed:
                    # ✅ Reactivate
                    existing_mouse.equipment_package = equipment_package
                    existing_mouse.mouse_brand_db = brand_instance
                    existing_mouse.mouse_model_db = mouse_model
                    existing_mouse.is_disposed = False
                    existing_mouse.save()
                else:
                    # Create new
                    MouseDetails.objects.create(
                        equipment_package=equipment_package,
                        mouse_sn_db=mouse_sn,
                        mouse_brand_db=brand_instance,
                        mouse_model_db=mouse_model,
                        is_disposed=False,
                    )

                msg = "✅ Mouse added successfully."

            base_url = reverse('desktop_details_view', kwargs={'package_id': equipment_package.pk})
            redirect_url = f"{base_url}#pills-mouse"

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': True, 'message': msg, 'redirect_url': redirect_url})

            messages.success(request, msg)
            return redirect(redirect_url)

        except Exception as e:
            msg = f"❌ Error: {str(e)}"
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': False, 'message': msg})
            messages.error(request, msg)
            return redirect("desktop_details_view", package_id=equipment_package.id)

    return redirect("desktop_details_view", package_id=equipment_package.id)



# 🧱 ADD UPS
@login_required
def add_ups_to_package(request, package_id):
    """Add or reassign UPS to equipment package"""
    equipment_package = get_object_or_404(Equipment_Package, pk=package_id)

    if request.method == "POST":
        salvaged_ups_id = request.POST.get("salvaged_ups_id")

        try:
            # CASE 1: Salvaged UPS
            if salvaged_ups_id:
                salvaged_ups = get_object_or_404(SalvagedUPS, id=salvaged_ups_id)
                
                if salvaged_ups.is_reassigned:
                    msg = "❌ This salvaged UPS has already been reassigned."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                sn_norm = salvaged_ups.ups_sn.strip().upper()
                existing_ups = UPSDetails.objects.filter(ups_sn_norm=sn_norm).first()
                
                if existing_ups:
                    if existing_ups.is_disposed:
                        # ✅ REACTIVATE
                        existing_ups.equipment_package = equipment_package
                        existing_ups.is_disposed = False
                        existing_ups.save()
                        msg = "✅ Salvaged UPS reactivated and reassigned successfully."
                    else:
                        msg = f"❌ UPS '{sn_norm}' is already active in another package."
                        if request.headers.get("x-requested-with") == "XMLHttpRequest":
                            return JsonResponse({'success': False, 'message': msg})
                        messages.error(request, msg)
                        return redirect("desktop_details_view", package_id=equipment_package.id)
                else:
                    # ✅ Create new
                    UPSDetails.objects.create(
                        equipment_package=equipment_package,
                        ups_sn_db=salvaged_ups.ups_sn,
                        ups_brand_db=Brand.objects.filter(name=salvaged_ups.ups_brand).first(),
                        ups_model_db=salvaged_ups.ups_model,
                        is_disposed=False,
                    )
                    msg = "✅ Salvaged UPS reassigned successfully."

                # Update salvaged record
                salvaged_ups.is_reassigned = True
                salvaged_ups.reassigned_to = equipment_package
                salvaged_ups.save()

                # Log history
                SalvagedUPSHistory.objects.create(
                    salvaged_ups=salvaged_ups,
                    reassigned_to=equipment_package,
                )

            # CASE 2: Manual Input
            else:
                ups_sn = request.POST.get("ups_sn", "").strip()
                ups_brand_id = request.POST.get("ups_brand_db")
                ups_model = request.POST.get("ups_model")

                if not ups_sn or not ups_model:
                    msg = "❌ Please fill in all required fields."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                sn_norm = ups_sn.upper()
                existing_ups = UPSDetails.objects.filter(ups_sn_norm=sn_norm).first()
                
                if existing_ups and not existing_ups.is_disposed:
                    msg = f"❌ UPS '{ups_sn}' is already active."
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse({'success': False, 'message': msg})
                    messages.error(request, msg)
                    return redirect("desktop_details_view", package_id=equipment_package.id)

                brand_instance = Brand.objects.filter(id=ups_brand_id).first() if ups_brand_id else None

                if existing_ups and existing_ups.is_disposed:
                    # ✅ Reactivate
                    existing_ups.equipment_package = equipment_package
                    existing_ups.ups_brand_db = brand_instance
                    existing_ups.ups_model_db = ups_model
                    existing_ups.is_disposed = False
                    existing_ups.save()
                else:
                    # Create new
                    UPSDetails.objects.create(
                        equipment_package=equipment_package,
                        ups_sn_db=ups_sn,
                        ups_brand_db=brand_instance,
                        ups_model_db=ups_model,
                        is_disposed=False,
                    )

                msg = "✅ UPS added successfully."

            base_url = reverse('desktop_details_view', kwargs={'package_id': equipment_package.pk})
            redirect_url = f"{base_url}#pills-ups"

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': True, 'message': msg, 'redirect_url': redirect_url})

            messages.success(request, msg)
            return redirect(redirect_url)

        except Exception as e:
            msg = f"❌ Error: {str(e)}"
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': False, 'message': msg})
            messages.error(request, msg)
            return redirect("desktop_details_view", package_id=equipment_package.id)

    return redirect("desktop_details_view", package_id=equipment_package.id)


#This function lists all disposed mice, assuming you have a DisposedMouse model similar to DisposedKeyboard.
def disposed_mice(request):
    # Get all disposed mice
    disposed_mice = DisposedMouse.objects.all()
    # Render the list of disposed mice to the template
    return render(request, 'disposed_mice.html', {'disposed_mice': disposed_mice})

from django.db.models.functions import Lower, Coalesce

@login_required
def add_equipment_package_with_details(request):

    employees = (
        Employee.objects
        .annotate(
            ln=Coalesce(Lower('employee_lname'), Value('zzzzzz')),
            fn=Coalesce(Lower('employee_fname'), Value(''))
        )
        .order_by('ln', 'fn')
    )

    desktop_brands = Brand.objects.filter(is_desktop=True)
    keyboard_brands = Brand.objects.filter(is_keyboard=True)
    mouse_brands = Brand.objects.filter(is_mouse=True)
    monitor_brands = Brand.objects.filter(is_monitor=True)
    ups_brands = Brand.objects.filter(is_ups=True)

    printer_brands = Brand.objects.filter(is_printer=True)

    if hasattr(Brand, "is_laptop"):
        laptop_brands = Brand.objects.filter(Q(is_laptop=True) | Q(is_desktop=True))
    else:
        laptop_brands = Brand.objects.filter(is_desktop=True)


    context = {
        'desktop_brands': desktop_brands,
        'keyboard_brands': keyboard_brands,
        'mouse_brands': mouse_brands,
        'monitor_brands': monitor_brands,
        'ups_brands': ups_brands,
        'printer_brands': printer_brands,
        'laptop_brands': laptop_brands,
        'employees': employees,
    }

    def normalize_sn(value):
        v = (value or "").strip()
        return v.upper() if v else ""

    def sn_exists(model, field_name, serial_value):
        norm = normalize_sn(serial_value)
        if not norm:
            return False
        return model.objects.annotate(
            sn_norm=Upper(Trim(F(field_name)))
        ).filter(sn_norm=norm).exists()

    def get_brand_or_none(field_name):
        brand_id = request.POST.get(field_name)
        if brand_id and brand_id.isdigit():
            return Brand.objects.filter(id=brand_id).first()
        return None

    def get_employee_or_none(field_name):
        emp_id = request.POST.get(field_name)
        return Employee.objects.filter(id=emp_id).first() if (emp_id and emp_id.isdigit()) else None

    if request.method == 'POST':
        context['post_data'] = request.POST
        equipment_type = request.POST.get("name_input")
        errors = []

        # ---------------- DESKTOP ----------------
        if equipment_type == "Desktop":
            desktop_sn = request.POST.get('desktop_serial_no')
            computer_name = request.POST.get('computer_name_input')
            desktop_processor = (request.POST.get('desktop_processor') or "").strip()
            monitor_sn = request.POST.get('monitor_sn_db')
            keyboard_sn = request.POST.get('keyboard_sn_db')
            mouse_sn = request.POST.get('mouse_sn_db')
            ups_sn = request.POST.get('ups_sn_db')

            # validations (unchanged from your original)
            if not desktop_sn: errors.append("Desktop serial number is required.")
            if not computer_name: errors.append("Computer name is required.")
            if not request.POST.get('desktop_brand_name'): errors.append("Desktop brand is required.")
            if not request.POST.get('desktop_model'): errors.append("Desktop model is required.")
            if not desktop_processor or desktop_processor in ["---", "0", "None", "N/A"]:
                errors.append("Processor is required.")
            if not request.POST.get('desktop_memory'): errors.append("Desktop memory is required.")
            if not request.POST.get('desktop_drive'): errors.append("Desktop drive is required.")

            # Monitor
            if not monitor_sn: errors.append("Monitor serial number is required.")
            if not request.POST.get('monitor_brand'): errors.append("Monitor brand is required.")
            if not request.POST.get('monitor_model'): errors.append("Monitor model is required.")
            if not request.POST.get('monitor_size'): errors.append("Monitor size is required.")

            # Keyboard
            if not keyboard_sn: errors.append("Keyboard serial number is required.")
            if not request.POST.get('keyboard_brand'): errors.append("Keyboard brand is required.")
            if not request.POST.get('keyboard_model'): errors.append("Keyboard model is required.")

            # Mouse
            if not mouse_sn: errors.append("Mouse serial number is required.")
            if not request.POST.get('mouse_brand'): errors.append("Mouse brand is required.")
            if not request.POST.get('mouse_model'): errors.append("Mouse model is required.")

            # UPS
            if not ups_sn: errors.append("UPS serial number is required.")
            if not request.POST.get('ups_brand'): errors.append("UPS brand is required.")
            if not request.POST.get('ups_model'): errors.append("UPS model is required.")
            if not request.POST.get('ups_capacity'): errors.append("UPS capacity is required.")

            # Documents
            if not request.POST.get('par_number_input'): errors.append("PAR Number is required.")
            if not request.POST.get('property_number_input'): errors.append("Property Number is required.")
            if not request.POST.get('acquisition_type_input'): errors.append("Acquisition type is required.")
            if not request.POST.get('value_desktop_input'): errors.append("Value is required.")
            if not request.POST.get('date_received_input'): errors.append("Date received is required.")
            if not request.POST.get('date_inspected_input'): errors.append("Date inspected is required.")
            if not request.POST.get('supplier_name_input'): errors.append("Supplier name is required.")
            if not request.POST.get('status_desktop_input'): errors.append("Status is required.")

            # UserDetails
            if not request.POST.get('enduser_input'): errors.append("End user is required.")
            if not request.POST.get('assetowner_input'): errors.append("Asset owner is required.")

            # Duplicate checks
            if desktop_sn and sn_exists(DesktopDetails, 'serial_no', desktop_sn):
                errors.append(f"Desktop SN '{desktop_sn}' already exists.")
            if monitor_sn and sn_exists(MonitorDetails, 'monitor_sn_db', monitor_sn):
                errors.append(f"Monitor SN '{monitor_sn}' already exists.")
            if keyboard_sn and sn_exists(KeyboardDetails, 'keyboard_sn_db', keyboard_sn):
                errors.append(f"Keyboard SN '{keyboard_sn}' already exists.")
            if mouse_sn and sn_exists(MouseDetails, 'mouse_sn_db', mouse_sn):
                errors.append(f"Mouse SN '{mouse_sn}' already exists.")
            if ups_sn and sn_exists(UPSDetails, 'ups_sn_db', ups_sn):
                errors.append(f"UPS SN '{ups_sn}' already exists.")

            if errors:
                for e in errors: messages.error(request, f"❌ {e}")
                return render(request, 'add_equipment_package_with_details.html', context)

            # save Desktop
            try:
                with transaction.atomic():
                    desktop_brand = get_brand_or_none('desktop_brand_name')
                    monitor_brand = get_brand_or_none('monitor_brand')
                    keyboard_brand = get_brand_or_none('keyboard_brand')
                    mouse_brand = get_brand_or_none('mouse_brand')
                    ups_brand = get_brand_or_none('ups_brand')

                    equipment_package = Equipment_Package.objects.create(is_disposed=False)

                    DesktopDetails.objects.create(
                        equipment_package=equipment_package,
                        serial_no=desktop_sn,
                        computer_name=computer_name,
                        brand_name=desktop_brand,
                        model=request.POST.get('desktop_model'),
                        processor=desktop_processor,
                        memory=request.POST.get('desktop_memory'),
                        drive=request.POST.get('desktop_drive'),
                        desktop_Graphics=request.POST.get('desktop_Graphics'),
                        desktop_Graphics_Size=request.POST.get('desktop_Graphics_Size'),
                        desktop_OS=request.POST.get('desktop_OS'),
                        desktop_Office=request.POST.get('desktop_Office'),
                        desktop_OS_keys=request.POST.get('desktop_OS_keys'),
                        desktop_Office_keys=request.POST.get('desktop_Office_keys'),
                    )

                    MonitorDetails.objects.create(
                        equipment_package=equipment_package,
                        monitor_sn_db=monitor_sn,
                        monitor_brand_db=monitor_brand,
                        monitor_model_db=request.POST.get('monitor_model'),
                        monitor_size_db=request.POST.get('monitor_size')
                    )

                    KeyboardDetails.objects.create(
                        equipment_package=equipment_package,
                        keyboard_sn_db=keyboard_sn,
                        keyboard_brand_db=keyboard_brand,
                        keyboard_model_db=request.POST.get('keyboard_model')
                    )

                    MouseDetails.objects.create(
                        equipment_package=equipment_package,
                        mouse_sn_db=mouse_sn,
                        mouse_brand_db=mouse_brand,
                        mouse_model_db=request.POST.get('mouse_model')
                    )

                    UPSDetails.objects.create(
                        equipment_package=equipment_package,
                        ups_sn_db=ups_sn,
                        ups_brand_db=ups_brand,
                        ups_model_db=request.POST.get('ups_model'),
                        ups_capacity_db=request.POST.get('ups_capacity')
                    )
                    
                    raw_value = request.POST.get('value_desktop_input', '')
                    clean_value = raw_value.replace('₱', '').replace(',', '')

                    DocumentsDetails.objects.create(
                        equipment_package=equipment_package,
                        docs_PAR=request.POST.get('par_number_input'),
                        docs_Propertyno=request.POST.get('property_number_input'),
                        docs_Acquisition_Type=request.POST.get('acquisition_type_input'),
                        docs_Value = clean_value,
                        docs_Datereceived=request.POST.get('date_received_input'),
                        docs_Dateinspected=request.POST.get('date_inspected_input'),
                        docs_Supplier=request.POST.get('supplier_name_input'),
                        docs_Status=request.POST.get('status_desktop_input')
                    )

                    enduser = get_employee_or_none('enduser_input')
                    assetowner = get_employee_or_none('assetowner_input')

                    UserDetails.objects.create(
                        equipment_package=equipment_package,
                        user_Enduser=enduser,
                        user_Assetowner=assetowner
                    )

                    if enduser and enduser.employee_office_section:
                        for schedule in PMSectionSchedule.objects.filter(section=enduser.employee_office_section):
                            PMScheduleAssignment.objects.get_or_create(
                                equipment_package=equipment_package,
                                pm_section_schedule=schedule
                            )

                    messages.success(request, "✅ Desktop Package added successfully.", extra_tags='equipment')
                    return redirect(f'/success_add/{equipment_package.id}/?type=Desktop')

            except IntegrityError as ie:
                messages.error(request, f"❌ Could not save: duplicate detected. Details: {ie}")
                return render(request, 'add_equipment_package_with_details.html', context)
            except Exception as e:
                messages.error(request, f"❌ Exception: {str(e)}")
                return render(request, 'add_equipment_package_with_details.html', context)

        ## ---------------- LAPTOP ----------------
        elif equipment_type == "Laptop":
            laptop_sn = request.POST.get('laptop_sn_db')
            laptop_name = request.POST.get('laptop_computer_name')
            laptop_brand = get_brand_or_none('laptop_brand_name')

            # validations
            if not laptop_sn:
                errors.append("Laptop serial number is required.")
            if not laptop_brand:
                errors.append("Laptop brand is required.")
            if not request.POST.get('laptop_model'):
                errors.append("Laptop model is required.")
            if not laptop_name:
                errors.append("Laptop computer name is required.")

            # Documents
            if not request.POST.get('par_number_input'): errors.append("PAR Number is required.")
            if not request.POST.get('property_number_input'): errors.append("Property Number is required.")
            if not request.POST.get('acquisition_type_input'): errors.append("Acquisition type is required.")
            if not request.POST.get('value_laptop_input'): errors.append("Value is required.")
            if not request.POST.get('date_received_input'): errors.append("Date received is required.")
            if not request.POST.get('date_inspected_input'): errors.append("Date inspected is required.")
            if not request.POST.get('supplier_name_input'): errors.append("Supplier name is required.")
            if not request.POST.get('status_laptop_input'): errors.append("Status is required.")

            # User
            if not request.POST.get('enduser_input'): errors.append("End user is required.")
            if not request.POST.get('assetowner_input'): errors.append("Asset owner is required.")

            if laptop_sn and sn_exists(LaptopDetails, 'laptop_sn_db', laptop_sn):
                errors.append(f"Laptop SN '{laptop_sn}' already exists.")

            if errors:
                for e in errors:
                    messages.error(request, f"❌ {e}")
                return render(request, 'add_equipment_package_with_details.html', context)

            try:
                with transaction.atomic():
                    # ✅ create laptop package
                    laptop_package = LaptopPackage.objects.create(is_disposed=False)

                    # ✅ laptop details
                    LaptopDetails.objects.create(
                        laptop_package=laptop_package,
                        laptop_sn_db=laptop_sn,
                        computer_name=laptop_name,
                        brand_name=laptop_brand,
                        model=request.POST.get('laptop_model'),
                        processor=request.POST.get('laptop_processor'),   # ✅ added
                        memory=request.POST.get('laptop_memory'),         # ✅ added
                        drive=request.POST.get('laptop_drive'),           # ✅ added
                        laptop_OS=request.POST.get('laptop_OS'),
                        laptop_Office=request.POST.get('laptop_Office'),
                        laptop_OS_keys=request.POST.get('laptop_OS_keys'),
                        laptop_Office_keys=request.POST.get('laptop_Office_keys'),
                    )

                    # ✅ documents
                    DocumentsDetails.objects.create(
                        laptop_package=laptop_package,
                        docs_PAR=request.POST.get('par_number_input'),
                        docs_Propertyno=request.POST.get('property_number_input'),
                        docs_Acquisition_Type=request.POST.get('acquisition_type_input'),
                        docs_Value=request.POST.get('value_laptop_input'),
                        docs_Datereceived=request.POST.get('date_received_input'),
                        docs_Dateinspected=request.POST.get('date_inspected_input'),
                        docs_Supplier=request.POST.get('supplier_name_input'),
                        docs_Status=request.POST.get('status_laptop_input')
                    )

                    # ✅ user details
                    enduser = get_employee_or_none('enduser_input')
                    assetowner = get_employee_or_none('assetowner_input')

                    UserDetails.objects.create(
                        laptop_package=laptop_package,
                        user_Enduser=enduser,
                        user_Assetowner=assetowner
                    )

                    # ✅ PM schedule
                    if enduser and enduser.employee_office_section:
                        for schedule in PMSectionSchedule.objects.filter(section=enduser.employee_office_section):
                            PMScheduleAssignment.objects.get_or_create(
                                laptop_package=laptop_package,
                                pm_section_schedule=schedule
                            )

                    messages.success(request, "✅ Laptop Package added successfully.", extra_tags='equipment')
                    return redirect(f'/success_add/{laptop_package.id}/?type=Laptop')

            except IntegrityError as ie:
                messages.error(request, f"❌ Could not save: duplicate detected. Details: {ie}")
                return render(request, 'add_equipment_package_with_details.html', context)
            except Exception as e:
                messages.error(request, f"❌❌ Exception: {str(e)}")
                return render(request, 'add_equipment_package_with_details.html', context)

        elif equipment_type == "Printer":
            printer_sn = request.POST.get("printer_sn_db")
            printer_brand = get_brand_or_none("printer_brand")
            printer_model = request.POST.get("printer_model")
            printer_type = request.POST.get("printer_type")
            printer_resolution = request.POST.get("printer_resolution")
            printer_monthly_duty = request.POST.get("printer_monthly_duty")
            printer_color = request.POST.get("printer_color") == "True"
            printer_duplex = request.POST.get("printer_duplex") == "True"

            # validations
            if not printer_sn:
                errors.append("Printer serial number is required.")
            if not printer_brand:
                errors.append("Printer brand is required.")
            if not printer_model:
                errors.append("Printer model is required.")

            # Documents
            if not request.POST.get('par_number_input'): errors.append("PAR Number is required.")
            if not request.POST.get('property_number_input'): errors.append("Property Number is required.")
            if not request.POST.get('acquisition_type_input'): errors.append("Acquisition type is required.")
            if not request.POST.get('value_printer_input'): errors.append("Value is required.")
            if not request.POST.get('date_received_input'): errors.append("Date received is required.")
            if not request.POST.get('date_inspected_input'): errors.append("Date inspected is required.")
            if not request.POST.get('supplier_name_input'): errors.append("Supplier name is required.")
            if not request.POST.get('status_printer_input'): errors.append("Status is required.")

            # User
            if not request.POST.get('enduser_input'): errors.append("End user is required.")
            if not request.POST.get('assetowner_input'): errors.append("Asset owner is required.")

            # Duplicate check
            if printer_sn and sn_exists(PrinterDetails, 'printer_sn_db', printer_sn):
                errors.append(f"Printer SN '{printer_sn}' already exists.")

            if errors:
                for e in errors:
                    messages.error(request, f"❌ {e}")
                return render(request, 'add_equipment_package_with_details.html', context)

            try:
                with transaction.atomic():
                    # ✅ Create printer package
                    printer_package = PrinterPackage.objects.create(is_disposed=False)

                    # ✅ Create printer details
                    printer = PrinterDetails.objects.create(
                        printer_package=printer_package,
                        printer_sn_db=printer_sn,
                        printer_brand_db=printer_brand,
                        printer_model_db=printer_model,
                        printer_type=printer_type,
                        printer_resolution=printer_resolution,
                        printer_monthly_duty=printer_monthly_duty,
                        printer_color=printer_color,
                        printer_duplex=printer_duplex
                    )

                    # ✅ Create documents
                    DocumentsDetails.objects.create(
                        printer_package=printer_package,
                        docs_PAR=request.POST.get('par_number_input'),
                        docs_Propertyno=request.POST.get('property_number_input'),
                        docs_Acquisition_Type=request.POST.get('acquisition_type_input'),
                        docs_Value=request.POST.get('value_printer_input'),
                        docs_Datereceived=request.POST.get('date_received_input'),
                        docs_Dateinspected=request.POST.get('date_inspected_input'),
                        docs_Supplier=request.POST.get('supplier_name_input'),
                        docs_Status=request.POST.get('status_printer_input')
                    )

                    # ✅ Create user details
                    enduser = get_employee_or_none('enduser_input')
                    assetowner = get_employee_or_none('assetowner_input')

                    UserDetails.objects.create(
                        printer_package=printer_package,
                        user_Enduser=enduser,
                        user_Assetowner=assetowner
                    )

                    # ✅ No PM logic for printers
                    messages.success(request, "✅ Printer added successfully.", extra_tags='equipment')
                    return redirect(reverse("success_page", args=[printer.id]) + "?type=Printer")


            except IntegrityError as ie:
                messages.error(request, f"❌ Could not save printer: duplicate detected. Details: {ie}")
                return render(request, 'add_equipment_package_with_details.html', context)
            except Exception as e:
                messages.error(request, f"❌ Exception while saving printer: {str(e)}")
                return render(request, 'add_equipment_package_with_details.html', context)

    return render(request, 'add_equipment_package_with_details.html', context)






#sa pag add ni sa desktop details, check if the computer name already exists in the database.
def check_computer_name(request):
    computer_name = request.GET.get('computer_name', '').strip()
    exists = DesktopDetails.objects.filter(computer_name=computer_name).exists()
    return JsonResponse({'exists': exists})






############### (RECENT at BASE)


def recent_it_equipment_and_count_base(request):
    recent_desktops = DesktopDetails.objects.select_related('equipment_package').order_by('-created_at')[:10]
    total_count = DesktopDetails.objects.count()
    disposed_count = DesktopDetails.objects.filter(is_disposed=True).count()
    disposal_trend = '100,90,95,110,120,130,125'  # static or computed

    return render(request, 'base.html', {
        'recent_desktops': recent_desktops,
        'total_count': total_count,
        'disposed_count': disposed_count,
        'disposal_trend': disposal_trend,
    })

#employees

def employee_list(request):
    if request.method == 'POST':
        # Get form data
        first_name = request.POST.get('firstName')
        middle_initial = request.POST.get('middleInitial', '')
        last_name = request.POST.get('lastName')
        position = request.POST.get('position', 'Unknown')
        office_section_id = request.POST.get('office_section')  # Now this is the foreign key
        status = request.POST.get('status')

        # Create new employee
        Employee.objects.create(
            employee_fname=first_name,
            employee_mname=middle_initial,
            employee_lname=last_name,
            employee_position=position,
            employee_office_section_id=office_section_id,
            employee_status=status
        )

        messages.success(request, f"✅ {first_name} {last_name} has been added successfully!", extra_tags='employee')
        return redirect('employee_list')

    employees = Employee.objects.all()
    office_sections = OfficeSection.objects.all()
    return render(request, 'employees.html', {
        'employees': employees,
        'office_sections': office_sections
    })


def update_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    if request.method == 'POST':
        employee.employee_fname = request.POST.get('firstName')
        employee.employee_mname = request.POST.get('middleInitial')
        employee.employee_lname = request.POST.get('lastName')
        employee.employee_position = request.POST.get('position')
        employee.employee_office_section_id = request.POST.get('office_section')  # FK update
        employee.employee_status = request.POST.get('status')
        employee.save()

        messages.success(request, f"✅ {employee.employee_fname} {employee.employee_lname} has been updated!", extra_tags='employee')
        return redirect('employee_list')

    office_sections = OfficeSection.objects.all()
    return render(request, 'edit_employee.html', {
        'employee': employee,
        'office_sections': office_sections
    })


def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    if request.method == 'POST':
        employee.delete()
        messages.success(request, f"✅ {employee.employee_fname} {employee.employee_lname} has been deleted!", extra_tags='employee')
        return redirect('employee_list')

    the_messages = get_messages(request)
    
    return render(request, 'delete_employee.html', {
        'employee': employee,
        'the_messages': the_messages
    })


# ==========================================
# EMPLOYEE PROFILE VIEW (Logged-in access)
# ==========================================
@login_required
def employee_profile_view(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    sections = OfficeSection.objects.all().order_by("name")

    # =============================
    # ASSETS I USE (I am End User)
    # =============================
    assets_i_use_desktops = (
        Equipment_Package.objects
        .filter(user_details__user_Enduser=employee, is_disposed=False)
        .exclude(user_details__user_Assetowner=employee)  # 👈 Exclude self-owned
        .prefetch_related('user_details__user_Assetowner', 'desktop_details')
        .distinct()
    )

    assets_i_use_laptops = (
        LaptopPackage.objects
        .filter(user_details__user_Enduser=employee, is_disposed=False)
        .exclude(user_details__user_Assetowner=employee)  # 👈 Exclude self-owned
        .prefetch_related('user_details__user_Assetowner', 'laptop_details')
        .distinct()
    )

    assets_i_use_printers = (
        PrinterDetails.objects
        .select_related("printer_package", "printer_brand_db")
        .filter(
            printer_package__user_details__user_Enduser=employee,
            printer_package__is_disposed=False,
            is_disposed=False
        )
        .exclude(printer_package__user_details__user_Assetowner=employee)  # 👈 Exclude self-owned
        .distinct()
    )

    # =============================
    # ASSETS I OWN (I am Asset Owner)
    # =============================
    assets_i_own_desktops = (
        Equipment_Package.objects
        .filter(user_details__user_Assetowner=employee, is_disposed=False)
        .prefetch_related('user_details__user_Enduser', 'desktop_details')
        .distinct()
    )

    assets_i_own_laptops = (
        LaptopPackage.objects
        .filter(user_details__user_Assetowner=employee, is_disposed=False)
        .prefetch_related('user_details__user_Enduser', 'laptop_details')
        .distinct()
    )

    assets_i_own_printers = (
        PrinterDetails.objects
        .select_related("printer_package", "printer_brand_db")
        .filter(
            printer_package__user_details__user_Assetowner=employee,
            printer_package__is_disposed=False,
            is_disposed=False
        )
        .distinct()
    )

    # (rest of your code unchanged — disposal history, history logs, etc.)

    # ============================================
    # DISPOSAL & HISTORY
    # ============================================
    all_packages = assets_i_use_desktops | assets_i_own_desktops
    disposed_desktops = DisposedDesktopDetail.objects.filter(
        desktop__equipment_package__in=all_packages
    ).select_related('desktop')

    disposed_laptops = DisposedLaptop.objects.filter(
        laptop__laptop_package__in=(assets_i_use_laptops | assets_i_own_laptops)
    ).select_related('laptop')

    desktop_ct = ContentType.objects.get_for_model(Equipment_Package)
    laptop_ct = ContentType.objects.get_for_model(LaptopPackage)

    desktop_ids = list(all_packages.values_list('id', flat=True))
    laptop_ids = list((assets_i_use_laptops | assets_i_own_laptops).values_list('id', flat=True))

    enduser_history = EndUserChangeHistory.objects.filter(
        Q(content_type=desktop_ct, object_id__in=desktop_ids)
        | Q(content_type=laptop_ct, object_id__in=laptop_ids)
    ).select_related('old_enduser', 'new_enduser', 'changed_by')[:20]

    assetowner_history = AssetOwnerChangeHistory.objects.filter(
        Q(content_type=desktop_ct, object_id__in=desktop_ids)
        | Q(content_type=laptop_ct, object_id__in=laptop_ids)
    ).select_related('old_assetowner', 'new_assetowner', 'changed_by')[:20]

    # Public profile link
    public_url = (
        request.build_absolute_uri(reverse('employee_assets_public', args=[employee.qr_token]))
        if employee.qr_token else None
    )

    return render(
        request,
        "employees/employee_profile.html",
        {
            "employee": employee,
            "sections": sections,
            "assets_i_use_desktops": assets_i_use_desktops,
            "assets_i_use_laptops": assets_i_use_laptops,
            "assets_i_use_printers": assets_i_use_printers,
            "assets_i_own_desktops": assets_i_own_desktops,
            "assets_i_own_laptops": assets_i_own_laptops,
            "assets_i_own_printers": assets_i_own_printers,
            "disposed_desktops": disposed_desktops,
            "disposed_laptops": disposed_laptops,
            "enduser_history": enduser_history,
            "assetowner_history": assetowner_history,
            "public_url": public_url,
        },
    )


# ==========================================
# UPDATE EMPLOYEE PROFILE
# ==========================================
@login_required
def update_employee_profile(request, employee_id):
    """Update employee profile information"""
    if request.method != "POST":
        return redirect("employee_profile_view", employee_id=employee_id)
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Update basic fields
    employee.employee_fname = request.POST.get("first_name", "").strip()
    employee.employee_mname = request.POST.get("middle_name", "").strip()
    employee.employee_lname = request.POST.get("last_name", "").strip()
    employee.employee_position = request.POST.get("position", "").strip()
    employee.email = request.POST.get("email", "").strip()
    employee.phone = request.POST.get("phone", "").strip()
    employee.bio = request.POST.get("bio", "").strip()
    employee.employee_level = request.POST.get("level", "").strip()
    employee.employee_status = request.POST.get("status", "Active")
    
    # Update office section
    section_id = request.POST.get("office_section_id")
    if section_id:
        try:
            employee.employee_office_section = OfficeSection.objects.get(id=section_id)
        except OfficeSection.DoesNotExist:
            employee.employee_office_section = None
    
    # Handle avatar upload
    if "avatar" in request.FILES:
        avatar = request.FILES["avatar"]
        employee.avatar = avatar
    
    # Handle date hired
    date_hired = request.POST.get("date_hired")
    if date_hired:
        from datetime import datetime
        try:
            employee.date_hired = datetime.strptime(date_hired, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    employee.save()
    messages.success(request, f"Profile updated for {employee.full_name}.")
    return redirect("employee_profile_view", employee_id=employee_id)


# ==========================================
# REGENERATE EMPLOYEE QR CODE
# ==========================================
@login_required
def regenerate_employee_qr(request, employee_id):
    """Regenerate QR code for employee"""
    if request.method != "POST":
        return redirect("employee_profile_view", employee_id=employee_id)
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Delete old QR code
    if employee.qr_code:
        employee.qr_code.delete(save=False)
    
    # Generate new token and QR
    import uuid
    employee.qr_token = uuid.uuid4()
    employee.qr_code = None
    employee.save()
    
    # Regenerate QR
    from inventory.models import ensure_employee_qr
    ensure_employee_qr(employee)
    
    messages.success(request, f"QR code regenerated for {employee.full_name}.")
    return redirect("employee_profile_view", employee_id=employee_id)


# ==========================================
# PUBLIC EMPLOYEE ASSETS VIEW (QR Access)
# ==========================================
def employee_assets_public(request, token):
    """
    Public page accessed via QR code - no login required
    Shows all IT assets assigned to this employee
    """
    employee = get_object_or_404(Employee, qr_token=token)
    
    # Get assigned packages WHERE EMPLOYEE IS ASSET OWNER
    packages = Equipment_Package.objects.none()
    laptops = LaptopPackage.objects.none()
    printers = PrinterDetails.objects.none()
    
    if employee:
        # 🖥 DESKTOP PACKAGES - Filter by ASSET OWNER
        packages = (
            Equipment_Package.objects
            .filter(user_details__user_Assetowner=employee, is_disposed=False)
            .distinct()
            .prefetch_related("desktop_details", "monitors", "keyboards", "mouse_db", "ups")
        )
        
        # 💻 LAPTOP PACKAGES - Filter by ASSET OWNER
        laptops = (
            LaptopPackage.objects
            .filter(user_details__user_Assetowner=employee, is_disposed=False)
            .distinct()
            .prefetch_related("laptop_details")
        )
        
        # 🖨 PRINTERS - Filter by ASSET OWNER
        printers = (
            PrinterDetails.objects
            .select_related("printer_package", "printer_brand_db")
            .filter(
                printer_package__user_details__user_Assetowner=employee,
                printer_package__is_disposed=False,
                is_disposed=False
            )
            .distinct()
        )
    
    # Get disposal history
    disposed_desktops = DisposedDesktopDetail.objects.filter(
        desktop__equipment_package__in=packages
    ).select_related('desktop')
    disposed_monitors = DisposedMonitor.objects.filter(
        monitor_disposed_db__equipment_package__in=packages
    )
    disposed_keyboards = DisposedKeyboard.objects.filter(
        keyboard_dispose_db__equipment_package__in=packages
    )
    disposed_mice = DisposedMouse.objects.filter(
        mouse_db__equipment_package__in=packages
    )
    disposed_ups = DisposedUPS.objects.filter(
        ups_db__equipment_package__in=packages
    )
    disposed_laptops = DisposedLaptop.objects.filter(
        laptop__laptop_package__in=laptops
    ).select_related('laptop')
    
    # Get change history
    desktop_ct = ContentType.objects.get_for_model(Equipment_Package)
    laptop_ct = ContentType.objects.get_for_model(LaptopPackage)
    
    desktop_ids = list(packages.values_list('id', flat=True))
    laptop_ids = list(laptops.values_list('id', flat=True))
    
    enduser_history = EndUserChangeHistory.objects.filter(
        Q(content_type=desktop_ct, object_id__in=desktop_ids) |
        Q(content_type=laptop_ct, object_id__in=laptop_ids)
    ).select_related('old_enduser', 'new_enduser', 'changed_by').order_by('-changed_at')[:20]
    
    assetowner_history = AssetOwnerChangeHistory.objects.filter(
        Q(content_type=desktop_ct, object_id__in=desktop_ids) |
        Q(content_type=laptop_ct, object_id__in=laptop_ids)
    ).select_related('old_assetowner', 'new_assetowner', 'changed_by').order_by('-changed_at')[:20]
    
    context = {
        "employee_owner": employee,
        "employee": employee,
        "packages": packages,
        "laptops": laptops,
        "printers": printers,
        "disposed_desktops": disposed_desktops,
        "disposed_monitors": disposed_monitors,
        "disposed_keyboards": disposed_keyboards,
        "disposed_mice": disposed_mice,
        "disposed_ups": disposed_ups,
        "disposed_laptops": disposed_laptops,
        "enduser_history": enduser_history,
        "assetowner_history": assetowner_history,
    }
    
    return render(request, "employees/employee_assets_public.html", context)




##update asset owner for desktop


@require_POST
def update_asset_owner(request, desktop_id):
    """AJAX: Update Asset Owner for Desktop Package"""
    try:
        with transaction.atomic():
            new_assetowner_id = request.POST.get('assetowner_input')
            if not new_assetowner_id:
                return JsonResponse({'success': False, 'error': 'Please select an asset owner.'}, status=400)

            new_assetowner = get_object_or_404(Employee, id=new_assetowner_id)
            user_details = get_object_or_404(UserDetails, equipment_package__id=desktop_id)
            old_assetowner = user_details.user_Assetowner

            # ✅ Update asset owner
            user_details.user_Assetowner = new_assetowner
            user_details.save()

            # ✅ LOG HISTORY using GenericForeignKey
            AssetOwnerChangeHistory.objects.create(
                device=user_details.equipment_package,  # ✅ Works with any model!
                old_assetowner=old_assetowner,
                new_assetowner=new_assetowner,
                changed_by=request.user if request.user.is_authenticated else None
            )

            return JsonResponse({
                'success': True,
                'message': 'Asset owner updated successfully!',
                'tab': 'user'
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'error': f"Error updating Asset Owner: {str(e)}"
        }, status=500)

#update end user for desktop    
@require_POST
def update_end_user(request, desktop_id):
    """AJAX: Update End User for Desktop Package"""
    try:
        with transaction.atomic():
            new_enduser_id = request.POST.get('enduser_input')
            if not new_enduser_id:
                return JsonResponse({'success': False, 'error': 'Please select an end user.'}, status=400)

            new_enduser = get_object_or_404(Employee, id=new_enduser_id)
            user_details = get_object_or_404(UserDetails, equipment_package__id=desktop_id)
            old_enduser = user_details.user_Enduser

            # ✅ Update end user
            user_details.user_Enduser = new_enduser
            user_details.save()

            # ✅ LOG HISTORY using GenericForeignKey
            EndUserChangeHistory.objects.create(
                device=user_details.equipment_package,  # ✅ Works with any model!
                old_enduser=old_enduser,
                new_enduser=new_enduser,
                changed_by=request.user if request.user.is_authenticated else None
            )

            # ✅ AUTO-TRANSFER PM SCHEDULE
            pm_message = auto_transfer_pm_schedule(
                user_details.equipment_package, 
                new_enduser
            )

            return JsonResponse({
                'success': True,
                'message': f'End user updated successfully. {pm_message}',
                'tab': 'user'
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'error': f"Error updating End User: {str(e)}"
        }, status=500)


@require_POST
def dispose_desktop(request, desktop_id):
    try:
        desktop_details = get_object_or_404(DesktopDetails, id=desktop_id)
        equipment_package = desktop_details.equipment_package
        user_details = UserDetails.objects.filter(equipment_package=equipment_package).first()

        reason = request.POST.get("reason", "")

        disposed_desktop = DisposedDesktopDetail.objects.create(
            desktop=desktop_details,
            serial_no=desktop_details.serial_no,
            brand_name=str(desktop_details.brand_name) if desktop_details.brand_name else None,
            model=desktop_details.model,
            asset_owner=user_details.user_Assetowner.full_name if user_details and user_details.user_Assetowner else None,
            reason=reason,
        )

        # ---------- Monitors ----------
        monitor_action = request.POST.get("monitor")
        if monitor_action == "dispose":
            for m in MonitorDetails.objects.filter(equipment_package=equipment_package, is_disposed=False):
                DisposedMonitor.objects.create(
                    monitor_disposed_db=m,
                    equipment_package=equipment_package,
                    disposed_under=disposed_desktop,
                    monitor_sn=m.monitor_sn_db,
                    monitor_brand=str(m.monitor_brand_db) if m.monitor_brand_db else None,
                    monitor_model=m.monitor_model_db,
                    monitor_size=m.monitor_size_db,
                    reason=reason,
                )
                m.is_disposed = True
                m.save()
        elif monitor_action == "salvage":
            for m in MonitorDetails.objects.filter(equipment_package=equipment_package, is_disposed=False):
                salvage_monitor_logic(m, notes="Salvaged instead of disposed")
                m.is_disposed = True
                m.save()

        # ---------- Keyboards ----------
        keyboard_action = request.POST.get("keyboard")
        if keyboard_action == "dispose":
            for kb in KeyboardDetails.objects.filter(equipment_package=equipment_package, is_disposed=False):
                DisposedKeyboard.objects.create(
                    keyboard_dispose_db=kb,
                    equipment_package=equipment_package,
                    disposed_under=disposed_desktop,
                )
                kb.is_disposed = True
                kb.save()
        elif keyboard_action == "salvage":
            for kb in KeyboardDetails.objects.filter(equipment_package=equipment_package, is_disposed=False):
                salvage_keyboard_logic(kb, notes="Salvaged instead of disposed")
                kb.is_disposed = True
                kb.save()

        # ---------- Mice ----------
        mouse_action = request.POST.get("mouse")
        if mouse_action == "dispose":
            for mouse in MouseDetails.objects.filter(equipment_package=equipment_package, is_disposed=False):
                DisposedMouse.objects.create(
                    mouse_db=mouse,
                    equipment_package=equipment_package,
                    disposed_under=disposed_desktop,
                )
                mouse.is_disposed = True
                mouse.save()
        elif mouse_action == "salvage":
            for mouse in MouseDetails.objects.filter(equipment_package=equipment_package, is_disposed=False):
                salvage_mouse_logic(mouse, notes="Salvaged instead of disposed")
                mouse.is_disposed = True
                mouse.save()

        # ---------- UPS ----------
        ups_action = request.POST.get("ups")
        if ups_action == "dispose":
            for ups in UPSDetails.objects.filter(equipment_package=equipment_package, is_disposed=False):
                DisposedUPS.objects.create(
                    ups_db=ups,
                    equipment_package=equipment_package,
                    disposed_under=disposed_desktop,
                )
                ups.is_disposed = True
                ups.save()
        elif ups_action == "salvage":
            for ups in UPSDetails.objects.filter(equipment_package=equipment_package, is_disposed=False):
                salvage_ups_logic(ups, notes="Salvaged instead of disposed")
                ups.is_disposed = True
                ups.save()

        # ---------- Desktop itself ----------
        desktop_details.is_disposed = True
        desktop_details.save()
        equipment_package.is_disposed = True
        equipment_package.disposal_date = timezone.now()
        equipment_package.save()

        # ---------- Delete related PM notifications AND schedules ----------
        # Get all PM assignments related to this equipment package
        pm_assignments = PMScheduleAssignment.objects.filter(
            equipment_package=equipment_package,
            is_completed=False  # Only delete pending/incomplete PM assignments
        )
        
        # Delete notifications linked to these PM assignments
        for assignment in pm_assignments:
            content_type = ContentType.objects.get_for_model(assignment)
            Notification.objects.filter(
                content_type=content_type,
                object_id=assignment.id,
                notification_type__in=['pm_due', 'pm_overdue']
            ).delete()
        
        # Delete the PM assignments themselves (or mark as cancelled)
        # Option 1: Delete them completely
        pm_assignments.delete()
        
        # Option 2: Mark as cancelled (if you want to keep history)
        # pm_assignments.update(is_completed=True, notes="Cancelled - Equipment Disposed")

        redirect_url = reverse('desktop_details_view', kwargs={'package_id': equipment_package.id})

        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({
                'success': True,
                'message': 'Desktop package disposal completed successfully!',
                'redirect_url': redirect_url
            })

        messages.success(request, "Desktop disposal process completed.")
        return redirect(redirect_url)

    except Exception as e:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({'success': False, 'message': f'Error disposing desktop: {str(e)}'})
        messages.error(request, f"Error disposing desktop: {str(e)}")
        return redirect('desktop_list_func')




#salvaged area overview
def salvage_overview(request):
    # --- Monitors ---
    all_monitors = SalvagedMonitor.objects.select_related("reassigned_to").order_by("-salvage_date", "-id")
    seen_monitors = set()
    deduped_monitors = []
    for sm in all_monitors:
        if sm.monitor_sn in seen_monitors:
            continue
        seen_monitors.add(sm.monitor_sn)

        if sm.reassigned_to:
            dd = DesktopDetails.objects.filter(equipment_package=sm.reassigned_to).first()
            sm.reassigned_computer_name = dd.computer_name if dd else "Unknown"
        else:
            sm.reassigned_computer_name = None

        deduped_monitors.append(sm)

    # --- Keyboards ---
    all_keyboards = SalvagedKeyboard.objects.select_related("reassigned_to").order_by("-salvage_date", "-id")
    seen_keyboards = set()
    deduped_keyboards = []
    for kb in all_keyboards:
        if kb.keyboard_sn in seen_keyboards:
            continue
        seen_keyboards.add(kb.keyboard_sn)

        if kb.reassigned_to:
            dd = DesktopDetails.objects.filter(equipment_package=kb.reassigned_to).first()
            kb.reassigned_computer_name = dd.computer_name if dd else "Unknown"
        else:
            kb.reassigned_computer_name = None

        deduped_keyboards.append(kb)

    # --- Mice ---
    all_mice = SalvagedMouse.objects.select_related("reassigned_to").order_by("-salvage_date", "-id")
    seen_mice = set()
    deduped_mice = []
    for m in all_mice:
        if m.mouse_sn in seen_mice:
            continue
        seen_mice.add(m.mouse_sn)

        if m.reassigned_to:
            dd = DesktopDetails.objects.filter(equipment_package=m.reassigned_to).first()
            m.reassigned_computer_name = dd.computer_name if dd else "Unknown"
        else:
            m.reassigned_computer_name = None

        deduped_mice.append(m)

    # --- UPS ---
    all_ups = SalvagedUPS.objects.select_related("reassigned_to").order_by("-salvage_date", "-id")
    seen_ups = set()
    deduped_ups = []
    for u in all_ups:
        if u.ups_sn in seen_ups:
            continue
        seen_ups.add(u.ups_sn)

        if u.reassigned_to:
            dd = DesktopDetails.objects.filter(equipment_package=u.reassigned_to).first()
            u.reassigned_computer_name = dd.computer_name if dd else "Unknown"
        else:
            u.reassigned_computer_name = None

        deduped_ups.append(u)

    # --- Context ---
    context = {
        "salvaged_monitors": deduped_monitors,
        "salvaged_keyboards": deduped_keyboards,
        "salvaged_mice": deduped_mice,
        "salvaged_ups": deduped_ups,
    }
    return render(request, "salvage_overview.html", context)




def salvaged_monitor_detail(request, pk):
    monitor = get_object_or_404(SalvagedMonitor, pk=pk)

    # history of reassignments
    history = (
        SalvagedMonitorHistory.objects
        .filter(salvaged_monitor=monitor)
        .select_related("reassigned_to")
        .order_by("-reassigned_at")
    )
    for entry in history:
        dd = DesktopDetails.objects.filter(equipment_package=entry.reassigned_to).only("computer_name").first()
        entry.reassigned_computer_name = dd.computer_name if dd else "Unknown"

    # ✅ add status label + class
    if monitor.is_disposed:
        status_label = f"Disposed on {monitor.disposed_date.strftime('%Y-%m-%d %H:%M') if monitor.disposed_date else ''}"
        status_class = "danger"
    elif monitor.is_reassigned:
        status_label = f"Reassigned → {monitor.reassigned_to}"
        status_class = "secondary"
    else:
        status_label = "Available"
        status_class = "success"

    return render(request, "salvage/salvaged_monitor_detail.html", {
        "monitor": monitor,
        "history": history,
        "status_label": status_label,
        "status_class": status_class,
    })

def salvaged_keyboard_detail(request, pk):
    keyboard = get_object_or_404(SalvagedKeyboard, pk=pk)

    history = (
        SalvagedKeyboardHistory.objects
        .filter(salvaged_keyboard=keyboard)
        .select_related("reassigned_to")
        .order_by("-reassigned_at")
    )
    for entry in history:
        dd = DesktopDetails.objects.filter(equipment_package=entry.reassigned_to).only("computer_name").first()
        entry.reassigned_computer_name = dd.computer_name if dd else "Unknown"

    if keyboard.is_disposed:
        status_label = f"Disposed on {keyboard.disposed_date.strftime('%Y-%m-%d %H:%M') if keyboard.disposed_date else ''}"
        status_class = "danger"
    elif keyboard.is_reassigned:
        status_label = f"Reassigned → {keyboard.reassigned_to}"
        status_class = "secondary"
    else:
        status_label = "Available"
        status_class = "success"

    return render(request, "salvage/salvaged_keyboard_detail.html", {
        "keyboard": keyboard,
        "history": history,
        "status_label": status_label,
        "status_class": status_class,
    })


def salvaged_mouse_detail(request, pk):
    mouse = get_object_or_404(SalvagedMouse, pk=pk)

    history = (
        SalvagedMouseHistory.objects
        .filter(salvaged_mouse=mouse)
        .select_related("reassigned_to")
        .order_by("-reassigned_at")
    )
    for entry in history:
        dd = DesktopDetails.objects.filter(equipment_package=entry.reassigned_to).only("computer_name").first()
        entry.reassigned_computer_name = dd.computer_name if dd else "Unknown"

    if mouse.is_disposed:
        status_label = f"Disposed on {mouse.disposed_date.strftime('%Y-%m-%d %H:%M') if mouse.disposed_date else ''}"
        status_class = "danger"
    elif mouse.is_reassigned:
        status_label = f"Reassigned → {mouse.reassigned_to}"
        status_class = "secondary"
    else:
        status_label = "Available"
        status_class = "success"

    return render(request, "salvage/salvaged_mouse_detail.html", {
        "mouse": mouse,
        "history": history,
        "status_label": status_label,
        "status_class": status_class,
    })


def salvaged_ups_detail(request, pk):
    ups = get_object_or_404(SalvagedUPS, pk=pk)

    history = (
        SalvagedUPSHistory.objects
        .filter(salvaged_ups=ups)
        .select_related("reassigned_to")
        .order_by("-reassigned_at")
    )
    for entry in history:
        dd = DesktopDetails.objects.filter(equipment_package=entry.reassigned_to).only("computer_name").first()
        entry.reassigned_computer_name = dd.computer_name if dd else "Unknown"

    if ups.is_disposed:
        status_label = f"Disposed on {ups.disposed_date.strftime('%Y-%m-%d %H:%M') if ups.disposed_date else ''}"
        status_class = "danger"
    elif ups.is_reassigned:
        status_label = f"Reassigned → {ups.reassigned_to}"
        status_class = "secondary"
    else:
        status_label = "Available"
        status_class = "success"

    return render(request, "salvage/salvaged_ups_detail.html", {
        "ups": ups,
        "history": history,
        "status_label": status_label,
        "status_class": status_class,
    })


def add_brand(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        is_desktop = 'is_desktop' in request.POST
        is_keyboard = 'is_keyboard' in request.POST
        is_mouse = 'is_mouse' in request.POST
        is_monitor = 'is_monitor' in request.POST
        is_ups = 'is_ups' in request.POST
        is_printer = 'is_printer' in request.POST    # ✅ NEW

        if not Brand.objects.filter(name=name).exists():
            Brand.objects.create(
                name=name,
                is_desktop=is_desktop,
                is_keyboard=is_keyboard,
                is_mouse=is_mouse,
                is_monitor=is_monitor,
                is_ups=is_ups,
                is_printer=is_printer   # ✅ NEW
            )
        return redirect('add_brand')

    brands = Brand.objects.all()
    return render(request, 'add_brand.html', {'brands': brands})


def edit_brand(request):
    if request.method == 'POST':
        brand_id = request.POST.get('id')
        brand = Brand.objects.get(pk=brand_id)
        brand.name = request.POST.get('name')
        brand.is_desktop = 'is_desktop' in request.POST
        brand.is_keyboard = 'is_keyboard' in request.POST
        brand.is_mouse = 'is_mouse' in request.POST
        brand.is_monitor = 'is_monitor' in request.POST
        brand.is_ups = 'is_ups' in request.POST
        brand.is_printer = 'is_printer' in request.POST    # ✅ NEW
        brand.save()
    return redirect('add_brand')

#print
from django.templatetags.static import static
from django.contrib.contenttypes.models import ContentType

def generate_desktop_pdf(request, package_id):
    # ✅ Use Equipment_Package instead of DesktopDetails
    equipment_package = get_object_or_404(Equipment_Package, id=package_id)
    desktop_details = equipment_package.desktop_details.first()

    if not desktop_details:
        raise Http404("No DesktopDetails found for this package.")

    # Current (non-disposed) components
    keyboard_details = KeyboardDetails.objects.filter(
        equipment_package=equipment_package, is_disposed=False
    ).first()
    mouse_details = MouseDetails.objects.filter(
        equipment_package=equipment_package, is_disposed=False
    ).first()
    monitor_details = MonitorDetails.objects.filter(
        equipment_package=equipment_package, is_disposed=False
    ).first()
    ups_details = UPSDetails.objects.filter(
        equipment_package=equipment_package, is_disposed=False
    ).first()

    # Documents
    documents_details = DocumentsDetails.objects.filter(equipment_package=equipment_package)

    # Current user assignment
    user_details = UserDetails.objects.filter(equipment_package=equipment_package).first()

    # ✅ FIXED: History data using GenericForeignKey pattern
    equipment_package_ct = ContentType.objects.get_for_model(Equipment_Package)
    
    asset_owner_history = AssetOwnerChangeHistory.objects.filter(
        content_type=equipment_package_ct,
        object_id=equipment_package.id
    ).order_by("-changed_at")
    
    enduser_history = EndUserChangeHistory.objects.filter(
        content_type=equipment_package_ct,
        object_id=equipment_package.id
    ).order_by("-changed_at")

    disposed_desktops = DisposedDesktopDetail.objects.filter(
        desktop__equipment_package=equipment_package
    ).order_by("-date_disposed")
    disposed_monitors = DisposedMonitor.objects.filter(
        equipment_package=equipment_package
    ).order_by("-disposal_date")
    disposed_keyboards = DisposedKeyboard.objects.filter(
        equipment_package=equipment_package
    ).order_by("-disposal_date")
    disposed_mice = DisposedMouse.objects.filter(
        equipment_package=equipment_package
    ).order_by("-disposal_date")
    disposed_ups = DisposedUPS.objects.filter(
        equipment_package=equipment_package
    ).order_by("-disposal_date")

    # QR code
    qr_code_url = None
    if equipment_package.qr_code:
        qr_code_url = request.build_absolute_uri(equipment_package.qr_code.url)

    # ✅ Logo fix – build absolute URL
    logo_url = request.build_absolute_uri(static('img/logo.png'))

    # ✅ Smart filename
    filename = f"desktop_{desktop_details.computer_name or equipment_package.id}_details.pdf"

    # Render PDF template
    html_string = render_to_string('pdf_template.html', {
        'desktop_detailsx': desktop_details,
        'equipment_package': equipment_package,
        'keyboard_detailse': keyboard_details,
        'mouse_detailse': mouse_details,
        'monitor_detailse': monitor_details,
        'ups_detailse': ups_details,
        'user_details': user_details,
        'documents_detailse': documents_details,
        'qr_code_url': qr_code_url,
        'logo_url': logo_url,

        # ✅ Added history
        'asset_owner_history': asset_owner_history,
        'enduser_history': enduser_history,
        'disposed_desktops': disposed_desktops,
        'disposed_monitors': disposed_monitors,
        'disposed_keyboards': disposed_keyboards,
        'disposed_mice': disposed_mice,
        'disposed_ups': disposed_ups,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename={filename}'
    HTML(string=html_string).write_pdf(response)

    return response
#export to excel
def export_equipment_packages_excel(request):
    template_path = 'static/excel_template/3f2e3faf-8c25-426f-b673-a2b5fb38e34a.xlsx'
    wb = load_workbook(template_path)
    ws = wb.active

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    start_row = 9
    row = start_row

    desktops = DesktopDetails.objects.select_related('equipment_package', 'brand_name').all()

    for i, desktop in enumerate(desktops, start=1):
        dp = desktop.equipment_package
        doc = DocumentsDetails.objects.filter(equipment_package=dp).first()
        user = UserDetails.objects.filter(equipment_package=dp).first()

        status = "Active" if not desktop.is_disposed else "Disposed"

        ws[f'A{row}'] = f"{i}a"
        ws[f'B{row}'] = "Desktop"
        ws[f'C{row}'] = doc.docs_Acquisition_Type if doc else ''
        ws[f'D{row}'] = desktop.processor
        ws[f'E{row}'] = desktop.memory
        ws[f'F{row}'] = desktop.drive
        ws[f'G{row}'] = desktop.desktop_OS
        ws[f'H{row}'] = desktop.desktop_Office
        ws[f'I{row}'] = status
        ws[f'J{row}'] = doc.docs_PAR if doc else ''
        ws[f'K{row}'] = desktop.serial_no
        ws[f'L{row}'] = doc.docs_Propertyno if doc else ''
        ws[f'M{row}'] = desktop.model
        ws[f'N{row}'] = desktop.brand_name.name if desktop.brand_name else ''
        ws[f'O{row}'] = doc.docs_Value if doc else ''
        ws[f'P{row}'] = f"{user.user_Enduser.employee_fname} {user.user_Enduser.employee_lname}" if user and user.user_Enduser else ''
        ws[f'Q{row}'] = user.user_Enduser.employee_position if user and user.user_Enduser else ''
        ws[f'R{row}'] = user.user_Enduser.employee_office_section.name if user and user.user_Enduser and user.user_Enduser.employee_office_section else ''
        ws[f'S{row}'] = "Region VIII"
        ws[f'T{row}'] = "Leyte 4th DEO"
        ws[f'U{row}'] = f"{user.user_Assetowner.employee_fname} {user.user_Assetowner.employee_lname}" if user and user.user_Assetowner else ''
        ws[f'V{row}'] = doc.docs_Datereceived if doc else ''
        ws[f'W{row}'] = doc.docs_Supplier if doc else ''
        ws[f'X{row}'] = doc.docs_Dateinspected if doc else ''
        ws[f'Y{row}'] = desktop.computer_name
        ws[f'Z{row}'] = status

        if status == "Disposed":
            for col in range(1, 27):
                ws.cell(row=row, column=col).fill = red_fill

        row += 1

        # Components - FIXED MODEL FIELD NAMES
        components = [
            ('Monitor', MonitorDetails.objects.filter(equipment_package=dp), 'b'),
            ('Keyboard', KeyboardDetails.objects.filter(equipment_package=dp), 'c'),
            ('Mouse', MouseDetails.objects.filter(equipment_package=dp), 'd'),
            ('UPS', UPSDetails.objects.filter(equipment_package=dp), 'e'),
        ]

        for label, items, suffix in components:
            for item in items:
                status = "Active" if not item.is_disposed else "Disposed"

                ws[f'A{row}'] = f"{i}{suffix}"
                ws[f'B{row}'] = label
                ws[f'C{row}'] = doc.docs_Acquisition_Type if doc else ''
                ws[f'D{row}'] = ws[f'E{row}'] = ws[f'F{row}'] = ws[f'G{row}'] = ws[f'H{row}'] = "N/A"
                ws[f'I{row}'] = status
                ws[f'J{row}'] = doc.docs_PAR if doc else ''
                ws[f'L{row}'] = doc.docs_Propertyno if doc else ''
                ws[f'O{row}'] = doc.docs_Value if doc else ''
                ws[f'P{row}'] = f"{user.user_Enduser.employee_fname} {user.user_Enduser.employee_lname}" if user and user.user_Enduser else ''
                ws[f'Q{row}'] = user.user_Enduser.employee_position if user and user.user_Enduser else ''
                ws[f'R{row}'] = user.user_Enduser.employee_office_section.name if user and user.user_Enduser and user.user_Enduser.employee_office_section else ''
                ws[f'S{row}'] = "Region VIII"
                ws[f'T{row}'] = "Leyte 4th DEO"
                ws[f'U{row}'] = f"{user.user_Assetowner.employee_fname} {user.user_Assetowner.employee_lname}" if user and user.user_Assetowner else ''
                ws[f'V{row}'] = doc.docs_Datereceived if doc else ''
                ws[f'W{row}'] = doc.docs_Supplier if doc else ''
                ws[f'X{row}'] = doc.docs_Dateinspected if doc else ''
                ws[f'Y{row}'] = "N/A"
                ws[f'Z{row}'] = status

                # Component-specific fields - FIXED FIELD NAMES
                if label == "Monitor":
                    ws[f'K{row}'] = item.monitor_sn_db  # ✅ Fixed from monitor_sn
                    ws[f'M{row}'] = item.monitor_model_db  # ✅ Fixed from monitor_model
                    ws[f'N{row}'] = item.monitor_brand_db.name if item.monitor_brand_db else ''  # ✅ Fixed from monitor_brand
                elif label == "Keyboard":
                    ws[f'K{row}'] = item.keyboard_sn_db  # ✅ Fixed from keyboard_sn
                    ws[f'M{row}'] = item.keyboard_model_db  # ✅ Fixed from keyboard_model
                    ws[f'N{row}'] = item.keyboard_brand_db.name if item.keyboard_brand_db else ''  # ✅ Fixed from keyboard_brand
                elif label == "Mouse":
                    ws[f'K{row}'] = item.mouse_sn_db  # ✅ Correct
                    ws[f'M{row}'] = item.mouse_model_db  # ✅ Correct
                    ws[f'N{row}'] = item.mouse_brand_db.name if item.mouse_brand_db else ''  # ✅ Correct
                elif label == "UPS":
                    ws[f'K{row}'] = item.ups_sn_db  # ✅ Correct
                    ws[f'M{row}'] = item.ups_model_db  # ✅ Correct
                    ws[f'N{row}'] = item.ups_brand_db.name if item.ups_brand_db else ''  # ✅ Correct

                if status == "Disposed":
                    for col in range(1, 27):
                        ws.cell(row=row, column=col).fill = red_fill

                row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=desktop_bundle_export_{datetime.today().date()}.xlsx'
    return response

def add_maintenance(request, desktop_id):
    desktop = get_object_or_404(Equipment_Package, id=desktop_id)

    if request.method == 'POST':
        maintenance_date = request.POST.get('maintenance_date')
        performed_by = request.POST.get('performed_by')
        notes = request.POST.get('notes')

        # Convert maintenance_date to a Python date object
        maintenance_date_obj = datetime.strptime(maintenance_date, '%Y-%m-%d').date()
        next_schedule = maintenance_date_obj + timedelta(days=30)

        PreventiveMaintenance.objects.create(
            equipment_package=desktop,
            maintenance_date=maintenance_date,
            next_schedule=next_schedule,
            performed_by=performed_by,
            notes=notes,
            is_completed=True
        )
        return redirect('maintenance_history', desktop_id=desktop.id)

    return render(request, 'maintenance/add_maintenance.html', {'desktop': desktop})

#maintenance History for desktop
def maintenance_history_view(request, desktop_id):
    # Get the desktop package and related details
    desktop = get_object_or_404(Equipment_Package, pk=desktop_id)
    desktop_details = DesktopDetails.objects.filter(equipment_package=desktop).first()
    user_details = UserDetails.objects.filter(equipment_package=desktop).first()

    # Completed maintenance history
    maintenance_history = (
        PreventiveMaintenance.objects
        .filter(equipment_package=desktop)
        .select_related('pm_schedule_assignment__pm_section_schedule__quarter_schedule')
        .order_by('date_accomplished')
    )

    # Get the latest completed PM (if any)
    latest_pm = maintenance_history.last()

    # Current (pending) PM assignments
    current_pm_schedule = PMScheduleAssignment.objects.filter(
        equipment_package=desktop
    ).select_related(
        'pm_section_schedule__quarter_schedule',
        'pm_section_schedule__section'
    ).order_by(
        'pm_section_schedule__quarter_schedule__year',
        'pm_section_schedule__quarter_schedule__quarter'
    )

        # ✅ Add overdue flag for each schedule
    today = timezone.now().date()
    for schedule in current_pm_schedule:
        end_date = schedule.pm_section_schedule.end_date
        schedule.is_overdue = (not schedule.is_completed) and (end_date < today)

    return render(request, 'maintenance/history.html', {
        'desktop': desktop,
        'desktop_details': desktop_details,
        'user_details': user_details,
        'maintenance_history': maintenance_history,
        'maintenance_records': maintenance_history,
        'pm': latest_pm,
        'current_pm_schedule': current_pm_schedule,  # New table data
    })
def maintenance_history_laptop(request, package_id):
    laptop_package = get_object_or_404(LaptopPackage, pk=package_id)

    # ✅ Fetch all maintenance history for this laptop
    maintenance_history = (
        PreventiveMaintenance.objects
        .filter(laptop_package=laptop_package)
        .select_related('pm_schedule_assignment__pm_section_schedule__quarter_schedule')
        .order_by('date_accomplished')
    )

    latest_pm = maintenance_history.last()

    # ✅ Fetch current PM schedule for this laptop
    current_pm_schedule = PMScheduleAssignment.objects.filter(
        laptop_package=laptop_package
    ).select_related(
        'pm_section_schedule__quarter_schedule',
        'pm_section_schedule__section'
    ).order_by(
        'pm_section_schedule__quarter_schedule__year',
        'pm_section_schedule__quarter_schedule__quarter'
    )

    # ✅ Fetch linked user details for this laptop
    user_details = UserDetails.objects.filter(laptop_package=laptop_package).select_related(
        'user_Enduser__employee_office_section',
        'user_Assetowner__employee_office_section'
    ).first()

    # ✅ Render the page with all needed context
    return render(request, 'maintenance/laptop_history.html', {
        'laptop_package': laptop_package,
        'maintenance_history': maintenance_history,
        'maintenance_records': maintenance_history,
        'pm': latest_pm,
        'current_pm_schedule': current_pm_schedule,
        'user_details': user_details,  # ✅ now added
    })


def get_schedule_date_range(request, quarter_id, section_id):
    try:
        schedule = PMSectionSchedule.objects.get(
            quarter_schedule_id=quarter_id,
            section_id=section_id
        )
        return JsonResponse({
            "start_date": schedule.start_date.strftime("%Y-%m-%d"),
            "end_date": schedule.end_date.strftime("%Y-%m-%d")
        })
    except PMSectionSchedule.DoesNotExist:
        return JsonResponse({"error": "Schedule not found"}, status=404)



# This function handles the checklist for preventive maintenance of a desktop package.


def checklist(request, desktop_id):
    desktop = get_object_or_404(Equipment_Package, pk=desktop_id)

    # quarter_schedules = QuarterSchedule.objects.all().order_by('-year', 'quarter')
    
     # ❌ If desktop is disposed, block PM and show message
    is_disposed = not DesktopDetails.objects.filter(equipment_package=desktop, is_disposed=False).exists()

    if is_disposed:
        messages.error(request, "❌ Desktop was already disposed and cannot be PM anymore.")
        return redirect('desktop_details_view', package_id=desktop.id)  # Make sure this matches your URL name
    
     # ❌ Block if no schedule assigned
    has_schedule = PMScheduleAssignment.objects.filter(equipment_package=desktop).exists()
    if not has_schedule:
        messages.error(request, "⚠ Please add a PM schedule first before conducting Preventive Maintenance.")
        return redirect('desktop_details_view', package_id=desktop.id)


    checklist_labels = {
        1: "Check if configured and connected to the DPWH domain",
        2: "Check if able to access the intranet services",
        3: "Check if installed with anti-virus software authorized by IMS",
        4: "Check if anti-virus definition files are up-to-date",
        5: "Perform full virus scan using updated virus removal tool",
        6: "Remove all un-authorized software installations",
        7: "Remove all un-authorized files (e.g. movies)",
        8: "Check working condition of hardware devices/components",
        9: "Clean hardware and components, and organize cables",
    }

    user_details = UserDetails.objects.filter(equipment_package=desktop).first()
    office = user_details.user_Enduser.employee_office_section if user_details and user_details.user_Enduser else ''
    end_user = f"{user_details.user_Enduser.employee_fname} {user_details.user_Enduser.employee_lname}" if user_details and user_details.user_Enduser else ''
    desktop_details = DesktopDetails.objects.filter(equipment_package=desktop).first()
    
    # ✅ NEW: extract the section_id safely
    section_id = (
        user_details.user_Enduser.employee_office_section.id
        if user_details and user_details.user_Enduser and user_details.user_Enduser.employee_office_section
        else None
    )
    
    # ✅ Only include quarters that actually have a schedule for this section
    quarter_schedules = QuarterSchedule.objects.filter(
        schedules__schedule_assignments__equipment_package=desktop
    
        
    ).exclude(
        schedules__schedule_assignments__maintenances__equipment_package=desktop
    ).distinct().order_by('-year', 'quarter')
        
        # ❌ If no section found, block PM
    if request.method == "POST":
        quarter_id = request.POST.get("quarter_schedule_id")
        date_accomplished = request.POST.get("date_accomplished")
        quarter = QuarterSchedule.objects.get(id=quarter_id) if quarter_id else None

        # ✅ Check if PM was already conducted for this quarter and desktop
        already_done = PreventiveMaintenance.objects.filter(
            equipment_package=desktop,
            pm_schedule_assignment__pm_section_schedule__quarter_schedule=quarter
        ).exists()

        if already_done:
            messages.warning(request, f"❌ for {quarter.get_quarter_display()} {quarter.year} is already conducted.")
            return redirect('checklist', desktop_id=desktop.id)

        matched_schedule = PMScheduleAssignment.objects.filter(
            equipment_package=desktop,
            pm_section_schedule__quarter_schedule=quarter,
            pm_section_schedule__start_date__lte=date_accomplished,
            pm_section_schedule__end_date__gte=date_accomplished
        ).first()

        # Auto-create schedule assignment if not found
        if not matched_schedule and quarter:
            section = user_details.user_Enduser.employee_office_section if user_details and user_details.user_Enduser else None
            pm_section_schedule = PMSectionSchedule.objects.filter(
                quarter_schedule=quarter,
                section=section
            ).first()

            # Auto-create PMSectionSchedule if missing
            if not pm_section_schedule and section:
                pm_section_schedule = PMSectionSchedule.objects.create(
                    quarter_schedule=quarter,
                    section=section,
                    start_date=date_accomplished,
                    end_date=date_accomplished,
                    notes="Auto-created from checklist"
                )

            # Auto-create PMScheduleAssignment
            if pm_section_schedule:
                matched_schedule = PMScheduleAssignment.objects.create(
                    equipment_package=desktop,
                    pm_section_schedule=pm_section_schedule,
                    is_completed=True,
                    remarks="Auto-assigned via checklist"
                )

        # Create PreventiveMaintenance record
        pm = PreventiveMaintenance.objects.create(
            equipment_package=desktop,
            pm_schedule_assignment=matched_schedule,
            office=office,
            end_user=end_user,
            maintenance_date=timezone.now().date(),
            date_accomplished=date_accomplished,
            performed_by=request.user.get_full_name() if request.user.is_authenticated else "Technician",
            is_completed=True,
            **{f"task_{i}": request.POST.get(f"task_{i}") == "on" for i in range(1, 10)},
            **{f"note_{i}": request.POST.get(f"note_{i}", "") for i in range(1, 10)},
        )

        # Save checklist items
        for i in range(1, 10):
            MaintenanceChecklistItem.objects.create(
                maintenance=pm,
                item_text=checklist_labels[i],
                is_checked=request.POST.get(f"task_{i}") == "on"
            )

        if matched_schedule:
            matched_schedule.is_completed = True
            matched_schedule.remarks = "Checklist completed"
            matched_schedule.save()

        return redirect('maintenance_history', desktop_id=desktop.id)

    return render(request, 'maintenance/checklist.html', {
        'desktop': desktop,
        'desktop_details': desktop_details,
        'checklist_labels': checklist_labels,
        'office': office,
        'end_user': end_user,
        'range': range(1, 10),
        'quarter_schedules': quarter_schedules,
        'section_id': section_id,  # ✅ pass section_id to template
    })


def generate_pm_excel_report(request, pm_id):
    pythoncom.CoInitialize()

    # Get preventive maintenance and related desktop details
    pm = get_object_or_404(PreventiveMaintenance, pk=pm_id)
    desktop_details = DesktopDetails.objects.filter(equipment_package=pm.equipment_package).first()

    # Define paths
    template_path = os.path.join(settings.BASE_DIR, 'static', 'excel_template', 'PM checklist.xlsx')
    output_xlsx_path = os.path.join(settings.MEDIA_ROOT, f'PM_Report_{pm.id}.xlsx')
    output_pdf_path = os.path.join(settings.MEDIA_ROOT, f'PM_Report_{pm.id}.pdf')

    # Load Excel template
    wb = load_workbook(template_path)
    ws = wb.active

    # Fill basic info
    ws['C7'] = pm.office or ''
    ws['C8'] = desktop_details.computer_name if desktop_details else ''
    ws['C9'] = pm.end_user or ''
    ws['C10'] = str(pm.date_accomplished or '')

    # Fill checklist: checkmarks and notes
    for i in range(1, 10):
        # Checkmark (✓) in column D
        is_checked = getattr(pm, f"task_{i}", False)
        check_cell = f'D{13 + i}'  # D14 to D22
        ws[check_cell] = "✓" if is_checked else ""

        # Notes in column E
        note_value = getattr(pm, f'note_{i}', '')
        note_cell = f'E{13 + i}'  # E14 to E22
        ws[note_cell] = note_value

    # Save updated Excel file
    wb.save(output_xlsx_path)

    # Convert to PDF using Excel COM
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    wb_com = excel.Workbooks.Open(output_xlsx_path)
    wb_com.ExportAsFixedFormat(0, output_pdf_path)  # 0 = PDF format
    wb_com.Close(False)
    excel.Quit()
    pythoncom.CoUninitialize()

    # Return the PDF file as download
    return FileResponse(open(output_pdf_path, 'rb'), as_attachment=True, filename=f'PM_Report_{pm.id}.pdf')


def pm_overview_view(request):
    pm_assignments = PMScheduleAssignment.objects.select_related(
        'equipment_package',
        'laptop_package',
        'pm_section_schedule__quarter_schedule',
        'pm_section_schedule__section'
    ).all()

    # Attach computer_name_display for assignments
    for assignment in pm_assignments:
        if assignment.equipment_package:
            desktop_detail = DesktopDetails.objects.filter(equipment_package=assignment.equipment_package).first()
            assignment.computer_name_display = desktop_detail.computer_name if desktop_detail else "N/A"
        elif assignment.laptop_package:
            laptop_detail = assignment.laptop_package.laptop_details.first()
            assignment.computer_name_display = laptop_detail.computer_name if laptop_detail else "N/A"
        else:
            assignment.computer_name_display = "N/A"

    # ✅ Build lookup for quarters already assigned (desktop + laptop)
    assigned_quarters_by_device = {}
    for a in pm_assignments:
        qid = a.pm_section_schedule.quarter_schedule_id
        if a.equipment_package:
            key = f"desktop-{a.equipment_package_id}"
            assigned_quarters_by_device.setdefault(key, set()).add(qid)
        if a.laptop_package_id:
            key = f"laptop-{a.laptop_package_id}"
            assigned_quarters_by_device.setdefault(key, set()).add(qid)

    # ✅ Desktops list with user + section
    desktops = list(
        Equipment_Package.objects.prefetch_related(
            Prefetch(
                'user_details',
                queryset=UserDetails.objects.select_related('user_Enduser__employee_office_section')
            )
        )
    )
    for desktop in desktops:
        desktop_detail = DesktopDetails.objects.filter(equipment_package=desktop).first()
        desktop.computer_name_display = desktop_detail.computer_name if desktop_detail else "N/A"
        u = desktop.user_details.first()
        desktop.section_name = u.user_Enduser.employee_office_section.name if u and u.user_Enduser and u.user_Enduser.employee_office_section else None
        desktop.enduser_name = u.user_Enduser.full_name if u and u.user_Enduser else None

    # ✅ Laptops list with user + section
    laptops = list(
        LaptopPackage.objects.prefetch_related(
            Prefetch(
                'user_details',
                queryset=UserDetails.objects.select_related('user_Enduser__employee_office_section')
            ),
            'laptop_details'
        )
    )
    for package in laptops:
        laptop_detail = package.laptop_details.first()
        package.computer_name_display = laptop_detail.computer_name if laptop_detail else "N/A"
        u = package.user_details.first()
        package.section_name = u.user_Enduser.employee_office_section.name if u and u.user_Enduser and u.user_Enduser.employee_office_section else None
        package.enduser_name = u.user_Enduser.full_name if u and u.user_Enduser else None

    # Schedules
    schedules = PMSectionSchedule.objects.select_related('section', 'quarter_schedule')
    schedules_by_section = {}
    for s in schedules:
        section_name = s.section.name if s.section else 'Other'
        schedules_by_section.setdefault(section_name, []).append(s)

    quarters = QuarterSchedule.objects.all().order_by('-year', 'quarter')
    sections = OfficeSection.objects.all()

    return render(request, 'maintenance/overview.html', {
        'pm_assignments': pm_assignments,
        'desktops': desktops,
        'laptops': laptops,
        'schedules': schedules,
        'schedules_by_section': schedules_by_section,
        'quarters': quarters,
        'sections': sections,
        'assigned_quarters_by_device_json': json.dumps(
            {k: sorted(list(v)) for k, v in assigned_quarters_by_device.items()}
        ),
        'today': timezone.now().date(),   # ✅ for "Due" check
    })

def assign_pm_schedule(request):
    if request.method == 'POST':
        device_type = request.POST.get('device_type')
        schedule_id = request.POST.get('schedule_id')

        schedule = get_object_or_404(PMSectionSchedule, pk=schedule_id)

        if device_type == "desktop":
            desktop_id = request.POST.get('equipment_package_id')
            desktop = get_object_or_404(Equipment_Package, pk=desktop_id)

            if PMScheduleAssignment.objects.filter(equipment_package=desktop, pm_section_schedule=schedule).exists():
                messages.warning(request, "This desktop is already assigned to this schedule.")
                return redirect('pm_overview')

            PMScheduleAssignment.objects.create(equipment_package=desktop, pm_section_schedule=schedule)
            messages.success(request, "Desktop successfully assigned to schedule.")

        elif device_type == "laptop":
            laptop_id = request.POST.get('laptop_package_id')   # 👈 must be LaptopPackage, not LaptopDetails
            laptop_package = get_object_or_404(LaptopPackage, pk=laptop_id)

            if PMScheduleAssignment.objects.filter(laptop_package=laptop_package, pm_section_schedule=schedule).exists():
                messages.warning(request, "This laptop is already assigned to this schedule.")
                return redirect('pm_overview')

            PMScheduleAssignment.objects.create(laptop_package=laptop_package, pm_section_schedule=schedule)
            messages.success(request, "Laptop successfully assigned to schedule.")

        else:
            messages.error(request, "Invalid device type selected.")
            return redirect('pm_overview')

        return redirect('pm_overview')

def section_schedule_list_view(request):
    schedules = PMSectionSchedule.objects.select_related('quarter_schedule', 'section').order_by('-start_date')
    quarters = QuarterSchedule.objects.all().order_by('-year')
    sections = OfficeSection.objects.all().order_by('name')

    if request.method == 'POST':
        quarter_id = request.POST.get('quarter_schedule_id')
        section_id = request.POST.get('section_id')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        quarter = get_object_or_404(QuarterSchedule, pk=quarter_id)
        section = get_object_or_404(OfficeSection, pk=section_id)

        # Prevent duplicate schedule for same section and quarter
        if PMSectionSchedule.objects.filter(quarter_schedule=quarter, section=section).exists():
            messages.warning(request, f"A schedule already exists for {section.name} in {quarter.get_quarter_display()} {quarter.year}.")
            return redirect('section_schedule_list')

        # Create schedule if not existing
        PMSectionSchedule.objects.create(
            quarter_schedule=quarter,
            section=section,
            start_date=start_date,
            end_date=end_date,
        )

        messages.success(request, "Schedule added successfully.")
        return redirect('section_schedule_list')

    return render(request, 'maintenance/section_schedule_list.html', {
        'schedules': schedules,
        'quarters': quarters,
        'sections': sections,
    })

# # View to handle office section list and addition    
def office_section_list(request):
    if request.method == 'POST':
        section_name = request.POST.get('section_name')
        if section_name:
            if not OfficeSection.objects.filter(name__iexact=section_name).exists():
                OfficeSection.objects.create(name=section_name)
                messages.success(request, f'Office Section "{section_name}" added successfully.')
            else:
                messages.error(request, f'Office Section "{section_name}" already exists.')
        return redirect('office_section_list')  # replace with your URL name

    office_sections = OfficeSection.objects.all()
    return render(request, 'office_section_list.html', {'office_sections': office_sections})



#disposal overview page

def disposal_overview(request):
    # Desktops
    latest_desktops = (
        DisposedDesktopDetail.objects
        .values("serial_no")
        .annotate(latest_id=Max("id"))
    )
    desktops = DisposedDesktopDetail.objects.filter(id__in=[row["latest_id"] for row in latest_desktops])

    # Monitors
    latest_monitors = (
        DisposedMonitor.objects
        .values("monitor_sn")
        .annotate(latest_id=Max("id"))
    )
    monitors = DisposedMonitor.objects.filter(id__in=[row["latest_id"] for row in latest_monitors])

    # Keyboards
    latest_keyboards = (
        DisposedKeyboard.objects
        .values("keyboard_dispose_db__keyboard_sn_db")
        .annotate(latest_id=Max("id"))
    )
    keyboards = DisposedKeyboard.objects.filter(id__in=[row["latest_id"] for row in latest_keyboards])

    # Mice
    latest_mice = (
        DisposedMouse.objects
        .values("mouse_db__mouse_sn_db")
        .annotate(latest_id=Max("id"))
    )
    mice = DisposedMouse.objects.filter(id__in=[row["latest_id"] for row in latest_mice])

    # UPS
    latest_ups = (
        DisposedUPS.objects
        .values("ups_db__ups_sn_db")
        .annotate(latest_id=Max("id"))
    )
    ups_list = DisposedUPS.objects.filter(id__in=[row["latest_id"] for row in latest_ups])

    categories = [
        {"label": "Disposed Desktops", "items": desktops, "category": "desktop"},
        {"label": "Disposed Monitors", "items": monitors, "category": "monitor"},
        {"label": "Disposed Keyboards", "items": keyboards, "category": "keyboard"},
        {"label": "Disposed Mice", "items": mice, "category": "mouse"},
        {"label": "Disposed UPS", "items": ups_list, "category": "ups"},
    ]

    return render(request, "disposal/disposal_overview.html", {"categories": categories})

def disposal_history(request):
    desktops = DisposedDesktopDetail.objects.all().order_by("-date_disposed")
    monitors = DisposedMonitor.objects.all().order_by("-disposal_date")
    keyboards = DisposedKeyboard.objects.all().order_by("-disposal_date")
    mice = DisposedMouse.objects.all().order_by("-disposal_date")
    ups_list = DisposedUPS.objects.all().order_by("-disposal_date")

    categories = [
        {"label": "Desktop Disposal History", "items": desktops, "category": "desktop"},
        {"label": "Monitor Disposal History", "items": monitors, "category": "monitor"},
        {"label": "Keyboard Disposal History", "items": keyboards, "category": "keyboard"},
        {"label": "Mouse Disposal History", "items": mice, "category": "mouse"},
        {"label": "UPS Disposal History", "items": ups_list, "category": "ups"},
    ]

    return render(request, "disposal/disposal_history.html", {"categories": categories})





#Dashboard chart view for disposal overview
# Map each model to the correct disposal date field


def get_monthly_count(model, label_name, date_field):
    data = (
        model.objects
        .annotate(month=TruncMonth(date_field))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    result = {}
    for entry in data:
        month_label = entry['month'].strftime("%b %Y")
        result[month_label] = entry['count']

    return label_name, result

def get_daily_count(model, label_name, date_field):
    data = (
        model.objects
        .annotate(day=TruncDay(date_field))  # ← Group by day
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )

    result = {}
    for entry in data:
        day_label = entry['day'].strftime("%Y-%m-%d")  # Format: YYYY-MM-DD
        result[day_label] = entry['count']

    return label_name, result



#http://127.0.0.1:8000/dashboard/chart/
def dashboard_view_chart(request):
    all_data = []

    models = [
        (DisposedDesktopDetail, 'Desktop', 'date_disposed'),
        (DisposedMonitor, 'Monitor', 'disposal_date'),
        (DisposedKeyboard, 'Keyboard', 'disposal_date'),
        (DisposedMouse, 'Mouse', 'disposal_date'),
        (DisposedUPS, 'UPS', 'disposal_date'),
    ]

    for model, label, date_field in models:
        category, data = get_daily_count(model, label, date_field)
        all_data.append({
            'label': category,
            'data': list(data.values()),
            'labels': list(data.keys()),
        })

    return render(request, 'dashboard_chart.html', {'chart_data': all_data})


#photo upload for monitor
@require_POST
def upload_monitor_photo(request, monitor_id):
    from django.contrib import messages
    monitor = get_object_or_404(MonitorDetails, id=monitor_id)
    if 'photo' in request.FILES:
        monitor.monitor_photo = request.FILES['photo']
        monitor.save()
        messages.success(request, 'Monitor photo uploaded successfully!')
    return redirect(f'/desktop_details_view/{monitor.equipment_package.id}/#pills-monitor')


@require_POST
def upload_keyboard_photo(request, keyboard_id):
    from django.contrib import messages
    keyboard = get_object_or_404(KeyboardDetails, id=keyboard_id)
    if 'photo' in request.FILES:
        keyboard.keyboard_photo = request.FILES['photo']
        keyboard.save()
        messages.success(request, 'Keyboard photo uploaded successfully!')
    return redirect(f'/desktop_details_view/{keyboard.equipment_package.id}/#pills-keyboard')


@require_POST
def upload_mouse_photo(request, mouse_id):
    from django.contrib import messages
    mouse = get_object_or_404(MouseDetails, id=mouse_id)
    if 'photo' in request.FILES:
        mouse.mouse_photo = request.FILES['photo']
        mouse.save()
        messages.success(request, 'Mouse photo uploaded successfully!')
    return redirect(f'/desktop_details_view/{mouse.equipment_package.id}/#pills-mouse')


@require_POST
def upload_ups_photo(request, ups_id):
    from django.contrib import messages
    ups = get_object_or_404(UPSDetails, id=ups_id)
    if 'photo' in request.FILES:
        ups.ups_photo = request.FILES['photo']
        ups.save()
        messages.success(request, 'UPS photo uploaded successfully!')
    return redirect(f'/desktop_details_view/{ups.equipment_package.id}/#pills-ups')


@require_POST
def upload_document_photo(request, document_id):
    """Upload multiple photos for supporting documents"""
    from inventory.models import DocumentPhoto
    from django.contrib import messages

    document = get_object_or_404(DocumentsDetails, id=document_id)

    # Handle multiple file uploads
    photos = request.FILES.getlist('photos')
    if photos:
        count = 0
        for photo in photos:
            DocumentPhoto.objects.create(
                document=document,
                photo=photo,
                caption=request.POST.get('caption', '')
            )
            count += 1

        if count == 1:
            messages.success(request, 'Document photo uploaded successfully!')
        else:
            messages.success(request, f'{count} document photos uploaded successfully!')

    # Redirect based on which package type this document belongs to
    if document.equipment_package:
        return redirect(f'/desktop_details_view/{document.equipment_package.id}/#pills-documents')
    elif document.laptop_package:
        return redirect(f'/laptops/{document.laptop_package.id}/#pills-documents')
    elif document.printer_package:
        return redirect(f'/printer_details_view/{document.printer_package.id}/#pills-documents')
    elif document.office_supplies_package:
        return redirect(f'/office_supplies_details_view/{document.office_supplies_package.id}/#pills-documents')
    else:
        return redirect('/')


@require_POST
def delete_document_photo(request, photo_id):
    """Delete a document photo"""
    from inventory.models import DocumentPhoto
    from django.contrib import messages

    photo = get_object_or_404(DocumentPhoto, id=photo_id)
    document = photo.document

    # Delete the photo file and database record
    photo.photo.delete()
    photo.delete()

    messages.success(request, 'Document photo deleted successfully!')

    # Redirect based on which package type this document belongs to
    if document.equipment_package:
        return redirect(f'/desktop_details_view/{document.equipment_package.id}/#pills-documents')
    elif document.laptop_package:
        return redirect(f'/laptops/{document.laptop_package.id}/#pills-documents')
    elif document.printer_package:
        return redirect(f'/printer_details_view/{document.printer_package.id}/#pills-documents')
    elif document.office_supplies_package:
        return redirect(f'/office_supplies_details_view/{document.office_supplies_package.id}/#pills-documents')
    else:
        return redirect('/')


@require_POST
def upload_laptop_photo(request, laptop_id):
    """Upload photo for laptop"""
    from inventory.models import LaptopDetails
    from django.http import JsonResponse

    try:
        laptop = get_object_or_404(LaptopDetails, id=laptop_id)

        if 'photo' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'No photo file provided.'
            }, status=400)

        laptop.laptop_photo = request.FILES['photo']
        laptop.save()

        return JsonResponse({
            'success': True,
            'message': 'Laptop photo uploaded successfully!',
            'photo_url': laptop.laptop_photo.url if laptop.laptop_photo else None
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to upload photo: {str(e)}'
        }, status=500)


@require_POST
def delete_laptop_photo(request, laptop_id):
    """Delete photo for laptop"""
    from inventory.models import LaptopDetails
    from django.http import JsonResponse

    try:
        laptop = get_object_or_404(LaptopDetails, id=laptop_id)

        if laptop.laptop_photo:
            laptop.laptop_photo.delete(save=False)
            laptop.laptop_photo = None
            laptop.save()

            return JsonResponse({
                'success': True,
                'message': 'Laptop photo deleted successfully!'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No photo to delete.'
            }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to delete photo: {str(e)}'
        }, status=500)


def export_salvage_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Salvaged Equipment"

    headers = [
        "Category", "SN", "Brand", "Model", "Size/Capacity",
        "Computer Name", "Asset Owner", "Date Salvaged", "Notes", "Status"
    ]
    ws.append(headers)

    def get_asset_owner(equipment_package):
        if not equipment_package:
            return ""
        user = UserDetails.objects.filter(equipment_package=equipment_package).first()
        if user and user.user_Assetowner:
            return user.user_Assetowner.full_name
        return ""

    def fmt_date(dt):
        if not dt:
            return ""
        try:
            return dt.strftime("%Y-%m-%d %H:%M")  # works if datetime
        except Exception:
            return dt.strftime("%Y-%m-%d")        # fallback for date

    # Monitors
    for m in DisposedMonitor.objects.all():
        computer_name = m.equipment_package.computer_name if m.equipment_package else ""
        asset_owner = get_asset_owner(m.equipment_package)
        ws.append([
            "Monitor", m.monitor_sn, m.monitor_brand, m.monitor_model, m.monitor_size,
            computer_name, asset_owner,
            fmt_date(m.disposal_date),
            m.reason or "", "Disposed"
        ])

    # Keyboards
    for k in DisposedKeyboard.objects.all():
        computer_name = k.equipment_package.computer_name if k.equipment_package else ""
        asset_owner = get_asset_owner(k.equipment_package)
        ws.append([
            "Keyboard",
            k.keyboard_dispose_db.keyboard_sn_db,
            str(k.keyboard_dispose_db.keyboard_brand_db),
            k.keyboard_dispose_db.keyboard_model_db,
            "",
            computer_name, asset_owner,
            fmt_date(k.disposal_date),
            "", "Disposed"
        ])

    # Mice
    for mo in DisposedMouse.objects.all():
        computer_name = mo.equipment_package.computer_name if mo.equipment_package else ""
        asset_owner = get_asset_owner(mo.equipment_package)
        ws.append([
            "Mouse",
            mo.mouse_db.mouse_sn_db,
            str(mo.mouse_db.mouse_brand_db),
            mo.mouse_db.mouse_model_db,
            "",
            computer_name, asset_owner,
            fmt_date(mo.disposal_date),
            "", "Disposed"
        ])

    # UPS
    for u in DisposedUPS.objects.all():
        computer_name = u.equipment_package.computer_name if u.equipment_package else ""
        asset_owner = get_asset_owner(u.equipment_package)
        ws.append([
            "UPS",
            u.ups_db.ups_sn_db,
            str(u.ups_db.ups_brand_db),
            u.ups_db.ups_model_db,
            u.ups_db.ups_capacity_db,
            computer_name, asset_owner,
            fmt_date(u.disposal_date),
            "", "Disposed"
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=salvaged_equipment.xlsx'
    return response



def print_salvage_overview(request):
    salvaged_monitors = DisposedMonitor.objects.all()
    salvaged_keyboards = DisposedKeyboard.objects.all()
    salvaged_mice = DisposedMouse.objects.all()
    salvaged_ups = DisposedUPS.objects.all()

    html_string = render_to_string("salvage/print_salvage.html", {
        "salvaged_monitors": salvaged_monitors,
        "salvaged_keyboards": salvaged_keyboards,
        "salvaged_mice": salvaged_mice,
        "salvaged_ups": salvaged_ups,
    })

    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'inline; filename=salvage_overview.pdf'
    HTML(string=html_string).write_pdf(response)
    return response



#the dashboard

# views_dashboard_snippet.py — paste this into your views.py



# Import your models below — adjust names if your app uses different model class names


def _months_back_labels(n=6):
    today = timezone.now().date()
    year = today.year
    month = today.month
    labels = []
    for _ in range(n):
        labels.append(f"{month_abbr[month]} {year}")
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    labels.reverse()
    return labels

def _monthly_counts_qs(qs, date_field, months=6):
    data = (qs.annotate(m=TruncMonth(date_field))
              .values('m')
              .annotate(c=Count('id'))
              .order_by('m'))
    # Map to 'Mon YYYY' -> count
    map_counts = {f"{month_abbr[row['m'].month]} {row['m'].year}": row['c'] for row in data if row['m']}
    labels = _months_back_labels(months)
    return labels, [map_counts.get(lbl, 0) for lbl in labels]


# ALL FOR DASHBOARD
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Q
from collections import defaultdict

@login_required
def dashboard_pro(request):
    """
    Dashboard with KPIs, charts, recent items, and audit trail.
    Updated to work with GenericForeignKey for change history.
    """
    
    # ===================== KPIs =====================
    total_packages = Equipment_Package.objects.count()
    active_packages = Equipment_Package.objects.filter(is_disposed=False).count()
    disposed_all = (
        DisposedDesktopDetail.objects.count() +
        DisposedKeyboard.objects.count() +
        DisposedMouse.objects.count() +
        DisposedMonitor.objects.count() +
        DisposedUPS.objects.count() +
        DisposedLaptop.objects.count() +
        DisposedPrinter.objects.count()
    )
    pm_pending = PMScheduleAssignment.objects.filter(is_completed=False).count()

    # Individual device counts
    total_desktops = DesktopDetails.objects.filter(is_disposed=False).count()
    total_monitors = MonitorDetails.objects.filter(is_disposed=False).count()
    total_keyboards = KeyboardDetails.objects.filter(is_disposed=False).count()
    total_mice = MouseDetails.objects.filter(is_disposed=False).count()
    total_ups = UPSDetails.objects.filter(is_disposed=False).count()
    total_laptops = LaptopDetails.objects.filter(is_disposed=False).count()
    total_printers = PrinterDetails.objects.filter(is_disposed=False).count()

    # Health score (simple calculation)
    health_score = 100
    if total_packages > 0:
        health_score = int((active_packages / total_packages) * 100)

    # ===================== CHARTS DATA =====================
    # Disposal trend - last 6 months
    from datetime import datetime, timedelta
    from django.db.models import Count
    from django.db.models.functions import TruncMonth

    months = 6
    today = timezone.now()
    start = today - timedelta(days=30 * months)

    # ✅ FIXED: Each model uses its correct field name
    # DisposedDesktopDetail uses 'date_disposed'
    desktop_disp = (
        DisposedDesktopDetail.objects.filter(date_disposed__gte=start)
        .annotate(month=TruncMonth("date_disposed"))
        .values("month")
        .annotate(cnt=Count("id"))
        .order_by("month")
    )
    
    # DisposedMouse uses 'disposal_date'
    mouse_disp = (
        DisposedMouse.objects.filter(disposal_date__gte=start)
        .annotate(month=TruncMonth("disposal_date"))
        .values("month")
        .annotate(cnt=Count("id"))
        .order_by("month")
    )
    
    # DisposedKeyboard uses 'disposal_date'
    keyboard_disp = (
        DisposedKeyboard.objects.filter(disposal_date__gte=start)
        .annotate(month=TruncMonth("disposal_date"))
        .values("month")
        .annotate(cnt=Count("id"))
        .order_by("month")
    )
    
    # DisposedUPS uses 'disposal_date'
    ups_disp = (
        DisposedUPS.objects.filter(disposal_date__gte=start)
        .annotate(month=TruncMonth("disposal_date"))
        .values("month")
        .annotate(cnt=Count("id"))
        .order_by("month")
    )

    # Build labels
    lbls = []
    month_keys = []
    for i in range(months, 0, -1):
        dt = today - timedelta(days=30 * i)
        lbls.append(dt.strftime("%b %Y"))
        month_keys.append(dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0))

    # Build series
    def build_series(queryset, month_keys):
        data_dict = {d["month"]: d["cnt"] for d in queryset}
        return [data_dict.get(mk, 0) for mk in month_keys]

    desktop_series = build_series(desktop_disp, month_keys)
    mouse_series = build_series(mouse_disp, month_keys)
    keyboard_series = build_series(keyboard_disp, month_keys)
    ups_series = build_series(ups_disp, month_keys)

    # Disposed by category (pie chart)
    disposed_labels = ["Desktop", "Monitor", "Keyboard", "Mouse", "UPS"]
    disposed_data = [
        DisposedDesktopDetail.objects.count(),
        DisposedMonitor.objects.count(),
        DisposedKeyboard.objects.count(),
        DisposedMouse.objects.count(),
        DisposedUPS.objects.count(),
    ]

    # Active vs Disposed stacked bar
    stack_labels = ["Desktop", "Monitor", "Keyboard", "Mouse", "UPS"]
    active_counts = [
        total_desktops,
        total_monitors,
        total_keyboards,
        total_mice,
        total_ups,
    ]
    disposed_counts = [
        DisposedDesktopDetail.objects.count(),
        DisposedMonitor.objects.count(),
        DisposedKeyboard.objects.count(),
        DisposedMouse.objects.count(),
        DisposedUPS.objects.count(),
    ]

    # Top 5 brands
    brand_counts = {}
    for desktop in DesktopDetails.objects.filter(is_disposed=False).select_related("brand_name"):
        if desktop.brand_name:
            brand_counts[desktop.brand_name.name] = brand_counts.get(desktop.brand_name.name, 0) + 1
    for laptop in LaptopDetails.objects.filter(is_disposed=False).select_related("brand_name"):
        if laptop.brand_name:
            brand_counts[laptop.brand_name.name] = brand_counts.get(laptop.brand_name.name, 0) + 1
    
    top_brands = sorted(brand_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    brand_labels = [b[0] for b in top_brands]
    brand_data = [b[1] for b in top_brands]

    # Assets by section
    section_counts = {}
    for user in UserDetails.objects.select_related("user_Enduser__employee_office_section"):
        if user.user_Enduser and user.user_Enduser.employee_office_section:
            sec_name = str(user.user_Enduser.employee_office_section)
            section_counts[sec_name] = section_counts.get(sec_name, 0) + 1

    section_labels = list(section_counts.keys())
    section_data = list(section_counts.values())

    # ===================== RECENT ITEMS (FIXED) =====================
    recent = []
    
    # Get recent Desktop packages
    recent_desktops = Equipment_Package.objects.filter(
        is_disposed=False
    ).prefetch_related('desktop_details__brand_name').order_by("-created_at")[:10]
    
    for pkg in recent_desktops:
        desktop = pkg.desktop_details.first()
        if desktop:
            recent.append({
                'id': pkg.id,
                'computer_name': desktop.computer_name or 'N/A',
                'serial_no': desktop.serial_no or 'N/A',
                'brand_name': desktop.brand_name.name if desktop.brand_name else 'N/A',
                'model': desktop.model or 'N/A',
                'created_at': pkg.created_at,
                'type': 'Desktop',
                'url_name': 'desktop_details_view'
            })
    
    # Get recent Laptop packages
    recent_laptops = LaptopPackage.objects.filter(
        is_disposed=False
    ).prefetch_related('laptop_details__brand_name').order_by("-created_at")[:10]
    
    for pkg in recent_laptops:
        laptop = pkg.laptop_details.first()
        if laptop:
            recent.append({
                'id': pkg.id,
                'computer_name': laptop.computer_name or 'N/A',
                'serial_no': laptop.laptop_sn_db or 'N/A',
                'brand_name': laptop.brand_name.name if laptop.brand_name else 'N/A',
                'model': laptop.model or 'N/A',
                'created_at': pkg.created_at,
                'type': 'Laptop',
                'url_name': 'laptop_details_view'
            })
    
    # Sort combined list by created_at (most recent first) and take top 10
    recent = sorted(recent, key=lambda x: x['created_at'], reverse=True)[:10]

    # ===================== UPCOMING PM =====================
    pm_upcoming = PMScheduleAssignment.objects.filter(
        is_completed=False
    ).select_related(
        "pm_section_schedule__section",
        "pm_section_schedule__quarter_schedule",
        "equipment_package"
    ).order_by("pm_section_schedule__quarter_schedule__year", "pm_section_schedule__quarter_schedule__quarter")[:5]

    # ===================== AUDIT TRAIL (FIXED FOR GenericForeignKey) =====================
    from django.contrib.contenttypes.models import ContentType
    
    # Get content types for Equipment_Package and LaptopPackage
    equipment_ct = ContentType.objects.get_for_model(Equipment_Package)
    laptop_ct = ContentType.objects.get_for_model(LaptopPackage)

    # ✅ FIXED: Only get records with content_type (skip old broken records)
    enduser = EndUserChangeHistory.objects.filter(
        content_type__isnull=False
    ).select_related(
        "new_enduser", "old_enduser", "changed_by", "content_type"
    ).order_by("-changed_at")[:5]

    assetowner = AssetOwnerChangeHistory.objects.filter(
        content_type__isnull=False
    ).select_related(
        "new_assetowner", "old_assetowner", "changed_by", "content_type"
    ).order_by("-changed_at")[:5]

    audit = []
    
    # Process End User changes
    for e in enduser:
        old_name = e.old_enduser.full_name if e.old_enduser else 'None'
        new_name = e.new_enduser.full_name if e.new_enduser else 'None'
        
        # ✅ Safety check for content_type
        if not e.content_type:
            continue
        
        # Determine device type
        if e.content_type == equipment_ct:
            device_type = "Desktop"
        elif e.content_type == laptop_ct:
            device_type = "Laptop"
        else:
            device_type = e.content_type.model.capitalize()
        
        audit.append({
            "type": "End User",
            "when": e.changed_at.strftime("%Y-%m-%d %H:%M"),
            "text": f"{device_type} Package #{e.object_id}: <strong>{old_name}</strong> → <strong>{new_name}</strong>",
        })
    
    # Process Asset Owner changes
    for a in assetowner:
        old_name = a.old_assetowner.full_name if a.old_assetowner else 'None'
        new_name = a.new_assetowner.full_name if a.new_assetowner else 'None'
        
        # ✅ Safety check for content_type
        if not a.content_type:
            continue
        
        # Determine device type
        if a.content_type == equipment_ct:
            device_type = "Desktop"
        elif a.content_type == laptop_ct:
            device_type = "Laptop"
        else:
            device_type = a.content_type.model.capitalize()
        
        audit.append({
            "type": "Asset Owner",
            "when": a.changed_at.strftime("%Y-%m-%d %H:%M"),
            "text": f"{device_type} Package #{a.object_id}: <strong>{old_name}</strong> → <strong>{new_name}</strong>",
        })
    
    # Sort by time and limit to 10 most recent
    audit = sorted(audit, key=lambda x: x["when"], reverse=True)[:10]

    # ===================== CONTEXT =====================
    context = {
        "kpis": {
            "total_packages": total_packages,
            "active_packages": active_packages,
            "disposed_all": disposed_all,
            "pm_pending": pm_pending,
            "total_desktops": total_desktops,
            "total_monitors": total_monitors,
            "total_keyboards": total_keyboards,
            "total_mice": total_mice,
            "total_ups": total_ups,
            "total_laptops": total_laptops,
            "total_printers": total_printers,
            "health_score": health_score,
        },
        "charts": {
            "months": months,
            "labels": lbls,
            "desktop": desktop_series,
            "mouse": mouse_series,
            "keyboard": keyboard_series,
            "ups": ups_series,
            "disposed_by_cat_labels": disposed_labels,
            "disposed_by_cat_data": disposed_data,
            "stack_labels": stack_labels,
            "stack_active": active_counts,
            "stack_disposed": disposed_counts,
            "brand_labels": brand_labels,
            "brand_data": brand_data,
            "section_labels": section_labels,
            "section_data": section_data,
        },
        "recent": recent,
        "audit": audit,
        "pm_upcoming": pm_upcoming,
    }

    return render(request, "dashboard.html", context)


#end all for dashboard

#QR code for Profile of the ADMIN only
@login_required
def _ensure_user_qr(profile, request):
    """Generate and persist a QR PNG for this profile if missing."""
    if profile.qr_code:
        return
    url = request.build_absolute_uri(reverse('user_assets_public', args=[str(profile.qr_token)]))
    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format='PNG')
    file_name = f"user_qr_{profile.user_id}.png"
    profile.qr_code.save(file_name, ContentFile(buf.getvalue()), save=True)

# views.py
@login_required
def profile_view(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    _ensure_user_qr(profile, request)  # auto-generate on first view

    sections = OfficeSection.objects.all().order_by("name")
    employees = Employee.objects.all().order_by("employee_lname", "employee_fname")

    assigned_packages = []
    assigned_laptops = []
    if profile.employee:
        # ✅ Filter by Asset Owner instead of Enduser
        assigned_packages = (
            Equipment_Package.objects
            .filter(user_details__user_Assetowner=profile.employee)
            .distinct()
            .prefetch_related("desktop_details")
        )

        # ✅ Include laptops where this employee is Asset Owner
        from inventory.models import LaptopPackage  # import if not already
        assigned_laptops = (
            LaptopPackage.objects
            .filter(user_details__user_Assetowner=profile.employee)
            .distinct()
        )

    return render(request, "account/profile.html", {
        "profile": profile,
        "sections": sections,
        "employees": employees,
        "assigned_packages": assigned_packages,
        "assigned_laptops": assigned_laptops,
    })

@login_required
def update_profile(request):
    if request.method != "POST":
        return redirect("profile")

    user = request.user
    profile = user.profile

    # Basic user fields
    user.first_name = request.POST.get("first_name", "").strip()
    user.last_name = request.POST.get("last_name", "").strip()
    # Optional: protect email change behind password check if you want
    email = request.POST.get("email", "").strip()
    if email:
        user.email = email
    user.save()

    # Profile fields
    profile.phone = request.POST.get("phone", "").strip()
    profile.position = request.POST.get("position", "").strip()
    profile.theme = request.POST.get("theme", "light")
    profile.timezone_str = request.POST.get("timezone_str", "Asia/Manila")
    profile.notify_pm_due = True if request.POST.get("notify_pm_due") == "on" else False

    # Optional FK links
    section_id = request.POST.get("office_section_id")
    if section_id:
        try:
            profile.office_section = OfficeSection.objects.get(id=section_id)
        except OfficeSection.DoesNotExist:
            profile.office_section = None

    employee_id = request.POST.get("employee_id")
    if employee_id:
        try:
            profile.employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            profile.employee = None

    # Avatar upload
    if "avatar" in request.FILES:
        avatar = request.FILES["avatar"]
        # Basic validation (size/type) can be added if you like
        profile.avatar = avatar

    profile.save()
    messages.success(request, "Profile updated.")
    return redirect("profile")


@login_required
def change_password(request):
    if request.method != "POST":
        return redirect("profile")

    current = request.POST.get("current_password")
    new1 = request.POST.get("new_password1")
    new2 = request.POST.get("new_password2")

    if not request.user.check_password(current):
        messages.error(request, "Current password is incorrect.")
        return redirect("profile")

    if new1 != new2:
        messages.error(request, "New passwords do not match.")
        return redirect("profile")

    try:
        validate_password(new1, user=request.user)
    except ValidationError as e:
        messages.error(request, " ".join(e.messages))
        return redirect("profile")

    request.user.set_password(new1)
    request.user.save()
    update_session_auth_hash(request, request.user)  # keep user logged in
    messages.success(request, "Password changed successfully.")
    return redirect("profile")

@login_required
def regenerate_user_qr(request):
    """Rotate token (revokes old QR) and re-generate image."""
    profile, _ = Profile.objects.get_or_create(user=request.user)
    import uuid, os
    # Optional: delete old image file
    if profile.qr_code and hasattr(profile.qr_code, 'path') and os.path.exists(profile.qr_code.path):
        os.remove(profile.qr_code.path)
        profile.qr_code = None
    profile.qr_token = uuid.uuid4()
    profile.save()
    _ensure_user_qr(profile, request)
    messages.success(request, "Your QR code has been regenerated.")
    return redirect('profile')


def user_assets_public(request, token):
    """
    Public (or internal) page: by QR token. No login required.
    Shows all IT assets assigned to this user + history.
    Includes Desktops, Laptops, and Printers.
    """
    profile = get_object_or_404(Profile, qr_token=token)
    employee = profile.employee

    packages = Equipment_Package.objects.none()
    laptops = LaptopPackage.objects.none()
    printers = PrinterDetails.objects.none()

    if employee:
        # 🖥 DESKTOP PACKAGES - Filter by ASSET OWNER
        packages = (
            Equipment_Package.objects
            .filter(user_details__user_Assetowner=employee, is_disposed=False)  # ✅ FIXED
            .distinct()
            .prefetch_related('desktop_details')
        )

        # 💻 LAPTOP PACKAGES - Filter by ASSET OWNER
        laptops = (
            LaptopPackage.objects
            .filter(user_details__user_Assetowner=employee, is_disposed=False)  # ✅ FIXED
            .distinct()
            .prefetch_related('laptop_details')
        )

        # 🖨 PRINTERS - Filter by ASSET OWNER
        printers = (
            PrinterDetails.objects
            .select_related("printer_package", "printer_brand_db")
            .filter(
                printer_package__is_disposed=False,
                printer_package__user_details__user_Assetowner=employee,  # ✅ FIXED
                is_disposed=False
            )
            .distinct()
        )

    # 🗑 Disposals (for all desktop-related assets)
    disposed_desktops = DisposedDesktopDetail.objects.filter(desktop__equipment_package__in=packages)
    disposed_monitors = DisposedMonitor.objects.filter(monitor_disposed_db__equipment_package__in=packages)
    disposed_keyboards = DisposedKeyboard.objects.filter(keyboard_dispose_db__equipment_package__in=packages)
    disposed_mice = DisposedMouse.objects.filter(mouse_db__equipment_package__in=packages)
    disposed_ups = DisposedUPS.objects.filter(ups_db__equipment_package__in=packages)
    
    # 🗑 Laptop disposals
    disposed_laptops = DisposedLaptop.objects.filter(laptop__laptop_package__in=laptops)
    
    # 🗑 Printer disposals (if you have this model)
    disposed_printers = DisposedPrinter.objects.filter(
        printer_package__user_details__user_Assetowner=employee
    ) if hasattr(globals(), 'DisposedPrinter') else []

    # 🕒 Change history - using GenericForeignKey filtering
    desktop_ct = ContentType.objects.get_for_model(Equipment_Package)
    laptop_ct = ContentType.objects.get_for_model(LaptopPackage)
    
    # Get package IDs
    desktop_ids = list(packages.values_list('id', flat=True))
    laptop_ids = list(laptops.values_list('id', flat=True))
    
    # Filter EndUserChangeHistory for both desktops and laptops
    enduser_history = EndUserChangeHistory.objects.filter(
        Q(content_type=desktop_ct, object_id__in=desktop_ids) |
        Q(content_type=laptop_ct, object_id__in=laptop_ids)
    ).select_related('old_enduser', 'new_enduser', 'changed_by').order_by('-changed_at')[:20]
    
    # Filter AssetOwnerChangeHistory for both desktops and laptops
    assetowner_history = AssetOwnerChangeHistory.objects.filter(
        Q(content_type=desktop_ct, object_id__in=desktop_ids) |
        Q(content_type=laptop_ct, object_id__in=laptop_ids)
    ).select_related('old_assetowner', 'new_assetowner', 'changed_by').order_by('-changed_at')[:20]

    return render(request, "account/user_assets_public.html", {
        "profile_owner": profile,
        "employee": employee,
        "packages": packages,
        "laptops": laptops,
        "printers": printers,
        "disposed_desktops": disposed_desktops,
        "disposed_monitors": disposed_monitors,
        "disposed_keyboards": disposed_keyboards,
        "disposed_mice": disposed_mice,
        "disposed_ups": disposed_ups,
        "disposed_laptops": disposed_laptops,
        "disposed_printers": disposed_printers,  # ✅ Added
        "enduser_history": enduser_history,
        "assetowner_history": assetowner_history,
    })

# =========================================Laptops ========================================
@login_required
def laptop_list(request):
    # Grab all LaptopPackages
    packages = LaptopPackage.objects.all().select_related()

    laptops = []
    for pkg in packages:
        # Current active details (if any)
        details = LaptopDetails.objects.filter(laptop_package=pkg).order_by('-created_at').first()

        # User assignment
        user = UserDetails.objects.filter(laptop_package=pkg).select_related("user_Enduser").first()

        laptops.append({
            "package": pkg,              # ✅ for item.package.id
            "details": details,          # ✅ for item.details.*
            "end_user": user.user_Enduser.full_name if user and user.user_Enduser else None,
        })

    return render(request, "laptop/laptop_list.html", {"laptops": laptops})


@login_required
def laptop_details_view(request, package_id):
    # Get the laptop package
    laptop_package = get_object_or_404(LaptopPackage, pk=package_id)

    # Related details
    laptop_details = LaptopDetails.objects.filter(laptop_package=laptop_package, is_disposed=False).first()
    user_details = UserDetails.objects.filter(laptop_package=laptop_package).first()
    documents_details = DocumentsDetails.objects.filter(laptop_package=laptop_package).first()
    disposed_laptops = DisposedLaptop.objects.filter(laptop__laptop_package=laptop_package).order_by('-date_disposed')
    brands = Brand.objects.all()
    employees = Employee.objects.all()

    # ✅ ADD CHANGE HISTORY (like desktop)
   
    enduser_history = get_end_user_history(laptop_package)
    assetowner_history = get_asset_owner_history(laptop_package)

    # PM assignments for this laptop
    pm_assignments = PMScheduleAssignment.objects.filter(laptop_package=laptop_package).select_related(
        'pm_section_schedule__quarter_schedule',
        'pm_section_schedule__section'
    )

    # ✅ Preventive Maintenance history
    maintenance_history = (
        PreventiveMaintenance.objects
        .filter(laptop_package=laptop_package)
        .select_related('pm_schedule_assignment__pm_section_schedule__quarter_schedule')
        .order_by('date_accomplished')
    )

    # ✅ Get the section of the end user
    section = (
        user_details.user_Enduser.employee_office_section
        if user_details and user_details.user_Enduser
        else None
    )

    # ✅ Filter schedules only for that section
    schedules = PMSectionSchedule.objects.select_related("quarter_schedule", "section")
    if section:
        schedules = schedules.filter(section=section)

    # ✅ Current (pending) PM assignments
    current_pm_schedule = PMScheduleAssignment.objects.filter(
        laptop_package=laptop_package
    ).select_related(
        'pm_section_schedule__quarter_schedule',
        'pm_section_schedule__section'
    ).order_by(
        'pm_section_schedule__quarter_schedule__year',
        'pm_section_schedule__quarter_schedule__quarter'
    )

    return render(request, 'laptop/laptop_details_view.html', {
        'laptop_package': laptop_package,
        'laptop_details': laptop_details,
        'user_details': user_details,
        'brands': brands,
        'documents_details': documents_details,
        'disposed_laptops': disposed_laptops,
        'pm_assignments': pm_assignments,
        'current_pm_schedule': current_pm_schedule,
        'maintenance_records': maintenance_history,
        'schedules': schedules,
        'section': section,
        'employees': employees,
        'enduser_history': enduser_history,  # ✅ NEW
        'assetowner_history': assetowner_history,  # ✅ NEW
    })
# AJAX handler to edit laptop details
@transaction.atomic
def edit_laptop(request, laptop_id):
    laptop = get_object_or_404(LaptopDetails, pk=laptop_id)

    if request.method == "POST":
        try:
            # ✅ Serial Number
            laptop.laptop_sn_db = request.POST.get("laptop_sn_db", laptop.laptop_sn_db)

            # ✅ Brand
            brand_id = request.POST.get("brand_name")
            if brand_id:
                brand_instance = Brand.objects.get(pk=brand_id)
                laptop.brand_name = brand_instance

            # ✅ Other basic fields
            laptop.model = request.POST.get("model", laptop.model)
            laptop.processor = request.POST.get("processor", laptop.processor)
            laptop.memory = request.POST.get("memory", laptop.memory)
            laptop.drive = request.POST.get("drive", laptop.drive)

            # ✅ Software details
            laptop.laptop_OS = request.POST.get("laptop_OS", laptop.laptop_OS)
            laptop.laptop_Office = request.POST.get("laptop_Office", laptop.laptop_Office)

            # ✅ NEW: Product keys
            laptop.laptop_OS_keys = request.POST.get("laptop_OS_keys", laptop.laptop_OS_keys)
            laptop.laptop_Office_keys = request.POST.get("laptop_Office_keys", laptop.laptop_Office_keys)

            # ✅ Save updates
            laptop.save()

            return JsonResponse({
                "success": True,
                "message": "Laptop details updated successfully, including software keys!"
            })

        except Brand.DoesNotExist:
            return JsonResponse({
                "success": False,
                "error": "Selected brand does not exist."
            })
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            })

    return JsonResponse({
        "success": False,
        "error": "Invalid request method."
    })


@require_POST
def dispose_laptop(request, package_id):
    """
    Dispose a laptop via AJAX POST request.
    Returns JSON response with success/error status.
    """
    try:
        laptop_package = get_object_or_404(LaptopPackage, id=package_id)
        laptop = LaptopDetails.objects.filter(
            laptop_package=laptop_package, 
            is_disposed=False
        ).first()

        if not laptop:
            return JsonResponse({
                'success': False,
                'error': 'No active laptop found for this package.'
            }, status=400)

        user_details = UserDetails.objects.filter(laptop_package=laptop_package).first()
        reason = request.POST.get("reason", "")

        # Create disposal record
        DisposedLaptop.objects.create(
            laptop=laptop,
            serial_no=laptop.laptop_sn_db,
            brand_name=str(laptop.brand_name) if laptop.brand_name else None,
            model=laptop.model,
            asset_owner=user_details.user_Assetowner.full_name if user_details and user_details.user_Assetowner else None,
            reason=reason,
        )

        # Mark laptop as disposed
        laptop.is_disposed = True
        laptop.save()

        # Mark package as disposed
        laptop_package.is_disposed = True
        laptop_package.disposal_date = timezone.now()
        laptop_package.save()

        # ---------- Delete related PM notifications AND schedules ----------
        # Get all PM assignments related to this laptop package
        pm_assignments = PMScheduleAssignment.objects.filter(
            laptop_package=laptop_package,
            is_completed=False  # Only delete pending/incomplete PM assignments
        )
        
        # Delete notifications linked to these PM assignments
        for assignment in pm_assignments:
            content_type = ContentType.objects.get_for_model(assignment)
            Notification.objects.filter(
                content_type=content_type,
                object_id=assignment.id,
                notification_type__in=['pm_due', 'pm_overdue']
            ).delete()
        
        # Delete the PM assignments themselves
        pm_assignments.delete()

        return JsonResponse({
            'success': True,
            'message': 'Laptop disposed successfully!'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=500)



def generate_laptop_pdf(request, package_id):
    from django.templatetags.static import static
    from django.template.loader import render_to_string
    from weasyprint import HTML

    laptop_package = get_object_or_404(LaptopPackage, id=package_id)
    laptop_details = laptop_package.laptop_details.first()

    if not laptop_details:
        raise Http404("No LaptopDetails found for this package.")

    # ✅ Documents
    documents_details = DocumentsDetails.objects.filter(laptop_package=laptop_package).first()

    # ✅ Current user
    user_details = UserDetails.objects.filter(laptop_package=laptop_package).first()

    # ✅ Disposed history
    disposed_laptops = DisposedLaptop.objects.filter(laptop__laptop_package=laptop_package).order_by("-date_disposed")

    # ✅ Maintenance history
    maintenance_records = PreventiveMaintenance.objects.filter(laptop_package=laptop_package).order_by("date_accomplished")

    # ✅ Current PM Schedule
    current_pm_schedule = PMScheduleAssignment.objects.filter(laptop_package=laptop_package).select_related(
        "pm_section_schedule__quarter_schedule", "pm_section_schedule__section"
    ).order_by(
        "pm_section_schedule__quarter_schedule__year",
        "pm_section_schedule__quarter_schedule__quarter"
    )

    # ✅ QR and logo
    qr_code_url = None
    if laptop_package.qr_code:
        qr_code_url = request.build_absolute_uri(laptop_package.qr_code.url)

    logo_url = request.build_absolute_uri(static('img/logo.png'))

    # ✅ Filename
    filename = f"laptop_{laptop_details.computer_name or laptop_package.id}_details.pdf"

    # ✅ Render HTML template
    html_string = render_to_string('laptop/pdf_template_laptop.html', {
        "laptop_package": laptop_package,
        "laptop_details": laptop_details,
        "user_details": user_details,
        "documents_details": documents_details,
        "disposed_laptops": disposed_laptops,
        "maintenance_records": maintenance_records,
        "current_pm_schedule": current_pm_schedule,
        "qr_code_url": qr_code_url,
        "logo_url": logo_url,
    })

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    HTML(string=html_string).write_pdf(response)
    return response


def generate_qr_for_laptop(instance):
    """Generate a QR code for LaptopPackage and attach it to the model."""
    qr = qrcode.make(f"{settings.SITE_URL}{reverse('laptop_details_view', args=[instance.id])}")
    qr_io = BytesIO()
    qr.save(qr_io, format='PNG')
    qr_filename = f"laptop_qr_{instance.id}.png"
    instance.qr_code.save(qr_filename, File(qr_io), save=False)


@login_required
def disposed_laptops(request):
    disposed_laptops = DisposedLaptop.objects.all().order_by("-date_disposed")
    return render(request, "laptop/disposed_laptops.html", {"disposed_laptops": disposed_laptops})


@login_required
def checklist_laptop(request, package_id):
    laptop_package = get_object_or_404(LaptopPackage, pk=package_id)
    laptop = LaptopDetails.objects.filter(laptop_package=laptop_package, is_disposed=False).first()

    quarter_schedules = QuarterSchedule.objects.filter(
        schedules__schedule_assignments__laptop_package=laptop_package
    ).exclude(
        schedules__schedule_assignments__maintenances__laptop_package=laptop_package
    ).distinct().order_by("-year", "quarter")

    if laptop and laptop.is_disposed:
        messages.error(request, "❌ Laptop was already disposed and cannot be PM anymore.")
        return redirect("maintenance_history_laptop", package_id=laptop_package.id)

    has_schedule = PMScheduleAssignment.objects.filter(laptop_package=laptop_package).exists()
    if not has_schedule:
        messages.error(request, "⚠ Please add a PM schedule first before conducting Preventive Maintenance.")
        return redirect("maintenance_history_laptop", package_id=laptop_package.id)

    checklist_labels = {
        1: "Check if configured and connected to the DPWH domain",
        2: "Check if able to access the intranet services",
        3: "Check if installed with anti-virus software authorized by IMS",
        4: "Check if anti-virus definition files are up-to-date",
        5: "Perform full virus scan using updated virus removal tool",
        6: "Remove all un-authorized     software installations",
        7: "Remove all un-authorized files (e.g. movies)",
        8: "Check working condition of hardware devices/components",
        9: "Clean hardware and components, and organize cables",
    }

    user_details = UserDetails.objects.filter(laptop_package=laptop_package).select_related("user_Enduser__employee_office_section").first()
    office = user_details.user_Enduser.employee_office_section if user_details and user_details.user_Enduser else ""
    end_user = f"{user_details.user_Enduser.employee_fname} {user_details.user_Enduser.employee_lname}" if user_details and user_details.user_Enduser else ""

    section_id = office.id if office else None

    if request.method == "POST":
        quarter_id = request.POST.get("quarter_schedule_id")
        date_accomplished = request.POST.get("date_accomplished")
        quarter = QuarterSchedule.objects.get(id=quarter_id) if quarter_id else None

        already_done = PreventiveMaintenance.objects.filter(
            laptop_package=laptop_package,
            pm_schedule_assignment__pm_section_schedule__quarter_schedule=quarter
        ).exists()
        if already_done:
            messages.warning(request, f"❌ for {quarter.get_quarter_display()} {quarter.year} is already conducted.")
            return redirect("checklist_laptop", package_id=laptop_package.id)

        matched_schedule = PMScheduleAssignment.objects.filter(
            laptop_package=laptop_package,
            pm_section_schedule__quarter_schedule=quarter,
            pm_section_schedule__start_date__lte=date_accomplished,
            pm_section_schedule__end_date__gte=date_accomplished
        ).first()

        if not matched_schedule and quarter:
            pm_section_schedule = PMSectionSchedule.objects.filter(
                quarter_schedule=quarter,
                section=office if office else None
            ).first()

            if not pm_section_schedule and office:
                pm_section_schedule = PMSectionSchedule.objects.create(
                    quarter_schedule=quarter,
                    section=office,
                    start_date=date_accomplished,
                    end_date=date_accomplished,
                    notes="Auto-created from laptop checklist"
                )

            if pm_section_schedule:
                matched_schedule = PMScheduleAssignment.objects.create(
                    laptop_package=laptop_package,
                    pm_section_schedule=pm_section_schedule,
                    is_completed=True,
                    remarks="Auto-assigned via laptop checklist"
                )

        pm = PreventiveMaintenance.objects.create(
            laptop_package=laptop_package,
            pm_schedule_assignment=matched_schedule,
            office=office,
            end_user=end_user,
            maintenance_date=timezone.now().date(),
            date_accomplished=date_accomplished,
            performed_by=request.user.get_full_name() if request.user.is_authenticated else "Technician",
            is_completed=True,
            **{f"task_{i}": request.POST.get(f"task_{i}") == "on" for i in range(1, 10)},
            **{f"note_{i}": request.POST.get(f"note_{i}", "") for i in range(1, 10)},
        )

        for i, label in checklist_labels.items():
            MaintenanceChecklistItem.objects.create(
                maintenance=pm,
                item_text=label,
                is_checked=request.POST.get(f"task_{i}") == "on"
            )

        if matched_schedule:
            matched_schedule.is_completed = True
            matched_schedule.remarks = "Checklist completed (Laptop)"
            matched_schedule.save()

        return redirect("maintenance_history_laptop", package_id=laptop_package.id)

    return render(request, "maintenance/checklist_laptop.html", {
        "laptop_package": laptop_package,
        "laptop": laptop,
        "checklist_labels": checklist_labels,
        "office": office,
        "end_user": end_user,
        "range": range(1, 10),
        "quarter_schedules": quarter_schedules,
        "section_id": section_id,
    })



# ==========================================
# STEP 3: Update views.py - Laptop Functions
# ==========================================

@require_POST
def update_asset_owner_laptop(request, package_id):
    """AJAX: Update Asset Owner for Laptop Package"""
    try:
        with transaction.atomic():
            new_assetowner_id = request.POST.get('assetowner_input')
            if not new_assetowner_id:
                return JsonResponse({'success': False, 'error': 'Please select an asset owner.'}, status=400)

            new_assetowner = get_object_or_404(Employee, id=new_assetowner_id)
            user_details = get_object_or_404(UserDetails, laptop_package__id=package_id)
            old_assetowner = user_details.user_Assetowner

            # ✅ Update asset owner
            user_details.user_Assetowner = new_assetowner
            user_details.save()

            # ✅ LOG HISTORY using GenericForeignKey
            AssetOwnerChangeHistory.objects.create(
                device=user_details.laptop_package,  # ✅ Works with any model!
                old_assetowner=old_assetowner,
                new_assetowner=new_assetowner,
                changed_by=request.user if request.user.is_authenticated else None
            )

            return JsonResponse({
                'success': True,
                'message': 'Asset owner updated successfully!',
                'tab': 'user'
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'error': f"Error updating Asset Owner: {str(e)}"
        }, status=500)


@require_POST
def update_end_user_laptop(request, package_id):
    """AJAX: Update End User for Laptop Package"""
    try:
        with transaction.atomic():
            new_enduser_id = request.POST.get('enduser_input')
            if not new_enduser_id:
                return JsonResponse({'success': False, 'error': 'Please select an end user.'}, status=400)

            new_enduser = get_object_or_404(Employee, id=new_enduser_id)
            user_details = get_object_or_404(UserDetails, laptop_package__id=package_id)
            old_enduser = user_details.user_Enduser

            # ✅ Update end user
            user_details.user_Enduser = new_enduser
            user_details.save()

            # ✅ LOG HISTORY using GenericForeignKey
            EndUserChangeHistory.objects.create(
                device=user_details.laptop_package,  # ✅ Works with any model!
                old_enduser=old_enduser,
                new_enduser=new_enduser,
                changed_by=request.user if request.user.is_authenticated else None
            )

            # ✅ AUTO-TRANSFER PM SCHEDULE
            pm_message = auto_transfer_pm_schedule_laptop(
                user_details.laptop_package, 
                new_enduser
            )

            return JsonResponse({
                'success': True,
                'message': f'End user updated successfully. {pm_message}',
                'tab': 'user'
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'error': f"Error updating End User: {str(e)}"
        }, status=500)
    
@require_POST
def update_documents_laptop(request, package_id):
    """AJAX: Update Documents for Laptop Package"""
    try:
        with transaction.atomic():
            documents = get_object_or_404(DocumentsDetails, laptop_package__id=package_id)

            # Update all fields
            fields = [
                "docs_PAR", "docs_Propertyno", "docs_Acquisition_Type", "docs_Value",
                "docs_Datereceived", "docs_Dateinspected", "docs_Supplier", "docs_Status"
            ]
            for f in fields:
                setattr(documents, f, request.POST.get(f))
            documents.save()

            return JsonResponse({
                "success": True,
                "message": "Documents updated successfully!",
                "tab": "documents"
            })

    except DocumentsDetails.DoesNotExist:
        return JsonResponse({
            "success": False,
            "error": "Documents record not found for this laptop.",
            "tab": "documents"
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": f"Error updating documents: {str(e)}",
            "tab": "documents"
        })

# ================================
# PRINTER VIEWS
# ================================

def printer_list(request):
    """List all printers, both active and disposed, with status badges."""
    
    printers = PrinterDetails.objects.select_related(
        'printer_package', 
        'printer_brand_db'
    ).prefetch_related(
        'printer_package__user_details',  # Prefetch UserDetails
        'printer_package__user_details__user_Enduser',  # Prefetch End User Employee
        'printer_package__user_details__user_Assetowner'  # Prefetch Asset Owner Employee
    )
    
    print("🖨️ VIEW CALLED — TOTAL PRINTER COUNT:", printers.count())
    return render(request, 'printer/printer_list.html', {'printers': printers})



def printer_details_view(request, printer_id):
    """Detailed view for a single printer with linked user and documents info."""
    printer = get_object_or_404(PrinterDetails, id=printer_id)

    # ✅ Fetch related details using printer_package
    user_details = UserDetails.objects.filter(printer_package=printer.printer_package).first()
    docs_details = DocumentsDetails.objects.filter(printer_package=printer.printer_package).first()

    # ✅ Fetch disposal history (based on printer_package)
    disposed_printers = DisposedPrinter.objects.filter(
        printer_package=printer.printer_package
    ).order_by('-disposal_date')

    context = {
        'printer': printer,
        'user_details': user_details,
        'docs_details': docs_details,
        'disposed_printers': disposed_printers,
    }
    return render(request, 'printer/printer_details_view.html', context)


def dispose_printer(request, printer_id):
    """Dispose a printer via AJAX confirmation."""
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request method."}, status=405)
    
    printer = get_object_or_404(PrinterDetails, id=printer_id)

    # ✅ Prevent double disposal
    if printer.is_disposed:
        return JsonResponse({"status": "error", "message": "This printer is already disposed."})

    reason = request.POST.get("reason", "No reason provided.")

    # ✅ Create disposal record
    DisposedPrinter.objects.create(
        printer_db=printer,
        printer_package=printer.printer_package,  # updated reference
        printer_sn=printer.printer_sn_db,
        printer_brand=str(printer.printer_brand_db) if printer.printer_brand_db else None,
        printer_model=printer.printer_model_db,
        printer_type=printer.printer_type,
        printer_resolution=printer.printer_resolution,
        printer_monthly_duty=printer.printer_monthly_duty,
        reason=reason,
    )

    # ✅ Mark the printer as disposed
    printer.is_disposed = True
    printer.save()

    # ✅ Mark the package as disposed too
    if printer.printer_package:
        printer.printer_package.is_disposed = True
        printer.printer_package.disposal_date = timezone.now()
        printer.printer_package.save()

    return JsonResponse({
        "status": "success",
        "message": f"Printer {printer.printer_model_db or 'Unknown'} disposed successfully."
    })


def disposed_printers(request):
    """List all disposed printers."""
    disposed = DisposedPrinter.objects.all().select_related('printer_db', 'printer_package')
    return render(request, 'printer/disposed_printers.html', {'disposed_printers': disposed})



# ================================ NOTIFICATIONS VIEWS ================================
@login_required
def notifications_center(request):
    """
    Main notifications center page
    """
    # Get filter parameters
    filter_type = request.GET.get('type', 'all')
    filter_status = request.GET.get('status', 'all')
    search_query = request.GET.get('q', '')
    
    # Base queryset
    notifications = Notification.objects.filter(user=request.user, is_archived=False)
    
    # Apply type filter
    if filter_type != 'all':
        notifications = notifications.filter(notification_type=filter_type)
    
    # Apply status filter
    if filter_status == 'unread':
        notifications = notifications.filter(is_read=False)
    elif filter_status == 'read':
        notifications = notifications.filter(is_read=True)
    
    # Apply search
    if search_query:
        notifications = notifications.filter(
            Q(title__icontains=search_query) | 
            Q(message__icontains=search_query)
        )
    
    # Get statistics
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    total_count = Notification.objects.filter(user=request.user, is_archived=False).count()
    
    # Get counts by type
    type_counts = Notification.objects.filter(
        user=request.user, 
        is_archived=False
    ).values('notification_type').annotate(count=Count('id'))
    
    # Today's notifications
    today = timezone.now().date()
    today_notifications = Notification.objects.filter(
        user=request.user,
        created_at__date=today
    ).count()
    
    # This week's notifications
    week_ago = timezone.now() - timedelta(days=7)
    week_notifications = Notification.objects.filter(
        user=request.user,
        created_at__gte=week_ago
    ).count()
    
    # Priority counts
    urgent_count = notifications.filter(priority='urgent', is_read=False).count()
    high_count = notifications.filter(priority='high', is_read=False).count()
    
    context = {
        'notifications': notifications,
        'unread_count': unread_count,
        'total_count': total_count,
        'today_count': today_notifications,
        'week_count': week_notifications,
        'urgent_count': urgent_count,
        'high_count': high_count,
        'type_counts': type_counts,
        'current_filter': filter_type,
        'current_status': filter_status,
        'search_query': search_query,
    }
    
    return render(request, 'notifications/notifications_center.html', context)


@login_required
def mark_notification_read(request, notification_id):
    """
    Mark a single notification as read
    """
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.mark_as_read()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    return redirect('notifications_center')


@login_required
def mark_all_read(request):
    """
    Mark all notifications as read
    """
    if request.method == 'POST':
        Notification.objects.filter(
            user=request.user, 
            is_read=False
        ).update(is_read=True, read_at=timezone.now())
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'All notifications marked as read'})
    
    return redirect('notifications_center')


@login_required
def delete_notification(request, notification_id):
    """
    Delete a single notification (archive it)
    """
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_archived = True
    notification.save()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    return redirect('notifications_center')


@login_required
def clear_all_notifications(request):
    """
    Clear all read notifications (archive them)
    """
    if request.method == 'POST':
        Notification.objects.filter(
            user=request.user,
            is_read=True
        ).update(is_archived=True)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'All read notifications cleared'})
    
    return redirect('notifications_center')


@login_required
def get_notification_count(request):
    """
    API endpoint to get unread notification count
    For updating navbar badge in real-time
    """
    unread_count = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).count()

    return JsonResponse({'unread_count': unread_count})


# ==================== OFFICE SUPPLIES VIEWS ====================

@login_required
def office_supplies_list(request):
    """List all office supplies"""
    supplies = OfficeSuppliesDetails.objects.select_related(
        'supplies_package', 'brand_name'
    ).filter(is_disposed=False).order_by('-created_at')

    return render(request, 'office_supplies/supplies_list.html', {'supplies': supplies})


@login_required
def office_supplies_details_view(request, package_id):
    """View details of office supplies package"""
    package = get_object_or_404(OfficeSuppliesPackage, id=package_id)
    supplies = package.supplies_details.first()
    user_details = package.user_details.first()
    docs = package.docs.first()

    return render(request, 'office_supplies/supplies_details_view.html', {
        'package': package,
        'supplies': supplies,
        'user_details': user_details,
        'docs': docs
    })


@login_required
def dispose_office_supplies(request, package_id):
    """Mark office supplies as disposed"""
    package = get_object_or_404(OfficeSuppliesPackage, id=package_id)
    supplies = package.supplies_details.first()

    if request.method == 'POST':
        reason = request.POST.get('reason', '')

        with transaction.atomic():
            DisposedOfficeSupplies.objects.create(
                supplies_db=supplies,
                supplies_package=package,
                item_type=supplies.item_type,
                quantity=supplies.quantity,
                unit=supplies.unit,
                reason=reason
            )
            package.is_disposed = True
            package.disposal_date = datetime.now().date()
            package.save()

            supplies.is_disposed = True
            supplies.save()

        messages.success(request, "✅ Office supplies disposed successfully.")
        return redirect('office_supplies_list')

    return render(request, 'office_supplies/dispose_supplies.html', {
        'package': package,
        'supplies': supplies
    })


@login_required
def add_office_supplies(request):
    """Add new office supplies (separate from IT equipment)"""
    from django.db.models.functions import Lower, Coalesce

    employees = (
        Employee.objects
        .annotate(
            ln=Coalesce(Lower('employee_lname'), Value('zzzzzz')),
            fn=Coalesce(Lower('employee_fname'), Value(''))
        )
        .order_by('ln', 'fn')
    )

    office_supplies_brands = Brand.objects.filter(is_office_supplies=True)

    context = {
        'office_supplies_brands': office_supplies_brands,
        'employees': employees,
    }

    def normalize_sn(value):
        v = (value or "").strip()
        return v.upper() if v else ""

    def sn_exists(model, field_name, serial_value):
        norm = normalize_sn(serial_value)
        if not norm:
            return False
        return model.objects.annotate(
            sn_norm=Upper(Trim(F(field_name)))
        ).filter(sn_norm=norm).exists()

    def get_brand_or_none(field_name):
        brand_id = request.POST.get(field_name)
        if brand_id and brand_id.isdigit():
            return Brand.objects.filter(id=brand_id).first()
        return None

    def get_employee_or_none(field_name):
        emp_id = request.POST.get(field_name)
        if emp_id and emp_id.isdigit():
            return Employee.objects.filter(id=emp_id).first()
        return None

    if request.method == 'POST':
        item_type = request.POST.get("item_type")
        supplies_sn = request.POST.get("supplies_sn_db")
        brand = get_brand_or_none("supplies_brand")
        description = request.POST.get("description")
        quantity_str = request.POST.get("quantity", "1")
        unit = request.POST.get("unit", "Unit")

        # validations
        errors = []
        if not item_type:
            errors.append("Item type is required.")

        try:
            quantity = int(quantity_str)
            if quantity < 1:
                errors.append("Quantity must be at least 1.")
        except (ValueError, TypeError):
            errors.append("Valid quantity is required.")
            quantity = 1

        # Documents
        if not request.POST.get('par_number_input'): errors.append("PAR Number is required.")
        if not request.POST.get('property_number_input'): errors.append("Property Number is required.")
        if not request.POST.get('acquisition_type_input'): errors.append("Acquisition type is required.")
        if not request.POST.get('value_supplies_input'): errors.append("Value is required.")
        if not request.POST.get('date_received_input'): errors.append("Date received is required.")
        if not request.POST.get('date_inspected_input'): errors.append("Date inspected is required.")
        if not request.POST.get('supplier_name_input'): errors.append("Supplier name is required.")
        if not request.POST.get('status_supplies_input'): errors.append("Status is required.")

        # User
        if not request.POST.get('enduser_input'): errors.append("End user is required.")
        if not request.POST.get('assetowner_input'): errors.append("Asset owner is required.")

        # Duplicate check (only if serial number provided)
        if supplies_sn and sn_exists(OfficeSuppliesDetails, 'supplies_sn_db', supplies_sn):
            errors.append(f"Office supplies SN '{supplies_sn}' already exists.")

        if errors:
            for e in errors:
                messages.error(request, f"❌ {e}")
            return render(request, 'office_supplies/add_office_supplies.html', context)

        try:
            with transaction.atomic():
                # ✅ Create office supplies package
                supplies_package = OfficeSuppliesPackage.objects.create(is_disposed=False)

                # ✅ Create office supplies details
                supplies = OfficeSuppliesDetails.objects.create(
                    supplies_package=supplies_package,
                    supplies_sn_db=supplies_sn,
                    item_type=item_type,
                    brand_name=brand,
                    description=description,
                    quantity=quantity,
                    unit=unit
                )

                # ✅ Create documents
                DocumentsDetails.objects.create(
                    office_supplies_package=supplies_package,
                    docs_PAR=request.POST.get('par_number_input'),
                    docs_Propertyno=request.POST.get('property_number_input'),
                    docs_Acquisition_Type=request.POST.get('acquisition_type_input'),
                    docs_Value=request.POST.get('value_supplies_input'),
                    docs_Datereceived=request.POST.get('date_received_input'),
                    docs_Dateinspected=request.POST.get('date_inspected_input'),
                    docs_Supplier=request.POST.get('supplier_name_input'),
                    docs_Status=request.POST.get('status_supplies_input')
                )

                # ✅ Create user details
                enduser = get_employee_or_none('enduser_input')
                assetowner = get_employee_or_none('assetowner_input')

                UserDetails.objects.create(
                    office_supplies_package=supplies_package,
                    user_Enduser=enduser,
                    user_Assetowner=assetowner
                )

                messages.success(request, "✅ Office supplies added successfully.")
                return redirect('office_supplies_list')

        except IntegrityError as ie:
            messages.error(request, f"❌ Could not save office supplies: duplicate detected. Details: {ie}")
            return render(request, 'office_supplies/add_office_supplies.html', context)
        except Exception as e:
            messages.error(request, f"❌ Exception while saving office supplies: {str(e)}")
            return render(request, 'office_supplies/add_office_supplies.html', context)

    return render(request, 'office_supplies/add_office_supplies.html', context)